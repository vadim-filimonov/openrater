# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""Brief 92 Phase 92.3 — the build.

Covers: the `Database.transaction()` scope (composition + rollback),
the builder end-to-end against a throwaway DB (with scoring faked at
the `score_once` seam — the same seam every runs test fakes), build
ATOMICITY (a mid-build failure leaves zero rows), the endpoint pair +
duplicate awareness on re-check, and — env-gated — the LIVE golden
path: the canonical nonprofit bundle built for real and its 20 test
cases scored through a running scoring service.
"""

from __future__ import annotations

import json
import os
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from openrater.persistence import Database
from openrater.rates.ingest.parser import parse_workbook
from openrater.rates.ingest.service import build_workbook
from tests.test_ingest_check import CANONICAL, build_mini, to_bytes


@pytest.fixture()
def tmp_db() -> Database:
    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as f:
        path = Path(f.name)
    db = Database(path)
    yield db
    for suffix in ("", "-wal", "-shm"):
        p = Path(str(path) + suffix)
        if p.exists():
            p.unlink()


def _fake_score(outputs: dict[str, float], tier: str | None = None):
    def fake(*, request, base_url=None):  # noqa: ANN001, ANN202
        views = {"premium": next(iter(outputs.values()), None)}
        if tier is not None:
            views["tier"] = tier
        return {"outputs": dict(outputs), "views": views, "row_status": "priced"}

    return fake


# ---------------------------------------------------------------------------
# The transaction scope.
# ---------------------------------------------------------------------------

def test_transaction_scope_composes_inner_begins(tmp_db: Database) -> None:
    conn = tmp_db.connection()
    conn.execute("CREATE TABLE t (v TEXT)")
    conn.close()

    with tmp_db.transaction() as tx:
        c = tx.connection()
        c.execute("INSERT INTO t VALUES ('a')")
        # A callee that manages "its own" transaction, domain-style:
        c.execute("BEGIN IMMEDIATE;")
        c.execute("INSERT INTO t VALUES ('b')")
        c.execute("COMMIT;")
        c.commit()  # a callee's commit() must NOT end the scope.
        c.close()  # nor its close().

    conn = tmp_db.connection()
    assert [r["v"] for r in conn.execute("SELECT v FROM t ORDER BY v")] == ["a", "b"]
    conn.close()


def test_transaction_scope_rolls_back_everything(tmp_db: Database) -> None:
    conn = tmp_db.connection()
    conn.execute("CREATE TABLE t (v TEXT)")
    conn.close()

    with pytest.raises(RuntimeError):
        with tmp_db.transaction() as tx:
            c = tx.connection()
            c.execute("INSERT INTO t VALUES ('a')")
            c.execute("BEGIN;")
            c.execute("INSERT INTO t VALUES ('b')")
            c.execute("COMMIT;")
            raise RuntimeError("mid-scope failure")

    conn = tmp_db.connection()
    assert conn.execute("SELECT COUNT(*) AS n FROM t").fetchone()["n"] == 0
    conn.close()


# ---------------------------------------------------------------------------
# The build, end to end (scoring faked at the seam).
# ---------------------------------------------------------------------------

def test_build_mini_workbook(tmp_db: Database, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    outcome = build_workbook(db=tmp_db, data=to_bytes(build_mini()), filename="mini.xlsx")

    conn = tmp_db.connection()
    plan_id = outcome.rating_plan_id
    kinds = sorted(
        r["stage_kind"]
        for r in conn.execute(
            "SELECT stage_kind FROM rating_plan_stages WHERE rating_plan_id = ?",
            (plan_id,),
        )
    )
    assert kinds == sorted(
        ["input_node", "input_node", "input_node", "multiplicative_chain"]
    )
    chain_cfg = conn.execute(
        "SELECT config_json FROM rating_plan_stages WHERE rating_plan_id = ? "
        "AND stage_kind = 'multiplicative_chain'",
        (plan_id,),
    ).fetchone()
    # The outputs sheet's chain-stage source renamed the tower's field.
    assert '"output_field": "building_premium"' in chain_cfg["config_json"]
    n_dims = conn.execute(
        "SELECT COUNT(*) AS n FROM plan_dimensions WHERE rating_plan_id = ?", (plan_id,)
    ).fetchone()["n"]
    n_tables = conn.execute(
        "SELECT COUNT(*) AS n FROM plan_factor_tables WHERE rating_plan_id = ?", (plan_id,)
    ).fetchone()["n"]
    n_cells = conn.execute(
        "SELECT COUNT(*) AS n FROM plan_factor_table_cells WHERE rating_plan_id = ?",
        (plan_id,),
    ).fetchone()["n"]
    plan_row = conn.execute(
        "SELECT product, jurisdiction, status, description FROM rating_plans "
        "WHERE rating_plan_id = ?",
        (plan_id,),
    ).fetchone()
    conn.close()

    assert (n_dims, n_tables, n_cells) == (2, 2, 6)
    assert plan_row["product"] == "bop"
    assert plan_row["jurisdiction"] == "IL"
    assert "mini-bop-demo-il-2026" in plan_row["description"]

    report = outcome.report
    assert report.workbook_plan_id == "mini-bop-demo-il-2026"
    assert report.vectors.status == "ran"
    assert (report.vectors.matched, report.vectors.mismatched) == (1, 0)
    assert report.manifest.counts.factor_cells == 6


def test_build_is_atomic(tmp_db: Database, monkeypatch: pytest.MonkeyPatch) -> None:
    def explode(**_kwargs):  # noqa: ANN003, ANN202
        raise RuntimeError("substrate write failed")

    monkeypatch.setattr(
        "openrater.rates.ingest.builder.bulk_upsert_factor_tables", explode
    )
    with pytest.raises(RuntimeError):
        build_workbook(db=tmp_db, data=to_bytes(build_mini()))

    conn = tmp_db.connection()
    for table in ("rating_plans", "rating_plan_stages", "plan_dimensions"):
        n = conn.execute(f"SELECT COUNT(*) AS n FROM {table}").fetchone()["n"]
        assert n == 0, f"{table} has {n} rows after a failed build"
    conn.close()


def test_build_rejects_dirty_workbook(tmp_db: Database) -> None:
    wb = build_mini()
    wb.remove(wb["inputs"])
    from openrater.rates.ingest.service import WorkbookNotCleanError

    with pytest.raises(WorkbookNotCleanError):
        build_workbook(db=tmp_db, data=to_bytes(wb))

    conn = tmp_db.connection()
    assert conn.execute("SELECT COUNT(*) AS n FROM rating_plans").fetchone()["n"] == 0
    conn.close()


# ---------------------------------------------------------------------------
# The endpoints.
# ---------------------------------------------------------------------------

def test_build_endpoint_report_and_duplicate_awareness(
    client, monkeypatch: pytest.MonkeyPatch  # noqa: ANN001
) -> None:
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    body = to_bytes(build_mini())

    resp = client.post("/api/v1/plans/ingest?filename=mini.xlsx", content=body)
    assert resp.status_code == 200, resp.text
    payload = resp.json()
    plan_id = payload["rating_plan_id"]
    assert payload["report"]["vectors"]["matched"] == 1

    report = client.get(f"/api/v1/plans/{plan_id}/build-report")
    assert report.status_code == 200
    assert report.json()["workbook_hash"] == payload["report"]["workbook_hash"]

    # CT-5 — re-checking the same bytes surfaces the existing plan.
    check = client.post("/api/v1/plans/ingest/check", content=body)
    assert check.status_code == 200
    assert check.json()["already_built"]["rating_plan_id"] == plan_id

    # The built plan is a live draft — the plans list knows it.
    listed = client.get(f"/api/v1/plans/{plan_id}")
    assert listed.status_code == 200

    missing = client.get("/api/v1/plans/nope/build-report")
    assert missing.status_code == 404
    assert missing.json()["error"]["code"] == "build_report_not_found"


def test_build_endpoint_dirty_workbook_is_422(client) -> None:  # noqa: ANN001
    wb = build_mini()
    wb.remove(wb["inputs"])
    resp = client.post("/api/v1/plans/ingest", content=to_bytes(wb))
    assert resp.status_code == 422
    assert resp.json()["error"]["code"] == "ingest_check_failed"


# ---------------------------------------------------------------------------
# The check=build contract on predicates + triggers (r4, 2026-07-15
# filing-digitization review). Zero builder-refusal paths were tested
# before this block.
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("op", "expected"),
    [("==", "eq"), ("!=", "ne"), ("<", "lt"), ("<=", "le"), (">", "gt"), (">=", "ge")],
)
def test_endorsement_trigger_scalar_operators_map(op: str, expected: str) -> None:
    from openrater.rates.ingest.builder import _endorsement_trigger

    t = _endorsement_trigger(f"form_input.building_age {op} 5")
    assert t == {"variable": "building_age", "op": expected, "value": 5}


def test_endorsement_trigger_in_builds_list_semantics() -> None:
    """`in`/`not-in` build REAL set semantics — before r4 the op_map
    silently rewrote any unmapped operator to `eq` (an endorsement
    attaching on the wrong condition, with no error anywhere)."""
    from openrater.rates.ingest.builder import _endorsement_trigger

    t = _endorsement_trigger(
        "form_input.construction_class in [frame, fire_resistive]"
    )
    assert t == {
        "variable": "construction_class",
        "op": "in",
        "value": ["frame", "fire_resistive"],
    }
    t = _endorsement_trigger("form_input.building_age not-in [5, 10.5, true]")
    assert t == {"variable": "building_age", "op": "nin", "value": [5, 10.5, True]}


def test_endorsement_trigger_inexpressible_is_refused_never_rewritten() -> None:
    from openrater.rates.ingest.builder import BuildError, _endorsement_trigger

    with pytest.raises(BuildError, match="cannot express"):
        _endorsement_trigger("form_input.building_age ~= 5")
    with pytest.raises(BuildError, match="list"):
        _endorsement_trigger("form_input.construction_class in frame")


@pytest.mark.parametrize("op", ["!=", "<", "<=", ">", ">=", "in", "not-in"])
def test_chain_predicate_beyond_equality_backstop_raises(op: str) -> None:
    """The builder backstop behind the R-190 check refusal: the
    domain's FactorPredicate is equality-only, so a non-`==` chain or
    loading predicate reaching the builder raises (never mis-gates)."""
    from openrater.rates.ingest.builder import BuildError, _equality_predicate

    value = "[frame]" if op in ("in", "not-in") else "5"
    with pytest.raises(BuildError, match="equality only"):
        _equality_predicate(f"form_input.building_age {op} {value}")
    assert _equality_predicate("form_input.building_age == 5") == {
        "path": "form_input.building_age",
        "equals": 5,
    }


def test_build_persists_in_trigger_with_real_semantics(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """End to end: an `in`-triggered factor endorsement lands in the
    stored plan with op `in` and the LIST value — not the pre-r4
    silent `eq` against the raw '[a, b]' string."""
    import json

    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    wb = build_mini()
    ws = wb.create_sheet("endorsements")
    ws.append(["endorsement_id", "kind", "form_number", "display_name",
               "factor", "amount", "coverage", "sublimit", "trigger"])
    ws.append(["frame_surcharge", "factor", "MS 20 01", "Frame surcharge",
               1.10, "", "", "",
               "form_input.construction_class in [frame, fire_resistive]"])
    outcome = build_workbook(db=tmp_db, data=to_bytes(wb), filename="mini.xlsx")

    conn = tmp_db.connection()
    row = conn.execute(
        "SELECT config_json FROM rating_plan_stages WHERE rating_plan_id = ? "
        "AND stage_kind = 'endorsement.factor'",
        (outcome.rating_plan_id,),
    ).fetchone()
    conn.close()
    assert row is not None
    config = json.loads(row["config_json"])
    assert config["trigger"] == {
        "variable": "construction_class",
        "op": "in",
        "value": ["frame", "fire_resistive"],
    }


def test_build_endpoint_builder_refusal_is_422_not_500(
    client, monkeypatch: pytest.MonkeyPatch
) -> None:  # noqa: ANN001
    """A BuildError escaping the builder maps to a structured 422
    `ingest_unbuildable` carrying the builder's message — before r4
    this was an opaque 500 internal_error. The natural repros are all
    check-refused now (R-190, R-146), so the backstop mapping is
    exercised by injecting the refusal at the route's build seam."""
    from openrater.app.routes import plan_ingest_route
    from openrater.rates.ingest.builder import BuildError

    def _refuse(**kwargs):  # noqa: ANN003, ANN202
        raise BuildError("the check passed but this construct cannot be expressed")

    monkeypatch.setattr(plan_ingest_route, "build_workbook", _refuse)
    resp = client.post("/api/v1/plans/ingest", content=to_bytes(build_mini()))
    assert resp.status_code == 422, resp.text
    err = resp.json()["error"]
    assert err["code"] == "ingest_unbuildable"
    assert "cannot be expressed" in err["message"]


def test_build_endpoint_coverage_total_without_round_is_check_refused(
    client,
) -> None:  # noqa: ANN001
    """The old natural repro for the builder backstop — coverage:total
    with no round row — now refuses at the CHECK (R-146), so the build
    endpoint reports it as `ingest_check_failed`, never reaching the
    builder."""
    wb = build_mini()
    wb["outputs"].append(
        ["out_total", "total_premium", "Total premium", "coverage:total"]
    )
    resp = client.post("/api/v1/plans/ingest", content=to_bytes(wb))
    assert resp.status_code == 422, resp.text
    err = resp.json()["error"]
    assert err["code"] == "ingest_check_failed"
    check = client.post("/api/v1/plans/ingest/check", content=to_bytes(wb))
    assert [e["rule"] for e in check.json()["errors"]] == ["R-146"]


# ---------------------------------------------------------------------------
# The LIVE golden path (env-gated; CI's contract job + local runs).
# ---------------------------------------------------------------------------

@pytest.mark.skipif(
    os.environ.get("RATER_INGEST_LIVE_SCORING") != "1",
    reason="needs a running scoring service (set RATER_INGEST_LIVE_SCORING=1)",
)
def test_canonical_bundle_builds_and_all_vectors_match(tmp_db: Database) -> None:
    """The Brief 92 acceptance seed: the canonical nonprofit bundle
    ingests hands-off and every one of the filing's 20 examples —
    premiums AND tiers — matches through the production engine."""
    outcome = build_workbook(
        db=tmp_db, data=CANONICAL.read_bytes(), filename=CANONICAL.name
    )
    v = outcome.report.vectors
    detail = [
        (c.case_id, c.field, c.expected, c.actual, c.status)
        for c in v.checks
        if c.status != "match"
    ]
    assert v.status == "ran", v.detail
    assert v.total_cases == 20
    assert v.mismatched == 0 and v.near == 0, detail
    assert v.matched == len(v.checks) == 80  # 20 cases × (3 premiums + tier)


def test_parse_reuse_for_build_matches_check() -> None:
    parsed, issues = parse_workbook(CANONICAL.read_bytes())
    assert not [i for i in issues if i.severity == "error"]
    assert parsed.plan_value("product") == "do"
    assert len(parsed.factor_tables) == 11


MERIDIAN = (
    Path(__file__).resolve().parents[2]
    / "docs" / "specs" / "examples" / "meridian-shopfront-bop"
    / "meridian_shopfront_bop.workbook.xlsx"
)


@pytest.mark.skipif(
    os.environ.get("RATER_INGEST_LIVE_SCORING") != "1",
    reason="needs a running scoring service (set RATER_INGEST_LIVE_SCORING=1)",
)
def test_meridian_all_constructs_builds_and_all_vectors_match(
    tmp_db: Database,
) -> None:
    """The all-constructs golden: exposure towers (the engine's
    own tower rounding), the 2-D matrix, geo ZIP→territory resolution,
    per-tip endorsements (the shared-node fix), loadings, the
    per-coverage clamp, and the composition-seam package floor — all
    40 checks (8 cases × 3 towers + total + tier) match."""
    outcome = build_workbook(
        db=tmp_db, data=MERIDIAN.read_bytes(), filename=MERIDIAN.name
    )
    v = outcome.report.vectors
    detail = [
        (c.case_id, c.field, c.expected, c.actual, c.status)
        for c in v.checks
        if c.status != "match"
    ]
    assert v.status == "ran", v.detail
    assert v.total_cases == 8
    assert v.mismatched == 0 and v.near == 0, detail
    assert v.matched == len(v.checks) == 40


# ---------------------------------------------------------------------------
# §4.6 binding grammar forms (R-127) — resolved at build, executed live.
# ---------------------------------------------------------------------------


def build_binding_forms() -> Workbook:
    """The §4.6 binding-grammar workbook: a binding-only literal base,
    a `literal:<n>` fixed exposure, an lcm bound `context.lcm` (the
    plan sheet's value, cited per block), and an lcm bound
    `literal:<n>` — every form R-127 admits beyond `form_input.*`.
    Expected premiums are hand-computed filed math, pinned engine-side
    by @openrater/ui `literalBindingForms.test.ts` with the same numbers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "plan"
    for row in (
        ("field", "value"),
        ("spec_version", "1.0"),
        ("rating_plan_id", "binding-forms-demo-il-2026"),
        ("display_name", "Binding grammar forms demo"),
        ("version", "1.0.0"),
        ("carrier", "Demo Mutual"),
        ("product", "bop"),
        ("jurisdiction_country", "US"),
        ("state", "IL"),
        ("effective_date", "2026-01-01"),
        ("coverages", "building,contents"),
        ("lcm", 1.30),
    ):
        ws.append(list(row))

    ws = wb.create_sheet("inputs")
    ws.append(["name", "label", "data_type", "required", "allowed_values", "default_value", "unit"])
    ws.append(["construction_class", "Construction class", "enum", True, "frame,fire_resistive", "", ""])
    ws.append(["cv", "Contents value", "currency", True, "", "", "USD"])

    ws = wb.create_sheet("dimensions")
    ws.append(["slug", "display_name", "shape", "role", "data_type", "dimension_type",
               "geo_granularity", "geo_scope", "axes"])
    ws.append(["construction_class", "Construction class", "categorical", "both", "enum", "standard", "", "", ""])

    ws = wb.create_sheet("dimension_levels")
    ws.append(["dimension_slug", "kind", "level_id", "label", "aliases", "min", "max", "territory_ref"])
    ws.append(["construction_class", "categorical", "frame", "Frame", "", "", "", ""])
    ws.append(["construction_class", "categorical", "fire_resistive", "Fire-resistive", "", "", "", ""])

    ws = wb.create_sheet("ft.construction_rel")
    for row in (
        ("table_id", "construction_rel"),
        ("display_name", "Construction factor"),
        ("dimensionality", "1d"),
        ("row_dimension", "construction_class"),
        ("lookup_method", "direct"),
        ("citation_rule", "Table 1.A"),
        ("citation_page", "p.1"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["level_id", "factor", "citation_rule", "citation_page"])
    ws.append(["frame", 1.00, "Table 1.A", "p.1"])
    ws.append(["fire_resistive", 0.78, "Table 1.A", "p.1"])

    ws = wb.create_sheet("chains")
    ws.append(["coverage", "order", "stage_kind", "stage_id", "factor_table", "dimension",
               "input_binding", "value", "exposure_divisor", "predicate"])
    # building — binding-only base, fixed literal exposure, context.lcm.
    ws.append(["building", 0, "base", "bld_base", "", "", "literal:4.000", "", "", ""])
    ws.append(["building", 1, "lookup.direct", "bld_constr", "ft.construction_rel",
               "construction_class", "", "", "", ""])
    ws.append(["building", 2, "exposure", "bld_exposure", "", "", "literal:250", "", 100, ""])
    ws.append(["building", 3, "lcm", "bld_lcm", "", "", "context.lcm", "", "", ""])
    # contents — per-risk exposure, literal lcm binding.
    ws.append(["contents", 0, "base", "cnt_base", "", "", "", 1.5, "", ""])
    ws.append(["contents", 1, "lookup.direct", "cnt_constr", "ft.construction_rel",
               "construction_class", "", "", "", ""])
    ws.append(["contents", 2, "exposure", "cnt_exposure", "", "", "form_input.cv", "", 100, ""])
    ws.append(["contents", 3, "lcm", "cnt_lcm", "", "", "literal:1.10", "", "", ""])

    ws = wb.create_sheet("outputs")
    ws.append(["output_id", "field_name", "display_name", "source"])
    ws.append(["out_building", "building_premium", "Building premium", "bld_base"])
    ws.append(["out_contents", "contents_premium", "Contents premium", "cnt_base"])

    ws = wb.create_sheet("test_cases")
    ws.append(["case_id", "name", "construction_class", "cv",
               "expected_building_premium", "expected_contents_premium"])
    # building: 4.000 × frame 1.0 → ×(250÷100) = 10 → ×1.30 = 13
    # contents: 1.500 × frame 1.0 → ×(10000÷100) = 150 → ×1.10 = 165
    ws.append(["tc_1", "Frame, $10k contents", "frame", 10000, 13.00, 165.00])
    # building: 4.000 × 0.78 = 3.120 → ×2.5 = 7.8 → ×1.30 = 10.14 → $10
    # contents: 1.500 × 0.78 = 1.170 → ×200 = 234 → ×1.10 = 257.4 → $257
    ws.append(["tc_2", "Fire-resistive, $20k contents", "fire_resistive", 20000, 10.00, 257.00])

    ws = wb.create_sheet("gaps_and_assumptions")
    ws.append(["kind", "description", "citation_rule", "citation_page", "impact", "related"])
    return wb


def test_binding_forms_resolve_at_build(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The builder resolves what only it can see: context.lcm → the
    plan sheet's value, literal lcm/base bindings → numbers. The
    literal EXPOSURE stays a passthrough — the projector executes it
    (there is no constant-exposure field in the chain contract)."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 13.0}),
    )
    outcome = build_workbook(
        db=tmp_db, data=to_bytes(build_binding_forms()), filename="forms.xlsx"
    )

    conn = tmp_db.connection()
    chain_cfg = conn.execute(
        "SELECT config_json FROM rating_plan_stages WHERE rating_plan_id = ? "
        "AND stage_kind = 'multiplicative_chain'",
        (outcome.rating_plan_id,),
    ).fetchone()
    conn.close()
    chains = {
        c["coverage_value"]: c for c in json.loads(chain_cfg["config_json"])["chains"]
    }

    bld = chains["building"]
    assert bld["base_value"] == 4.0  # binding-only base resolved
    assert bld["exposure_input"] == "literal:250"  # projector executes this
    assert bld["apply_exposure"] is True
    assert bld["lcm"]["value"] == 1.30  # context.lcm ← plan sheet
    assert bld["lcm"]["input_path"] is None

    cnt = chains["contents"]
    assert cnt["base_value"] == 1.5
    assert cnt["exposure_input"] == "form_input.cv"
    assert cnt["lcm"]["value"] == 1.10  # literal binding canonicalized
    assert cnt["lcm"]["input_path"] is None

    assert outcome.report.vectors.status == "ran"
    assert outcome.report.vectors.total_cases == 2


@pytest.mark.skipif(
    os.environ.get("RATER_INGEST_LIVE_SCORING") != "1",
    reason="needs a running scoring service (set RATER_INGEST_LIVE_SCORING=1)",
)
def test_binding_forms_vectors_match_live(tmp_db: Database) -> None:
    """The grammar-forms golden: a workbook whose building tower has NO
    per-risk exposure input (fixed `literal:250`) and whose LCMs arrive
    via `context.lcm` + `literal:1.10` rates hands-off, and both filed
    examples reproduce exactly through the production engine."""
    outcome = build_workbook(
        db=tmp_db, data=to_bytes(build_binding_forms()), filename="forms.xlsx"
    )
    v = outcome.report.vectors
    detail = [
        (c.case_id, c.field, c.expected, c.actual, c.status)
        for c in v.checks
        if c.status != "match"
    ]
    assert v.status == "ran", v.detail
    assert v.total_cases == 2
    assert v.mismatched == 0 and v.near == 0, detail
    assert v.matched == len(v.checks) == 4  # 2 cases × 2 premiums


@pytest.mark.skipif(
    not os.environ.get("RATER_ACCEPTANCE_WORKBOOK"),
    reason="set RATER_ACCEPTANCE_WORKBOOK=/path/to/workbook.xlsx (needs live scoring)",
)
def test_acceptance_workbook_hands_off(tmp_db: Database) -> None:
    """Optionally ingest any local spec-v1.0 workbook and print the full
    vector table. The workbook is read by path and never enters the
    repository. Mismatches, unavailable scoring, and error-status checks
    all fail visibly."""
    path = Path(os.environ["RATER_ACCEPTANCE_WORKBOOK"])
    outcome = build_workbook(db=tmp_db, data=path.read_bytes(), filename=path.name)
    v = outcome.report.vectors
    print(
        f"\nacceptance: {path.name} -> {outcome.rating_plan_id} · "
        f"{v.matched} match / {v.near} near / {v.mismatched} mismatch "
        f"of {len(v.checks)} checks ({v.total_cases} cases)"
    )
    for c in v.checks:
        if c.status != "match":
            print(f"  {c.status.upper():<9} {c.case_id} {c.field}: "
                  f"expected {c.expected} got {c.actual} (Δ {c.delta})")
    assert v.status == "ran", v.detail
    assert not [c for c in v.checks if c.status == "error"]


def test_ingest_sources_carry_no_program_literals() -> None:
    """The ingest pipeline contains no program-specific knowledge;
    every program-shaped value must arrive as workbook data."""
    import re

    root = Path(__file__).resolve().parents[1] / "src" / "openrater" / "rates" / "ingest"
    # Word-bounded program tokens. Bare "iso" is exempted where it means
    # the DATE FORMAT ("ISO 8601", "ISO date", _iso_date) — the
    # insurance bureau reads as "iso bop" / "insurance services office".
    forbidden = re.compile(
        r"\b(iso[ _-]?bop|insurance services office|meridian|kansas|"
        r"cincinnati|nonprofit|ntee|bowling|bmut)\b",
        re.IGNORECASE,
    )
    hits: list[str] = []
    for path in sorted(root.glob("*.py")):
        for m in forbidden.finditer(path.read_text(encoding="utf-8")):
            hits.append(f"{path.name}: {m.group(0)!r}")
    assert hits == [], f"program literals in ingest sources: {hits}"


def test_template_is_alive_build_side(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The packaged starter-kit template builds a real plan and its
    spec-§9 vector verifies ($390.00)."""
    from importlib import resources

    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    data = (
        resources.files("openrater.rates.ingest")
        .joinpath("assets")
        .joinpath("openrater_workbook_template.xlsx")
        .read_bytes()
    )
    outcome = build_workbook(
        db=tmp_db, data=data, filename="openrater_workbook_template.xlsx"
    )
    vectors = outcome.report.vectors
    assert vectors.status == "ran"
    assert vectors.mismatched == 0 and vectors.matched >= 1, vectors


# ---------------------------------------------------------------------------
# Gate-default citation, envelope verdict, and single-parse guarantees.
# ---------------------------------------------------------------------------

def _add_gates_with_cited_default(wb) -> None:  # noqa: ANN001 — Workbook
    ws = wb.create_sheet("gates")
    ws.append(["order", "rule_id", "variable", "op", "value",
               "variable_2", "op_2", "value_2",
               "variable_3", "op_3", "value_3",
               "tier", "reasoning", "citation_rule", "citation_page"])
    ws.append([1, "old_building", "building_age", "ge", 40,
               "", "", "", "", "", "",
               "submit", "Buildings 40+ years refer.", "Rule 6.A", "p.61"])
    ws.append([2, "__default__", "building_age", "ge", 0,
               "", "", "", "", "", "",
               "standard", "Everything else writes standard.", "Rule 6.C", "p.62"])


def test_gate_default_citation_lands_in_config(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The filed `__default__` row's citation was
    silently dropped (construct-audit gap 7); it now lands in the gate
    config's `default_citation` — typed, additive."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}, tier="standard"),
    )
    wb = build_mini()
    _add_gates_with_cited_default(wb)
    outcome = build_workbook(db=tmp_db, data=to_bytes(wb), filename="mini.xlsx")

    conn = tmp_db.connection()
    row = conn.execute(
        "SELECT config_json FROM rating_plan_stages WHERE rating_plan_id = ? "
        "AND stage_kind = 'eligibility.gate'",
        (outcome.rating_plan_id,),
    ).fetchone()
    assert row is not None
    cfg = json.loads(row["config_json"])
    assert cfg["default_tier"] == "standard"
    assert cfg["default_citation"] == "Rule 6.C"
    assert cfg["rules"][0]["citation"] == "Rule 6.A"


def test_verification_verdict_mapping() -> None:
    """The envelope verdict reports each of its five states."""
    from openrater.rates.ingest.reports import VectorsSummary, verification_verdict

    ran = {"status": "ran", "total_cases": 2}
    assert verification_verdict(VectorsSummary(**ran, matched=4)) == "all_match"
    assert verification_verdict(VectorsSummary(**ran, matched=3, near=1)) == "near"
    assert (
        verification_verdict(VectorsSummary(**ran, matched=3, near=1, mismatched=1))
        == "mismatches"
    )
    assert verification_verdict(VectorsSummary(status="none")) == "none"
    assert (
        verification_verdict(VectorsSummary(status="unavailable", detail="down"))
        == "unavailable"
    )


def test_build_response_carries_verification(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    outcome = build_workbook(db=tmp_db, data=to_bytes(build_mini()), filename="m.xlsx")
    assert outcome.verification == "all_match"


def test_build_parses_the_bytes_once(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The check's parsed model is the build's input —
    the bytes were previously parsed twice per build."""
    from openrater.rates.ingest import service as ingest_service

    calls = {"n": 0}
    real = ingest_service.parse_workbook

    def counting(data: bytes):  # noqa: ANN202
        calls["n"] += 1
        return real(data)

    monkeypatch.setattr(ingest_service, "parse_workbook", counting)
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    build_workbook(db=tmp_db, data=to_bytes(build_mini()), filename="m.xlsx")
    assert calls["n"] == 1, f"parse_workbook ran {calls['n']}× per build"


def test_personal_lines_product_builds_end_to_end(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """`homeowners` is a first-class product
    code — the check accepts it (R-028), the builder maps it, and the
    INSERT passes migration 048's widened CHECK."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    wb = build_mini()
    ws = wb["plan"]
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "product":
            ws.cell(row=r, column=2).value = "homeowners"
            break
    outcome = build_workbook(db=tmp_db, data=to_bytes(wb), filename="ho.xlsx")

    conn = tmp_db.connection()
    row = conn.execute(
        "SELECT product FROM rating_plans WHERE rating_plan_id = ?",
        (outcome.rating_plan_id,),
    ).fetchone()
    assert row["product"] == "homeowners"


# ---------------------------------------------------------------------------
# Brief 92.R phase 92R.1 — the base: bytes stored (D1), revision
# discovery (D2), report history.
# ---------------------------------------------------------------------------

def _bump_mini_revision(wb) -> None:  # noqa: ANN001 — Workbook
    """The mini workbook, re-issued: version bumped per spec §4.1 and
    one factor changed — same workbook_plan_id, different bytes."""
    ws = wb["plan"]
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "version":
            ws.cell(row=r, column=2).value = "1.1.0"
    ft = wb["ft.construction_class"]
    for r in range(1, ft.max_row + 1):
        if ft.cell(row=r, column=1).value == "frame":
            ft.cell(row=r, column=2).value = 1.05


def test_build_stores_workbook_bytes_and_version(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """D1: the exact bytes land on the report row (the base of every
    future diff); the workbook's own version is captured; and neither
    rides the read models."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    data = to_bytes(build_mini())
    outcome = build_workbook(db=tmp_db, data=data, filename="mini.xlsx")

    conn = tmp_db.connection()
    row = conn.execute(
        "SELECT workbook_blob, workbook_version FROM plan_build_reports "
        "WHERE rating_plan_id = ?",
        (outcome.rating_plan_id,),
    ).fetchone()
    assert row["workbook_blob"] == data
    assert row["workbook_version"] == "1.0.0"
    assert outcome.report.workbook_version == "1.0.0"
    assert not hasattr(outcome.report, "workbook_blob")


def test_workbook_export_round_trips_hash_identical(
    monkeypatch: pytest.MonkeyPatch, client  # noqa: ANN001
) -> None:
    """The canonical workbook container comes back out:
    GET /workbook serves the EXACT ingested bytes with the recorded
    hash, and re-ingesting the download answers already_built (the
    round-trip honesty test). A plan with no stored workbook 404s
    BY NAME."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    original = to_bytes(build_mini())
    built = client.post(
        "/api/v1/plans/ingest?filename=mini.xlsx", content=original
    )
    assert built.status_code == 200
    plan_id = built.json()["rating_plan_id"]

    res = client.get(f"/api/v1/plans/{plan_id}/workbook")
    assert res.status_code == 200
    assert res.content == original
    assert res.headers["X-Workbook-Hash"]
    assert 'filename="mini.xlsx"' in res.headers["Content-Disposition"]

    # Round trip: the export re-ingested is byte-identical.
    dup = client.post("/api/v1/plans/ingest/check", content=res.content).json()
    assert dup["already_built"] is not None

    # No workbook stored → the boundary is NAMED, never a bare 404.
    missing = client.get("/api/v1/plans/never-built/workbook")
    assert missing.status_code == 404
    assert missing.json()["error"]["code"] == "workbook_not_stored"


def test_revision_discovery_on_check(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch, client  # noqa: ANN001
) -> None:
    """D2 end-to-end over the endpoint: byte-identical → already_built
    (never a fake revision); same id + new bytes → revises with the
    plan's name and v1.0.0 → v1.1.0; an unseen id → neither."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    original = to_bytes(build_mini())
    resp = client.post("/api/v1/plans/ingest?filename=mini.xlsx", content=original)
    assert resp.status_code == 200

    # Byte-identical drop: already_built, revises stays empty (CT-6 of
    # the brief: no fake revision).
    dup = client.post("/api/v1/plans/ingest/check", content=original).json()
    assert dup["already_built"] is not None
    assert dup["revises"] is None

    # The re-issued workbook: a revision candidate, versions named.
    revised = build_mini()
    _bump_mini_revision(revised)
    rev = client.post("/api/v1/plans/ingest/check", content=to_bytes(revised)).json()
    assert rev["already_built"] is None
    assert rev["revises"] is not None
    assert rev["revises"]["display_name"] == "Mini BOP demo"
    assert rev["revises"]["version_from"] == "1.0.0"
    assert rev["revises"]["version_to"] == "1.1.0"

    # A different workbook identity: plain build, no revision claim.
    stranger = build_mini()
    ws = stranger["plan"]
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "rating_plan_id":
            ws.cell(row=r, column=2).value = "someone-else-entirely-2026"
    other = client.post("/api/v1/plans/ingest/check", content=to_bytes(stranger)).json()
    assert other["already_built"] is None
    assert other["revises"] is None


def test_build_report_history_endpoint(
    monkeypatch: pytest.MonkeyPatch, client  # noqa: ANN001
) -> None:
    """The history read model: newest first; a plan with no reports is
    an empty list, not a 404 (collection semantics). Everything rides
    the client's own DB (the fixture provisions a throwaway)."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    built = client.post(
        "/api/v1/plans/ingest?filename=a.xlsx", content=to_bytes(build_mini())
    ).json()
    plan_id = built["rating_plan_id"]
    # Re-ingest lands in 92R.3; a second row via the insert seam stands
    # in for it here (reports are append-only by design).
    from openrater.rates.ingest.reports import (
        Manifest,
        VectorsSummary,
        insert_build_report,
    )

    insert_build_report(
        db=client.app.state.db,
        rating_plan_id=plan_id,
        workbook_hash="deadbeef",
        filename="b.xlsx",
        spec_version="1.0",
        workbook_plan_id="mini-bop-demo-il-2026",
        manifest=Manifest.model_validate(built["report"]["manifest"]),
        issues=[],
        vectors=VectorsSummary.model_validate(built["report"]["vectors"]),
        gaps=[],
        workbook_version="1.1.0",
    )
    resp = client.get(f"/api/v1/plans/{plan_id}/build-reports")
    assert resp.status_code == 200
    payload = resp.json()
    assert [r["filename"] for r in payload] == ["b.xlsx", "a.xlsx"]
    assert payload[0]["workbook_version"] == "1.1.0"
    assert "workbook_blob" not in payload[0]

    empty = client.get("/api/v1/plans/never_built_anything/build-reports")
    assert empty.status_code == 200
    assert empty.json() == []


# ---------------------------------------------------------------------------
# Brief 95 A2 — the workbook's rating_plan_id IS the built plan's id.
# ---------------------------------------------------------------------------


def test_build_pins_the_workbook_plan_id(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Same workbook → same plan id, on any box (Brief 95 A2)."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    outcome = build_workbook(db=tmp_db, data=to_bytes(build_mini()), filename="m.xlsx")
    assert outcome.rating_plan_id == "mini-bop-demo-il-2026"

    conn = tmp_db.connection()
    row = conn.execute(
        "SELECT display_name FROM rating_plans WHERE rating_plan_id = ?",
        ("mini-bop-demo-il-2026",),
    ).fetchone()
    conn.close()
    assert row is not None and row["display_name"] == "Mini BOP demo"


def test_identical_rebuild_refuses_already_built(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Identical bytes against the pinned id: nothing to do — refuse
    with the existing plan named, never a duplicate (A2/A3)."""
    from openrater.rates.ingest.service import PlanAlreadyBuiltError

    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    data = to_bytes(build_mini())
    build_workbook(db=tmp_db, data=data, filename="m.xlsx")

    with pytest.raises(PlanAlreadyBuiltError) as exc:
        build_workbook(db=tmp_db, data=data, filename="m.xlsx")
    assert exc.value.already.rating_plan_id == "mini-bop-demo-il-2026"

    conn = tmp_db.connection()
    n = conn.execute("SELECT COUNT(*) AS n FROM rating_plans").fetchone()["n"]
    conn.close()
    assert n == 1, "the refused rebuild must not create a duplicate plan"


def test_taken_id_with_different_bytes_points_at_reingest(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """A revision dropped on the BUILD door refuses toward the
    re-ingest door — the diff flow updates the plan; build never
    silently duplicates it (A2)."""
    from openrater.rates.ingest.service import PlanIdTakenError

    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    build_workbook(db=tmp_db, data=to_bytes(build_mini()), filename="m.xlsx")

    revised = build_mini()
    _bump_mini_revision(revised)
    with pytest.raises(PlanIdTakenError) as exc:
        build_workbook(db=tmp_db, data=to_bytes(revised), filename="m2.xlsx")
    assert "reingest" in str(exc.value)
    assert exc.value.rating_plan_id == "mini-bop-demo-il-2026"

    conn = tmp_db.connection()
    n = conn.execute("SELECT COUNT(*) AS n FROM rating_plans").fetchone()["n"]
    conn.close()
    assert n == 1


def test_build_endpoint_conflicts_are_409s(
    monkeypatch: pytest.MonkeyPatch, client  # noqa: ANN001
) -> None:
    """The route maps the two A2 refusals to structured 409s."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    original = to_bytes(build_mini())
    assert client.post("/api/v1/plans/ingest", content=original).status_code == 200

    dup = client.post("/api/v1/plans/ingest", content=original)
    assert dup.status_code == 409
    assert dup.json()["error"]["code"] == "ingest_already_built"

    revised = build_mini()
    _bump_mini_revision(revised)
    taken = client.post("/api/v1/plans/ingest", content=to_bytes(revised))
    assert taken.status_code == 409
    assert taken.json()["error"]["code"] == "ingest_plan_id_taken"
    assert "reingest" in taken.json()["error"]["hint"]


# ---------------------------------------------------------------------------
# Geo territory_ref → level_id join (found by the transcription eval).
# ---------------------------------------------------------------------------


def build_geo_ref_spelling() -> Workbook:
    """A minimal geo program whose dimension_levels declare the join
    with a DIFFERENT spelling than the level ids: level t1/t2, ref
    T1/T2, geo rows written in the ref spelling. R-082 accepts this
    (the join is declared); the factor table rates by LEVEL id — so
    the builder must translate territory codes through the join, or
    every lookup dies at rate time with `unknown_key T1`."""
    wb = Workbook()
    ws = wb.active
    ws.title = "plan"
    for row in (
        ("field", "value"),
        ("spec_version", "1.0"),
        ("rating_plan_id", "geo-ref-spelling-ne-2026"),
        ("display_name", "Geo ref spelling demo"),
        ("version", "1.0.0"),
        ("carrier", "Demo Mutual"),
        ("product", "bop"),
        ("jurisdiction_country", "US"),
        ("state", "NE"),
        ("effective_date", "2026-01-01"),
        ("coverages", "building"),
        ("lcm", 1.0),
    ):
        ws.append(list(row))

    ws = wb.create_sheet("inputs")
    ws.append(["name", "label", "data_type", "required", "allowed_values", "default_value", "unit"])
    ws.append(["zip", "ZIP code", "string", True, "", "", ""])
    ws.append(["building_limit", "Building limit", "currency", True, "", "", "USD"])

    ws = wb.create_sheet("dimensions")
    ws.append(["slug", "display_name", "shape", "role", "data_type", "dimension_type",
               "geo_granularity", "geo_scope", "axes"])
    ws.append(["territory", "Territory", "geographic", "rating-input", "string",
               "geographic", "zip", "subset:NE", ""])

    ws = wb.create_sheet("dimension_levels")
    ws.append(["dimension_slug", "kind", "level_id", "label", "aliases", "min", "max", "territory_ref"])
    ws.append(["territory", "geographic", "t1", "Territory 1", "", "", "", "T1"])
    ws.append(["territory", "geographic", "t2", "Territory 2", "", "", "", "T2"])

    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code", "citation_rule", "citation_page"])
    ws.append(["68001", "T1", "Rule D.1", "p.6"])
    ws.append(["68102", "T2", "Rule D.1", "p.6"])

    ws = wb.create_sheet("ft.territory_rel")
    for row in (
        ("table_id", "territory_rel"),
        ("display_name", "Territory factor"),
        ("dimensionality", "1d"),
        ("row_dimension", "territory"),
        ("lookup_method", "direct"),
        ("citation_rule", "Rule C.6"),
        ("citation_page", "p.5"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["level_id", "factor", "citation_rule", "citation_page"])
    ws.append(["t1", 0.90, "Rule C.6", "p.5"])
    ws.append(["t2", 1.10, "Rule C.6", "p.5"])

    ws = wb.create_sheet("chains")
    ws.append(["coverage", "order", "stage_kind", "stage_id", "factor_table", "dimension",
               "input_binding", "value", "exposure_divisor", "predicate"])
    ws.append(["building", 0, "base", "bld_base", "", "", "", 0.20, "", ""])
    ws.append(["building", 1, "lookup.direct", "bld_terr", "ft.territory_rel",
               "territory", "", "", "", ""])
    ws.append(["building", 2, "exposure", "bld_exposure", "", "",
               "form_input.building_limit", "", 100, ""])

    ws = wb.create_sheet("outputs")
    ws.append(["output_id", "field_name", "display_name", "source"])
    ws.append(["out_building", "building_premium", "Building premium", "bld_base"])

    ws = wb.create_sheet("test_cases")
    ws.append(["case_id", "name", "zip", "building_limit", "expected_building_premium"])
    ws.append(["tc_1", "T2 risk", "68102", 100000, 220.00])

    ws = wb.create_sheet("gaps_and_assumptions")
    ws.append(["kind", "description", "citation_rule", "citation_page", "impact", "related"])
    return wb


def test_geo_territory_ref_join_translates_to_level_ids(
    tmp_db: Database, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The built territory GROUP ids are the LEVEL ids (what factor
    tables rate by), never the geo sheet's ref spelling."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 220.0}),
    )
    outcome = build_workbook(
        db=tmp_db, data=to_bytes(build_geo_ref_spelling()), filename="geo.xlsx"
    )

    conn = tmp_db.connection()
    row = conn.execute(
        "SELECT geo_territories_json FROM plan_dimensions "
        "WHERE rating_plan_id = ? AND dim_id = 'territory'",
        (outcome.rating_plan_id,),
    ).fetchone()
    conn.close()
    territories = {t["id"]: t for t in json.loads(row[0])}
    assert set(territories) == {"t1", "t2"}, (
        "territory group ids must be the LEVEL ids — the ref spelling "
        f"leaked through: {sorted(territories)}"
    )
    assert territories["t2"]["members"] == ["68102"]
    assert territories["t1"]["label"] == "Territory 1"
