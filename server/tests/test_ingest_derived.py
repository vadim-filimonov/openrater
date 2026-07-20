# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""Brief 95 phase 95.4 — submitter ergonomics.

C2 (derived inputs): the inputs sheet's `derived_from = sum(a,b,…)`
column — grammar + operand validation (R-045), never-row-supplied
(R-046), lookup-axes-only usage (R-047), and the build side (input
stage `source="derived"` + ComputedExpr, lookup axes bound as
computed sums).

D1 (book template): `GET /plans/{id}/book-template.csv` — headers in
stage order, derived inputs excluded, the filing's own first test case
as the example row.

D2 (sample seeding): test-case INPUTS persisted on the build report
(`vectors.cases`), and pre-95.4 reports parsing with `cases=[]`.
"""

from __future__ import annotations

import csv
import io

import pytest
from openpyxl import Workbook

from openrater.rates.ingest.reports import VectorsSummary
from tests.test_ingest_build import _fake_score, tmp_db  # noqa: F401 — fixture
from tests.test_ingest_check import (
    _add_valid_gates,
    build_mini,
    run_check,
    to_bytes,
)


# ---------------------------------------------------------------------------
# The mini workbook, extended with one derived input feeding a 1-D
# lookup: total_limit = sum(tiv, bpp_limit) keys ft.limit_band.
# ---------------------------------------------------------------------------

def build_derived() -> Workbook:
    wb = build_mini()

    ws = wb["inputs"]
    ws.cell(row=1, column=8, value="maps_to_dimension")
    ws.cell(row=1, column=9, value="derived_from")
    ws.append(["bpp_limit", "BPP limit", "currency", True, "", "", "USD", "", ""])
    ws.append([
        "total_limit", "Total property limit", "currency", False, "", "", "USD",
        "limit_band", "sum(tiv,bpp_limit)",
    ])

    ws = wb["dimensions"]
    ws.append(["limit_band", "Limit band", "banded", "rating-input", "number",
               "standard", "", "", ""])
    ws = wb["dimension_levels"]
    ws.append(["limit_band", "banded", "band_lo", "≤ $500k", "", 0, 500000, ""])
    ws.append(["limit_band", "banded", "band_hi", "> $500k", "", 500000, "+inf", ""])

    ws = wb.create_sheet("ft.limit_band")
    for row in (
        ("table_id", "limit_band"),
        ("display_name", "Limit band factor"),
        ("dimensionality", "1d"),
        ("row_dimension", "limit_band"),
        ("lookup_method", "direct"),
        ("citation_rule", "Table 9.Z"),
        ("citation_page", "p.99"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["level_id", "factor", "citation_rule", "citation_page"])
    ws.append(["band_lo", 1.00, "Table 9.Z", "p.99"])
    ws.append(["band_hi", 1.10, "Table 9.Z", "p.99"])

    idx = wb.sheetnames.index("chains")
    wb.remove(wb["chains"])
    ws = wb.create_sheet("chains", idx)
    ws.append(["coverage", "order", "stage_kind", "stage_id", "factor_table",
               "dimension", "input_binding", "value", "exposure_divisor",
               "predicate"])
    for row in (
        ("building", 0, "base", "bld_base", "", "", "literal:0.150", 0.150, "", ""),
        ("building", 1, "lookup.direct", "bld_constr", "ft.construction_class",
         "construction_class", "", "", "", ""),
        ("building", 2, "lookup.multi", "bld_constr_age", "ft.constr_x_age",
         "construction_class", "", "", "", ""),
        ("building", 3, "lookup.direct", "bld_limit_band", "ft.limit_band",
         "limit_band", "", "", "", ""),
        ("building", 4, "exposure", "bld_exposure", "", "", "form_input.tiv",
         "", 100, ""),
        ("building", 5, "lcm", "bld_lcm", "", "", "", 1.30, "", ""),
    ):
        ws.append(list(row))

    # tc_1 supplies the operands, never the derived value.
    ws = wb["test_cases"]
    ws.cell(row=1, column=7, value="bpp_limit")
    ws.cell(row=2, column=7, value=100000)

    return wb


def _errors_for(result, rule: str) -> list[str]:
    return [i.message for i in result.errors if i.rule == rule]


def _set_derived_from(wb: Workbook, expr: str) -> None:
    ws = wb["inputs"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "total_limit":
            ws.cell(row=r, column=9, value=expr)
            return
    raise AssertionError("total_limit row not found")


# ---------------------------------------------------------------------------
# C2 — the clean path.
# ---------------------------------------------------------------------------

def test_derived_workbook_checks_clean() -> None:
    result = run_check(build_derived())
    assert result.errors == [], [(i.rule, i.message) for i in result.errors]
    assert result.ok is True
    assert result.manifest is not None
    assert result.manifest.counts.inputs == 5  # incl. the derived one


# ---------------------------------------------------------------------------
# C2 — R-045: the grammar + operand contract.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize(
    ("expr", "needle"),
    [
        ("div(payroll,100)", "named-deferred"),
        ("sum tiv bpp_limit", "doesn't parse"),
        ("sum(tiv)", "at least two operands"),
        ("sum(tiv,ghost_input)", "not a declared input"),
        ("sum(tiv,construction_class)", "must be number or currency"),
    ],
)
def test_r045_rejects_bad_expressions(expr: str, needle: str) -> None:
    wb = build_derived()
    _set_derived_from(wb, expr)
    result = run_check(wb)
    messages = _errors_for(result, "R-045")
    assert any(needle in m for m in messages), (expr, messages)


def test_r045_rejects_derived_operand() -> None:
    """A derived input cannot feed another derived input."""
    wb = build_derived()
    ws = wb["inputs"]
    ws.append([
        "stacked", "Stacked", "currency", False, "", "", "USD", "",
        "sum(tiv,total_limit)",
    ])
    result = run_check(wb)
    assert any(
        "itself derived" in m for m in _errors_for(result, "R-045")
    ), _errors_for(result, "R-045")


def test_r045_rejects_non_numeric_derived_type() -> None:
    wb = build_derived()
    ws = wb["inputs"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "total_limit":
            ws.cell(row=r, column=3, value="enum")
    result = run_check(wb)
    assert any(
        "its data_type must be number or currency" in m
        for m in _errors_for(result, "R-045")
    )


# ---------------------------------------------------------------------------
# C2 — R-046: derived means never row-supplied.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("column,value", [(4, True), (6, 123456)])
def test_r046_rejects_required_or_default(column: int, value) -> None:  # noqa: ANN001
    wb = build_derived()
    ws = wb["inputs"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "total_limit":
            ws.cell(row=r, column=column, value=value)
    result = run_check(wb)
    assert _errors_for(result, "R-046"), [
        (i.rule, i.message) for i in result.errors
    ]


# ---------------------------------------------------------------------------
# C2 — R-047: lookup axes only.
# ---------------------------------------------------------------------------

def test_r047_refuses_derived_in_chain_binding() -> None:
    wb = build_derived()
    ws = wb["chains"]
    ws.cell(row=6, column=7, value="form_input.total_limit")  # the exposure row
    result = run_check(wb)
    assert any(
        "input_binding" in m for m in _errors_for(result, "R-047")
    ), _errors_for(result, "R-047")


def test_r047_refuses_derived_in_predicate() -> None:
    wb = build_derived()
    ws = wb["chains"]
    ws.cell(row=5, column=10, value="form_input.total_limit == 100000")
    result = run_check(wb)
    assert any("predicate" in m for m in _errors_for(result, "R-047"))


def test_r047_refuses_derived_in_gates_variable() -> None:
    wb = build_derived()
    _add_valid_gates(wb)
    wb["gates"].cell(row=2, column=3, value="total_limit")
    result = run_check(wb)
    assert any("gates variable" in m for m in _errors_for(result, "R-047"))


def test_r047_warns_on_derived_test_case_column() -> None:
    wb = build_derived()
    ws = wb["test_cases"]
    ws.cell(row=1, column=8, value="total_limit")
    ws.cell(row=2, column=8, value=300000)
    result = run_check(wb)
    warned = [i for i in result.warnings if i.rule == "R-047"]
    assert warned and "ignored" in warned[0].message
    assert result.ok is True  # a warning, not a blocker


# ---------------------------------------------------------------------------
# C2 + D2 — the build side.
# ---------------------------------------------------------------------------

def test_build_derived_input_stage_and_computed_axis(
    tmp_db, monkeypatch: pytest.MonkeyPatch  # noqa: ANN001
) -> None:
    import json

    from openrater.rates.ingest.service import build_workbook

    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    outcome = build_workbook(
        db=tmp_db, data=to_bytes(build_derived()), filename="derived.xlsx"
    )
    plan_id = outcome.rating_plan_id

    conn = tmp_db.connection()
    stages = {
        json.loads(r["config_json"]).get("name"): json.loads(r["config_json"])
        for r in conn.execute(
            "SELECT config_json FROM rating_plan_stages "
            "WHERE rating_plan_id = ? AND stage_kind = 'input_node'",
            (plan_id,),
        )
    }
    chain = json.loads(
        conn.execute(
            "SELECT config_json FROM rating_plan_stages "
            "WHERE rating_plan_id = ? AND stage_kind = 'multiplicative_chain'",
            (plan_id,),
        ).fetchone()["config_json"]
    )
    conn.close()

    derived = stages["total_limit"]
    assert derived["source"] == "derived"
    assert derived["required"] is False
    assert derived["derived_rule"] == "sum(tiv,bpp_limit)"
    assert derived["derived_expr"] == {
        "kind": "op",
        "op": "+",
        "left": {"kind": "input", "name": "tiv"},
        "right": {"kind": "input", "name": "bpp_limit"},
    }
    assert stages["bpp_limit"]["source"] == "form"

    tower = chain["chains"][0]
    lookups = [
        f for f in tower["factor_lookups"] if f.get("factor_kind") == "limit_band"
    ]
    assert lookups, tower["factor_lookups"]
    binding = lookups[0]["dimensions"]["limit_band"]
    assert binding == {
        "source": "computed", "op": "sum", "fields": ["tiv", "bpp_limit"],
    }

    # D2 — the report carries the case INPUTS (operands only).
    cases = outcome.report.vectors.cases
    assert [c.case_id for c in cases] == ["tc_1"]
    assert cases[0].inputs["tiv"] == 200000
    assert cases[0].inputs["bpp_limit"] == 100000
    assert "total_limit" not in cases[0].inputs


def test_pre_95_4_vectors_summary_parses_without_cases() -> None:
    old = VectorsSummary.model_validate_json(
        '{"status": "ran", "total_cases": 1, "matched": 1}'
    )
    assert old.cases == []


# ---------------------------------------------------------------------------
# D1 — the book-template CSV endpoint.
# ---------------------------------------------------------------------------

def test_book_template_csv(client, monkeypatch: pytest.MonkeyPatch) -> None:  # noqa: ANN001
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    resp = client.post(
        "/api/v1/plans/ingest?filename=derived.xlsx",
        content=to_bytes(build_derived()),
    )
    assert resp.status_code == 200, resp.text
    plan_id = resp.json()["rating_plan_id"]

    csv_resp = client.get(f"/api/v1/plans/{plan_id}/book-template.csv")
    assert csv_resp.status_code == 200
    assert "attachment" in csv_resp.headers["content-disposition"]
    assert plan_id in csv_resp.headers["content-disposition"]

    rows = list(csv.reader(io.StringIO(csv_resp.text)))
    header, example = rows[0], rows[1]
    # Declared inputs in sheet order; the derived input is NOT a column.
    assert header == ["tiv", "construction_class", "building_age", "bpp_limit"]
    assert "total_limit" not in header
    # The example row is the filing's own first test case.
    assert example[header.index("tiv")] == "200000"
    assert example[header.index("bpp_limit")] == "100000"
    assert example[header.index("construction_class")] == "frame"


def test_book_template_unknown_plan_404(client) -> None:  # noqa: ANN001
    resp = client.get("/api/v1/plans/nope/book-template.csv")
    assert resp.status_code == 404
    assert resp.json()["error"]["code"] == "plan_not_found"


def test_book_template_plan_without_inputs_422(client) -> None:  # noqa: ANN001
    created = client.post(
        "/api/v1/plans",
        json={"display_name": "Bare plan", "product": "bop"},
    )
    assert created.status_code in (200, 201), created.text
    plan_id = created.json()["rating_plan_id"]
    resp = client.get(f"/api/v1/plans/{plan_id}/book-template.csv")
    assert resp.status_code == 422
    assert resp.json()["error"]["code"] == "book_template_no_inputs"
