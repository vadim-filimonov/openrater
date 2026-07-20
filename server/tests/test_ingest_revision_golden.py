# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""Brief 92.R phase 92R.5 — the revised-nonprofit acceptance golden.

The whole revision loop, exercised on the canonical example bundle:

  base    = the canonical nonprofit workbook + one endorsement whose
            trigger (revenue >= $10M; the examples top out at $3.2M)
            never fires — so the base still verifies 80/80.
  revised = v1.1.0: +5% on three factors (ONE of which — ntee_do
            religion — is traversed by exactly the two religion cases),
            one added NTEE level with its cells in both NTEE tables,
            the endorsement removed, and the two affected cases'
            expectations revised by the filing itself.

Structural tests always run (the diff engine is pure — no scoring).
The live golden (RATER_INGEST_LIVE_SCORING=1) builds the base 80/80,
applies the revision, and asserts 80/80 on the NEW expectations plus
the drift pins — all through the production engine.

Every revised expectation below is re-derived from the filed math
(the same derivation generate_workbook.py self-verifies against the
source CSVs), not scaled from rounded values:

  NP-001 Faith Community Church   do 658 -> 691   total 1054 -> 1087
  NP-020 Faith Youth Mission      do 883 -> 927   total 1235 -> 1279

Drift (engine actuals): both do towers move exactly +5.0%; the
round-once package totals move +3.1% (1054->1087) and +3.6%
(1235->1279). pcts = [5.0, 3.1, 5.0, 3.6] -> median 4.3, max 5.0.
"""

from __future__ import annotations

import io
import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from test_ingest_build import _fake_score, tmp_db  # noqa: E402, F401
from test_ingest_check import CANONICAL, find_row  # noqa: E402

from openrater.rates.ingest.diff import diff_workbooks  # noqa: E402
from openrater.rates.ingest.parser import parse_workbook  # noqa: E402
from openrater.rates.ingest.service import (  # noqa: E402
    build_workbook,
    check_workbook,
    reingest_apply,
    reingest_check,
)

CITE = "Nonprofit 990 D&O+GL rating v1 (source workbook)"


def _bytes(wb) -> bytes:  # noqa: ANN001
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_base_bytes() -> bytes:
    """The canonical bundle + one never-firing endorsement (so the
    revision has something real to remove; premiums are untouched
    because no example's revenue reaches the $10M trigger)."""
    wb = load_workbook(io.BytesIO(CANONICAL.read_bytes()))
    ws = wb.create_sheet("endorsements")
    ws.append(["endorsement_id", "kind", "form_number", "display_name",
               "factor", "amount", "coverage", "sublimit", "trigger",
               "citation_rule", "citation_page"])
    ws.append(["cyber_liab_ext", "factor", "NP 07 88",
               "Cyber liability extension", 1.25, "", "", "",
               "form_input.revenue >= 10000000", CITE, "Endorsements"])
    return _bytes(wb)


def _bump(ws, level_id: str, pct: float = 1.05) -> None:  # noqa: ANN001
    """+5% a 1-D factor grid row (level ids live in column 1)."""
    row = find_row(ws, 1, level_id)
    cell = ws.cell(row=row, column=2)
    cell.value = round(cell.value * pct, 4)


def build_revision_bytes() -> bytes:
    """v1.1.0 — the scripted revision the brief's phase table names."""
    wb = load_workbook(io.BytesIO(build_base_bytes()))
    wb.remove(wb["endorsements"])

    ws = wb["plan"]
    ws.cell(row=find_row(ws, 1, "version"), column=2).value = "1.1.0"

    # +5% on three factors. Only religion is traversed by any example
    # (exactly two: np_001 and np_020) — the other two are quiet.
    _bump(wb["ft.ntee_do"], "religion")          # 1.20 -> 1.26
    _bump(wb["ft.ntee_gl"], "public_societal")   # 1.00 -> 1.05
    _bump(wb["ft.state_do"], "az")               # 1.00 -> 1.05

    # One added level + its cells in both tables that key the dimension.
    wb["dimension_levels"].append(
        ["ntee_major", "categorical", "digital_media",
         "Digital Media & Online Communities", "", "", "", "",
         CITE, "Dim - NTEE Major"])
    wb["ft.ntee_do"].append(["digital_media", 1.1, CITE, "Dim - NTEE Major"])
    wb["ft.ntee_gl"].append(["digital_media", 1.1, CITE, "Dim - NTEE Major"])

    # The filing revises the two affected expectations itself.
    tc = wb["test_cases"]
    for case_id, new_do, new_total in (
        ("np_001", 691, 1087),
        ("np_020", 927, 1279),
    ):
        row = find_row(tc, 1, case_id)
        tc.cell(row=row, column=10).value = new_do      # expected_do_premium
        tc.cell(row=row, column=12).value = new_total   # expected_total_premium
    return _bytes(wb)


# ---------------------------------------------------------------------------
# Structural — the diff engine is pure; these always run.
# ---------------------------------------------------------------------------


def _parsed(data: bytes):  # noqa: ANN202
    parsed, issues = parse_workbook(data)
    assert not [i for i in issues if i.severity == "error"], issues
    return parsed


def test_base_and_revision_both_check_clean() -> None:
    for data in (build_base_bytes(), build_revision_bytes()):
        result = check_workbook(data=data, filename="golden.xlsx")
        assert result.ok, [e.message for e in result.errors]


def test_revision_diff_reads_exactly() -> None:
    diff = diff_workbooks(_parsed(build_base_bytes()),
                          _parsed(build_revision_bytes()))

    assert diff.totals.added == 1
    assert diff.totals.changed == 6  # plan + 3 tables + 2 expectations
    assert diff.totals.removed == 1
    assert diff.totals.sections_changed == 5

    by_section = {s.section: s for s in diff.sections}

    plan = by_section["plan"]
    assert (plan.added, plan.changed, plan.removed) == (0, 1, 0)
    [version] = plan.items[0].changes or []
    assert version.field == "version"
    assert (version.from_, version.to) == ("1.0.0", "1.1.0")

    dims = by_section["dimensions"]
    assert (dims.added, dims.changed, dims.removed) == (1, 0, 0)
    assert "digital_media" in dims.items[0].summary

    fts = by_section["factor_tables"]
    assert (fts.added, fts.changed, fts.removed) == (0, 3, 0)
    by_key = {i.key: i for i in fts.items}
    assert set(by_key) == {"ft.ntee_do", "ft.ntee_gl", "ft.state_do"}

    ntee_do = {c.field: c for c in by_key["ft.ntee_do"].changes or []}
    assert set(ntee_do) == {"religion", "digital_media"}
    assert ntee_do["religion"].pct == 5.0
    assert (ntee_do["religion"].from_, ntee_do["religion"].to) == (1.2, 1.26)
    assert ntee_do["digital_media"].from_ is None
    assert ntee_do["digital_media"].to == 1.1

    endo = by_section["endorsements"]
    assert (endo.added, endo.changed, endo.removed) == (0, 0, 1)
    [removed] = endo.items
    assert removed.state == "removed"
    assert "cyber_liab_ext" in removed.summary
    assert "NP 07 88" in removed.summary
    assert "trigger form_input.revenue >= 10000000" in removed.summary

    tcs = by_section["test_cases"]
    assert (tcs.added, tcs.changed, tcs.removed) == (0, 2, 0)
    assert {i.key for i in tcs.items} == {"np_001", "np_020"}
    np_001 = {c.field: c for c in
              next(i for i in tcs.items if i.key == "np_001").changes or []}
    assert set(np_001) == {"expected_do_premium", "expected_total_premium"}
    assert (np_001["expected_do_premium"].from_,
            np_001["expected_do_premium"].to) == (658, 691)
    assert np_001["expected_do_premium"].pct == 5.0


def test_revision_diff_is_deterministic() -> None:
    """CT-8 — the same base+new pair, byte-identical diff JSON, twice."""
    base, new = build_base_bytes(), build_revision_bytes()
    one = diff_workbooks(_parsed(base), _parsed(new))
    two = diff_workbooks(_parsed(base), _parsed(new))
    assert one.model_dump_json(by_alias=True) == two.model_dump_json(by_alias=True)


def test_apply_lands_the_revision_on_the_substrate(
    client, monkeypatch  # noqa: ANN001
) -> None:
    """Fake-scored end-to-end: after apply the SAME plan carries the
    bumped cell, the new level, and no endorsement stage."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"do_premium": 0.0, "gl_premium": 0.0, "total_premium": 0.0}),
    )
    resp = client.post(
        "/api/v1/plans/ingest?filename=base.xlsx", content=build_base_bytes()
    )
    assert resp.status_code == 200, resp.text
    plan_id = resp.json()["rating_plan_id"]

    resp = client.post(
        f"/api/v1/plans/{plan_id}/reingest?filename=rev.xlsx",
        content=build_revision_bytes(),
    )
    assert resp.status_code == 200, resp.text
    report = resp.json()["report"]
    assert report["workbook_version"] == "1.1.0"
    assert report["diff"]["totals"] == {
        "added": 1, "changed": 6, "removed": 1, "sections_changed": 5,
    }
    assert report["drift"]["expectations_revised"] == 2

    conn = client.app.state.db.connection()
    religion = conn.execute(
        "SELECT value FROM plan_factor_table_cells WHERE rating_plan_id = ?"
        " AND table_id = 'ntee_do' AND cell_key = 'religion'",
        (plan_id,),
    ).fetchone()
    assert religion["value"] == 1.26
    added = conn.execute(
        "SELECT value FROM plan_factor_table_cells WHERE rating_plan_id = ?"
        " AND table_id = 'ntee_gl' AND cell_key = 'digital_media'",
        (plan_id,),
    ).fetchone()
    assert added["value"] == 1.1
    stages = conn.execute(
        "SELECT stage_id FROM rating_plan_stages WHERE rating_plan_id = ?"
        " AND stage_id LIKE 'endorsement_%'",
        (plan_id,),
    ).fetchall()
    assert stages == []


# ---------------------------------------------------------------------------
# The LIVE golden (env-gated; the contract job boots both services).
# ---------------------------------------------------------------------------


@pytest.mark.skipif(
    os.environ.get("RATER_INGEST_LIVE_SCORING") != "1",
    reason="needs a running scoring service (set RATER_INGEST_LIVE_SCORING=1)",
)
def test_revised_nonprofit_golden_live(tmp_db) -> None:  # noqa: ANN001, F811
    """92R.5 — the acceptance golden: the base verifies 80/80, the
    revision applies onto the SAME plan, the NEW expectations verify
    80/80 through the production engine, and the drift is the filed
    move exactly."""
    outcome = build_workbook(
        db=tmp_db, data=build_base_bytes(), filename="nonprofit_base.xlsx"
    )
    plan_id = outcome.rating_plan_id
    v = outcome.report.vectors
    assert v.status == "ran", v.detail
    assert v.matched == len(v.checks) == 80, [
        (c.case_id, c.field, c.expected, c.actual, c.status)
        for c in v.checks if c.status != "match"
    ]

    checked = reingest_check(
        db=tmp_db, rating_plan_id=plan_id,
        data=build_revision_bytes(), filename="nonprofit_v1_1.xlsx",
    )
    assert checked.check.ok
    assert checked.diff is not None
    assert checked.diff.totals.changed == 6
    assert not checked.hand_edited_since_build

    applied = reingest_apply(
        db=tmp_db, rating_plan_id=plan_id,
        data=build_revision_bytes(), filename="nonprofit_v1_1.xlsx",
        if_match=checked.plan_content_hash,
    )
    assert applied.rating_plan_id == plan_id

    v = applied.report.vectors
    assert v.status == "ran", v.detail
    assert v.mismatched == 0 and v.near == 0, [
        (c.case_id, c.field, c.expected, c.actual, c.status)
        for c in v.checks if c.status != "match"
    ]
    assert v.matched == len(v.checks) == 80

    drift = applied.report.drift
    assert drift is not None
    assert drift.compared == 80
    assert drift.expectations_revised == 2
    moved = {(c.case_id, c.field): c.pct for c in drift.cases}
    assert moved == {
        ("np_001", "do_premium"): 5.0,
        ("np_001", "total_premium"): 3.1,
        ("np_020", "do_premium"): 5.0,
        ("np_020", "total_premium"): 3.6,
    }
    assert drift.median_pct == 4.3
    assert drift.max_pct == 5.0
