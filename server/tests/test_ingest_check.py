# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""Brief 92 Phase 92.2 — the workbook check.

Three layers:

  1. The conformant MINI workbook (the spec §9 example, programmatic)
     passes clean, and every `R-###` rule fires from a targeted
     mutation of it — one parametrized case per rule id.
  2. The canonical example bundle
     (docs/specs/examples/nonprofit-do-gl/) checks green with the
     expected manifest counts — the golden path 92.3's build test
     will extend.
  3. The endpoint + CLI wrap the same service (raw-body POST; exit
     codes), and the packaged capability registry is byte-identical
     to the docs copy (the no-drift guard the spec §6 promises).
"""

from __future__ import annotations

import io
import json
from collections.abc import Callable
from importlib import resources
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from openrater.rates.ingest import check_workbook
from openrater.rates.ingest.__main__ import main as cli_main

REPO_ROOT = Path(__file__).resolve().parents[2]
CANONICAL = REPO_ROOT / "docs" / "specs" / "examples" / "nonprofit-do-gl" / "nonprofit_do_gl.workbook.xlsx"
DOCS_REGISTRY = REPO_ROOT / "docs" / "specs" / "transcription-capability-registry.json"


# ---------------------------------------------------------------------------
# The conformant mini workbook (spec §9, programmatic).
# ---------------------------------------------------------------------------

def build_mini() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "plan"
    for row in (
        ("field", "value"),
        ("spec_version", "1.0"),
        ("rating_plan_id", "mini-bop-demo-il-2026"),
        ("display_name", "Mini BOP demo"),
        ("version", "1.0.0"),
        ("carrier", "Demo Mutual"),
        ("product", "bop"),
        ("jurisdiction_country", "US"),
        ("state", "IL"),
        ("effective_date", "2026-01-01"),
        ("coverages", "building"),
    ):
        ws.append(list(row))

    ws = wb.create_sheet("inputs")
    ws.append(["name", "label", "data_type", "required", "allowed_values", "default_value", "unit"])
    ws.append(["tiv", "Total insured value", "currency", True, "", "", "USD"])
    ws.append(["construction_class", "Construction class", "enum", True, "frame,fire_resistive", "", ""])
    ws.append(["building_age", "Building age", "number", True, "", "", "years"])

    ws = wb.create_sheet("dimensions")
    ws.append(["slug", "display_name", "shape", "role", "data_type", "dimension_type",
               "geo_granularity", "geo_scope", "axes"])
    ws.append(["construction_class", "Construction class", "categorical", "both", "enum", "standard", "", "", ""])
    ws.append(["building_age", "Building age (yrs)", "banded", "rating-input", "number", "standard", "", "", ""])

    ws = wb.create_sheet("dimension_levels")
    ws.append(["dimension_slug", "kind", "level_id", "label", "aliases", "min", "max", "territory_ref"])
    ws.append(["construction_class", "categorical", "frame", "Frame (type c1)", "wood,c1", "", "", ""])
    ws.append(["construction_class", "categorical", "fire_resistive", "Fire-resistive (type c2)", "fr,c2", "", "", ""])
    ws.append(["building_age", "banded", "age_0_25", "0-25 yrs", "", 0, 25, ""])
    ws.append(["building_age", "banded", "age_25_plus", "25+ yrs", "", 25, "+inf", ""])

    ws = wb.create_sheet("ft.construction_class")
    for row in (
        ("table_id", "construction_class"),
        ("display_name", "Construction factor"),
        ("dimensionality", "1d"),
        ("row_dimension", "construction_class"),
        ("lookup_method", "direct"),
        ("citation_rule", "Table 5.A"),
        ("citation_page", "p.51"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["level_id", "factor", "citation_rule", "citation_page"])
    ws.append(["frame", 1.00, "Table 5.A", "p.51"])
    ws.append(["fire_resistive", 0.78, "Table 5.A", "p.51"])

    ws = wb.create_sheet("ft.constr_x_age")
    for row in (
        ("table_id", "constr_x_age"),
        ("display_name", "Construction x Age factor"),
        ("dimensionality", "2d"),
        ("row_dimension", "construction_class"),
        ("col_dimension", "building_age"),
        ("lookup_method", "direct"),
        ("citation_rule", "Table 5.B"),
        ("citation_page", "p.52"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["row\\col", "age_0_25", "age_25_plus"])
    ws.append(["frame", 1.00, 1.20])
    ws.append(["fire_resistive", 0.95, 1.05])

    ws = wb.create_sheet("chains")
    ws.append(["coverage", "order", "stage_kind", "stage_id", "factor_table", "dimension",
               "input_binding", "value", "exposure_divisor", "predicate"])
    ws.append(["building", 0, "base", "bld_base", "", "", "literal:0.150", 0.150, "", ""])
    ws.append(["building", 1, "lookup.direct", "bld_constr", "ft.construction_class",
               "construction_class", "", "", "", ""])
    ws.append(["building", 2, "lookup.multi", "bld_constr_age", "ft.constr_x_age",
               "construction_class", "", "", "", ""])
    ws.append(["building", 3, "exposure", "bld_exposure", "", "", "form_input.tiv", "", 100, ""])
    ws.append(["building", 4, "lcm", "bld_lcm", "", "", "", 1.30, "", ""])

    ws = wb.create_sheet("outputs")
    ws.append(["output_id", "field_name", "display_name", "source"])
    ws.append(["out_building", "building_premium", "Building premium", "bld_exposure"])

    ws = wb.create_sheet("test_cases")
    ws.append(["case_id", "name", "construction_class", "building_age", "tiv",
               "expected_building_premium"])
    ws.append(["tc_1", "Frame, 10 yrs, $200k", "frame", 10, 200000, 390.00])

    ws = wb.create_sheet("gaps_and_assumptions")
    ws.append(["kind", "description", "citation_rule", "citation_page", "impact", "related"])

    return wb


def to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_check(wb: Workbook):
    return check_workbook(to_bytes(wb), filename="mini.xlsx")


def find_row(ws: Worksheet, col: int, value) -> int:
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == value:
            return r
    raise AssertionError(f"{value!r} not found in {ws.title} col {col}")


def rules_of(result) -> set[str]:
    return {i.rule for i in (result.errors + result.warnings + result.notices)}


# ---------------------------------------------------------------------------
# 1a. The mini workbook is conformant.
# ---------------------------------------------------------------------------

def test_mini_workbook_checks_clean() -> None:
    result = run_check(build_mini())
    assert result.errors == []
    assert result.ok is True
    assert result.manifest is not None
    c = result.manifest.counts
    assert (c.dimensions, c.factor_tables, c.factor_cells) == (2, 2, 6)
    assert (c.chains, c.chain_stages, c.outputs, c.test_cases) == (1, 5, 1, 1)
    assert (c.inputs, c.inputs_with_defaults, c.declared_gaps) == (3, 0, 0)
    assert result.manifest.provenance.carrier == "Demo Mutual"
    assert result.workbook_hash and len(result.workbook_hash) == 64


# ---------------------------------------------------------------------------
# 1b. Every rule fires — one mutation per rule id.
# ---------------------------------------------------------------------------

Mutator = Callable[[Workbook], None]


def _add_valid_gates(wb: Workbook) -> None:
    ws = wb.create_sheet("gates")
    ws.append(["order", "rule_id", "variable", "op", "value",
               "variable_2", "op_2", "value_2", "variable_3", "op_3", "value_3",
               "tier", "reasoning"])
    ws.append([1, "big_tiv", "tiv", "gt", 5000000, "", "", "", "", "", "",
               "submit", "Large schedule"])
    ws.append([99, "__default__", "", "", "", "", "", "", "", "", "",
               "standard", "Standard appetite"])


def _add_valid_modifiers(wb: Workbook) -> None:
    ws = wb.create_sheet("modifiers")
    ws.append(["schedule_id", "schedule_name", "scope", "total_cap_pct",
               "category_id", "category_name", "range_pct", "tier_filter"])
    ws.append(["irpm", "IRPM", "package", 25, "management", "Management", 10, ""])


def _add_valid_endorsements(wb: Workbook) -> None:
    ws = wb.create_sheet("endorsements")
    ws.append(["endorsement_id", "kind", "form_number", "display_name",
               "factor", "amount", "coverage", "sublimit", "trigger"])
    ws.append(["spoilage", "factor", "MS 10 13", "Spoilage", 1.05, "", "", "", ""])


def _add_valid_loadings(wb: Workbook) -> None:
    ws = wb.create_sheet("loadings")
    ws.append(["loading_id", "factor_kind", "display_name", "factor", "applies_to", "predicate"])
    ws.append(["expense", "expense", "Expense load", 1.18, "", ""])


def _add_valid_adjustments(wb: Workbook) -> None:
    ws = wb.create_sheet("final_adjustments")
    ws.append(["adjustment_id", "kind", "order", "applies_to", "min_value", "max_value",
               "round_increment", "round_min"])
    ws.append(["min_premium", "clamp", 1, "", 500, "", "", ""])


def _plan_field_row(wb: Workbook, field: str) -> int:
    return find_row(wb["plan"], 1, field)


def _set_interpolation_linear(wb: Workbook) -> None:
    ws = wb["ft.construction_class"]
    row = find_row(ws, 1, "lookup_method")
    ws.insert_rows(row + 1)
    ws.cell(row=row + 1, column=1, value="interpolation")
    ws.cell(row=row + 1, column=2, value="linear")


# (rule, expected severity bucket, mutator)
CASES: list[tuple[str, str, Mutator]] = [
    ("R-003", "errors", lambda wb: wb["inputs"].cell(row=1, column=1, value="nom")),
    ("R-004", "errors", lambda wb: wb["chains"].cell(row=2, column=8, value="=6*100")),
    ("R-005", "errors", lambda wb: wb["chains"].cell(row=2, column=8, value="$0.150")),
    ("R-006", "errors", lambda wb: wb["dimensions"].cell(row=2, column=1, value="Bad Slug")),
    ("R-008", "errors", lambda wb: wb["dimension_levels"].cell(row=2, column=3, value="__default__")),
    ("R-020", "errors", lambda wb: wb.remove(wb["inputs"])),
    ("R-021", "errors", lambda wb: wb.remove(wb["plan"])),
    ("R-022", "errors", lambda wb: wb.remove(wb["dimensions"])),
    ("R-023", "errors", lambda wb: wb.remove(wb["chains"])),
    ("R-024", "errors", lambda wb: wb.remove(wb["outputs"])),
    ("R-025", "errors", lambda wb: wb.remove(wb["test_cases"])),
    ("R-026", "errors", lambda wb: wb.remove(wb["gaps_and_assumptions"])),
    ("R-027", "errors", lambda wb: wb["plan"].delete_rows(_plan_field_row(wb, "carrier"))),
    ("R-028", "errors", lambda wb: wb["plan"].cell(row=_plan_field_row(wb, "product"), column=2, value="boat")),
    ("R-029", "errors", lambda wb: wb["plan"].cell(row=_plan_field_row(wb, "state"), column=2, value="Illinois")),
    ("R-030", "errors", lambda wb: wb["plan"].cell(row=_plan_field_row(wb, "effective_date"), column=2, value="July 1, 2026")),
    ("R-031", "errors", lambda wb: wb["plan"].cell(row=_plan_field_row(wb, "coverages"), column=2, value=" , ")),
    ("R-032", "errors", lambda wb: wb["plan"].cell(row=_plan_field_row(wb, "spec_version"), column=2, value="0.9")),
    ("R-040", "errors", lambda wb: wb["inputs"].append(["tiv", "Dup", "currency", True, "", "", ""])),
    ("R-041", "errors", lambda wb: wb["inputs"].cell(row=2, column=3, value="text")),
    ("R-042", "errors", lambda wb: wb["inputs"].cell(row=3, column=5, value="")),
    ("R-043", "errors", lambda wb: wb["inputs"].cell(row=3, column=6, value="brick")),
    ("R-044", "errors", lambda wb: (
        wb["test_cases"].cell(row=1, column=7, value="mystery_field"),
        wb["test_cases"].cell(row=2, column=7, value=42),
    )),
    ("R-060", "errors", lambda wb: wb["dimensions"].append(
        ["construction_class", "Dup", "categorical", "both", "enum", "standard", "", "", ""])),
    ("R-061", "errors", lambda wb: wb["dimensions"].cell(row=2, column=3, value="circle")),
    ("R-062", "errors", lambda wb: wb["dimension_levels"].cell(
        row=find_row(wb["dimension_levels"], 3, "age_25_plus"), column=6, value=20)),
    ("R-063", "errors", lambda wb: (
        wb["dimension_levels"].cell(
            row=find_row(wb["dimension_levels"], 3, "age_0_25"), column=7, value="+inf"),
    )),
    ("R-064", "errors", lambda wb: wb["dimensions"].append(
        ["state", "State", "geographic", "rating-input", "string", "geographic", "", "", ""])),
    ("R-065", "errors", lambda wb: wb["dimensions"].append(
        ["combo", "Combo", "composite", "structural", "string", "standard", "", "", "construction_class"])),
    ("R-066", "errors", lambda wb: wb["dimensions"].append(
        ["dep_band", "Deductible band", "banded", "rating-input", "number", "standard", "", "", ""])),
    ("R-067", "errors", lambda wb: wb["dimension_levels"].append(
        ["construction_class", "categorical", "frame", "Dup frame", "", "", "", ""])),
    ("R-068", "errors", lambda wb: wb["dimension_levels"].append(
        ["ghost", "categorical", "phantom", "Phantom", "", "", "", ""])),
    ("R-069", "errors", lambda wb: wb["dimension_levels"].cell(
        row=find_row(wb["dimension_levels"], 3, "age_0_25"), column=7, value=0)),
    ("R-100", "errors", lambda wb: wb["ft.construction_class"].delete_rows(
        find_row(wb["ft.construction_class"], 1, "row_dimension"))),
    ("R-101", "errors", lambda wb: wb["ft.construction_class"].cell(
        row=find_row(wb["ft.construction_class"], 1, "table_id"), column=2, value="wrong")),
    ("R-102", "errors", lambda wb: wb["ft.construction_class"].cell(
        row=find_row(wb["ft.construction_class"], 1, "dimensionality"), column=2, value="3d")),
    ("R-103", "errors", lambda wb: wb["ft.construction_class"].cell(
        row=find_row(wb["ft.construction_class"], 1, "row_dimension"), column=2, value="ghost")),
    ("R-104", "errors", lambda wb: wb["ft.construction_class"].append(["brick", 1.10, "", ""])),
    ("R-105", "errors", lambda wb: wb["ft.constr_x_age"].cell(
        row=find_row(wb["ft.constr_x_age"], 1, "row\\col"), column=2, value="age_bad")),
    ("R-106", "errors", lambda wb: wb["ft.construction_class"].append(["frame", 1.05, "", ""])),
    ("R-107", "errors", lambda wb: wb["ft.construction_class"].cell(
        row=find_row(wb["ft.construction_class"], 1, "frame"), column=2, value=-1)),
    ("R-108", "errors", lambda wb: (
        wb["ft.construction_class"].append(["__default__", 1.0, "", ""]),
        wb["ft.construction_class"].append(["__default__", 1.1, "", ""]),
    )),
    ("R-109", "errors", lambda wb: wb["ft.construction_class"].delete_rows(
        find_row(wb["ft.construction_class"], 1, "level_id"), 3)),
    # Brief 95 C5 (registry r9) — interpolation is supported; R-111 is a
    # notice naming the interpolating table, no longer a caveat.
    ("R-111", "notices", lambda wb: _set_interpolation_linear(wb)),
    ("R-120", "errors", lambda wb: wb["chains"].cell(row=2, column=1, value="boat")),
    ("R-121", "errors", lambda wb: wb["chains"].cell(row=3, column=5, value="ft.sprinkler")),
    ("R-122", "errors", lambda wb: wb["chains"].cell(row=3, column=6, value="ghost")),
    ("R-123", "errors", lambda wb: wb["chains"].cell(row=2, column=2, value=9)),
    ("R-124", "errors", lambda wb: (
        wb["chains"].cell(row=2, column=7, value=""),
        wb["chains"].cell(row=2, column=8, value=""),
    )),
    ("R-125", "errors", lambda wb: wb["chains"].cell(row=3, column=3, value="multiply")),
    ("R-126", "errors", lambda wb: wb["chains"].cell(row=3, column=4, value="bld_base")),
    # A bare field name in a binding cell (the form_input. prefix is the
    # grammar) — the classic transcriber slip R-127 exists to catch.
    ("R-127", "errors", lambda wb: wb["chains"].cell(row=5, column=7, value="tiv")),
    ("R-128", "errors", lambda wb: wb["chains"].cell(row=3, column=10, value="tiv >>> 5")),
    ("R-140", "errors", lambda wb: wb["test_cases"].cell(row=2, column=6, value="three ninety")),
    ("R-141", "errors", lambda wb: wb["outputs"].cell(row=2, column=4, value="ghost_stage")),
    ("R-142", "errors", lambda wb: wb["outputs"].append(
        ["out_building", "building_premium", "Dup", "bld_exposure"])),
    ("R-143", "errors", lambda wb: wb["test_cases"].cell(row=2, column=5, value="")),
    ("R-144", "errors", lambda wb: (
        wb["test_cases"].cell(row=1, column=7, value="expected_ghost"),
        wb["test_cases"].cell(row=2, column=7, value=1),
    )),
    ("R-145", "errors", lambda wb: wb["test_cases"].delete_rows(2)),
    ("R-146", "errors", lambda wb: wb["outputs"].cell(row=2, column=4, value="coverage:total")),
    ("R-160", "errors", lambda wb: (_add_valid_gates(wb), wb["gates"].delete_rows(3))),
    ("R-161", "errors", lambda wb: (_add_valid_gates(wb), wb["gates"].cell(row=2, column=4, value="equals"))),
    ("R-162", "errors", lambda wb: (_add_valid_gates(wb), wb["gates"].cell(row=2, column=12, value="gold"))),
    ("R-163", "errors", lambda wb: (_add_valid_gates(wb), wb["gates"].append(
        [2, "big_tiv", "tiv", "gt", 1, "", "", "", "", "", "", "submit", "Dup"]))),
    ("R-164", "errors", lambda wb: (_add_valid_modifiers(wb), wb["modifiers"].append(
        ["irpm", "IRPM", "package", 25, "management", "Dup category", 5, ""]))),
    ("R-165", "errors", lambda wb: (_add_valid_modifiers(wb), wb["modifiers"].cell(row=2, column=7, value=-5))),
    ("R-166", "errors", lambda wb: (_add_valid_endorsements(wb), wb["endorsements"].cell(row=2, column=5, value=""))),
    ("R-167", "errors", lambda wb: (_add_valid_loadings(wb), wb["loadings"].cell(row=2, column=5, value="boat"))),
    ("R-168", "errors", lambda wb: (_add_valid_adjustments(wb), wb["final_adjustments"].cell(row=2, column=5, value=""))),
    ("R-169", "errors", lambda wb: (_add_valid_gates(wb), wb["gates"].cell(row=2, column=6, value="tiv"))),
    ("R-170", "errors", lambda wb: (_add_valid_adjustments(wb), wb["final_adjustments"].cell(row=2, column=4, value="boat"))),
    ("R-180", "errors", lambda wb: wb["gaps_and_assumptions"].append(
        ["wish", "Something", "", "", "Impact", ""])),
    ("R-201", "warnings", lambda wb: (
        wb["ft.construction_class"].cell(
            row=find_row(wb["ft.construction_class"], 1, "citation_rule"), column=1, value="was_citation"),
        wb["ft.construction_class"].cell(
            row=find_row(wb["ft.construction_class"], 1, "frame"), column=3, value=""),
        wb["ft.construction_class"].cell(
            row=find_row(wb["ft.construction_class"], 1, "fire_resistive"), column=3, value=""),
    )),
    ("R-202", "warnings", lambda wb: wb["inputs"].cell(row=1, column=8, value="wibble")),
    ("R-203", "notices", lambda wb: wb.create_sheet("Notes").append(["prose"])),
]


@pytest.mark.parametrize(("rule", "bucket", "mutate"), CASES, ids=[c[0] for c in CASES])
def test_rule_fires(rule: str, bucket: str, mutate: Mutator) -> None:
    wb = build_mini()
    mutate(wb)
    result = run_check(wb)
    hits = {i.rule for i in getattr(result, bucket)}
    assert rule in hits, (
        f"{rule} did not fire; errors={[(i.rule, i.message) for i in result.errors]} "
        f"warnings={[(i.rule, i.message) for i in result.warnings]}"
    )
    if bucket == "errors":
        assert result.ok is False and result.manifest is None
    else:
        assert result.ok is True


def test_unreadable_bytes_is_r001() -> None:
    result = check_workbook(b"definitely not a workbook")
    assert [i.rule for i in result.errors] == ["R-001"]
    assert result.ok is False and result.manifest is None


def test_r127_binding_grammar_shapes() -> None:
    """The §4.6 input_binding grammar, branch by branch. The clean
    shapes MUST pass — they are exactly what the builder resolves and
    the projector executes (spec §4.6 / R-127)."""
    # context.lcm on a non-lcm row is meaningless.
    wb = build_mini()
    wb["chains"].cell(row=5, column=7, value="context.lcm")
    assert "R-127" in {i.rule for i in run_check(wb).errors}

    # context.lcm on the lcm row of a plan with NO plan-sheet lcm —
    # nothing to resolve from.
    wb = build_mini()
    wb["chains"].cell(row=6, column=7, value="context.lcm")
    wb["chains"].cell(row=6, column=8, value="")
    assert "R-127" in {i.rule for i in run_check(wb).errors}

    # ...and a TEXT plan-sheet lcm would crash the builder's float().
    wb = build_mini()
    wb["plan"].append(["lcm", "one point three"])
    wb["chains"].cell(row=6, column=7, value="context.lcm")
    wb["chains"].cell(row=6, column=8, value="")
    assert "R-005" in {i.rule for i in run_check(wb).errors}

    # A malformed literal fires even when the value cell wins.
    wb = build_mini()
    wb["chains"].cell(row=6, column=7, value="literal:abc")
    assert "R-127" in {i.rule for i in run_check(wb).errors}

    # Clean: context.lcm on the lcm row with a numeric plan-sheet lcm.
    wb = build_mini()
    wb["plan"].append(["lcm", 1.30])
    wb["chains"].cell(row=6, column=7, value="context.lcm")
    wb["chains"].cell(row=6, column=8, value="")
    result = run_check(wb)
    assert result.ok is True, [(i.rule, i.message) for i in result.errors]

    # Clean: a binding-only base (R-124 allows value OR binding) and a
    # literal exposure — the forms the builder/projector now execute.
    wb = build_mini()
    wb["chains"].cell(row=2, column=8, value="")  # base keeps literal:0.150
    wb["chains"].cell(row=5, column=7, value="literal:250")
    result = run_check(wb)
    assert result.ok is True, [(i.rule, i.message) for i in result.errors]


def test_expected_tier_requires_gates_sheet() -> None:
    wb = build_mini()
    wb["test_cases"].cell(row=1, column=7, value="expected_tier")
    wb["test_cases"].cell(row=2, column=7, value="standard")
    result = run_check(wb)
    assert "R-144" in {i.rule for i in result.errors}

    wb = build_mini()
    _add_valid_gates(wb)
    wb["test_cases"].cell(row=1, column=7, value="expected_tier")
    wb["test_cases"].cell(row=2, column=7, value="standard")
    result = run_check(wb)
    assert result.ok is True, [(i.rule, i.message) for i in result.errors]


def test_issue_cites_sheet_and_cell() -> None:
    wb = build_mini()
    wb["ft.construction_class"].append(["brick", 1.10, "", ""])
    result = run_check(wb)
    hit = next(i for i in result.errors if i.rule == "R-104")
    assert hit.sheet == "ft.construction_class"
    assert hit.cell is not None and hit.cell.startswith("A")
    assert "brick" in hit.message and "construction_class" in hit.message


# ---------------------------------------------------------------------------
# 2. The canonical bundle is green.
# ---------------------------------------------------------------------------

def test_canonical_bundle_checks_clean() -> None:
    result = check_workbook(CANONICAL.read_bytes(), filename=CANONICAL.name)
    assert result.errors == [], [(i.rule, i.sheet, i.message) for i in result.errors]
    assert result.ok is True
    m = result.manifest
    assert m is not None
    assert m.counts.dimensions == 7
    assert m.counts.factor_tables == 11
    assert m.counts.test_cases == 20
    assert m.counts.declared_gaps == 5
    assert m.counts.inputs == 7 and m.counts.inputs_with_defaults == 2
    assert m.counts.gates == 6
    assert m.counts.final_adjustments == 1
    assert m.provenance.product == "do"
    assert m.gap_kinds == {"unsupported": 3, "assumption": 1, "gap": 1}


# ---------------------------------------------------------------------------
# 3. Endpoint, CLI, registry sync.
# ---------------------------------------------------------------------------

def test_check_endpoint_roundtrip(client) -> None:  # noqa: ANN001 — conftest fixture
    body = to_bytes(build_mini())
    resp = client.post(
        "/api/v1/plans/ingest/check?filename=mini.xlsx",
        content=body,
        headers={"Content-Type": "application/octet-stream"},
    )
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["ok"] is True
    assert payload["filename"] == "mini.xlsx"
    assert payload["manifest"]["counts"]["factor_tables"] == 2


def test_check_endpoint_reports_errors_as_200(client) -> None:  # noqa: ANN001
    wb = build_mini()
    wb.remove(wb["inputs"])
    resp = client.post("/api/v1/plans/ingest/check", content=to_bytes(wb))
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["ok"] is False
    assert "R-020" in {e["rule"] for e in payload["errors"]}
    assert payload["manifest"] is None


def test_check_endpoint_empty_body_is_400(client) -> None:  # noqa: ANN001
    resp = client.post("/api/v1/plans/ingest/check", content=b"")
    assert resp.status_code == 400
    assert resp.json()["error"]["code"] == "ingest_empty_body"


def test_cli_exit_codes(tmp_path: Path, capsys) -> None:  # noqa: ANN001
    good = tmp_path / "good.xlsx"
    good.write_bytes(to_bytes(build_mini()))
    assert cli_main(["check", str(good)]) == 0
    out = capsys.readouterr().out
    assert "PASSED" in out and "2 tables" in out

    wb = build_mini()
    wb.remove(wb["inputs"])
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(to_bytes(wb))
    assert cli_main(["check", str(bad)]) == 1
    out = capsys.readouterr().out
    assert "FAILED" in out and "R-020" in out

    assert cli_main(["check", str(tmp_path / "missing.xlsx")]) == 2


def test_packaged_registry_matches_docs_copy() -> None:
    packaged = (
        resources.files("openrater.rates.ingest")
        .joinpath("capability_registry.json")
        .read_text(encoding="utf-8")
    )
    assert json.loads(packaged) == json.loads(DOCS_REGISTRY.read_text(encoding="utf-8"))


MERIDIAN = (
    REPO_ROOT / "docs" / "specs" / "examples" / "meridian-shopfront-bop"
    / "meridian_shopfront_bop.workbook.xlsx"
)


def test_meridian_all_constructs_bundle_checks_clean() -> None:
    """The 92.5 all-constructs bundle: 2-D matrix, geo ZIP detail,
    exposure towers, endorsements, modifiers, loadings, per-coverage
    clamp, floored round, compound gates — zero errors, zero warnings
    (the interpolation flag is a NOTICE since Brief 95 C5)."""
    result = check_workbook(MERIDIAN.read_bytes(), filename=MERIDIAN.name)
    assert result.ok, [i.model_dump() for i in result.errors]
    assert result.warnings == []
    assert "R-111" in {n.rule for n in result.notices}
    m = result.manifest
    assert m is not None
    assert m.counts.factor_tables == 8
    assert m.counts.chains == 3 and m.counts.chain_stages == 21
    # The reference program covers 40 classes and 6 territories; its vectors
    # and pre-existing factors remain unchanged.
    assert m.counts.geo_rows == 36
    assert m.counts.dimension_levels == 64
    assert m.counts.factor_cells == 115
    assert m.counts.factor_cells_cited == 115
    assert m.counts.endorsements == 2 and m.counts.modifier_categories == 2
    assert m.counts.loadings == 1 and m.counts.final_adjustments == 2
    assert m.provenance.state == "NE"


def test_chain_flat_factor_is_registry_refused() -> None:
    """r3 chain_flat_factor — the projector cannot key an unkeyed
    constant, so the CHECK refuses it (never a silent skip)."""
    wb = build_mini()
    ws = wb["chains"]
    ws.append(["building", 9, "flat_factor", "sneaky_constant", "", "", "",
               1.25, "", "", ""])
    result = check_workbook(to_bytes(wb))
    assert not result.ok
    assert any(
        e.rule == "R-190" and "constant chain factor" in e.message.lower()
        for e in result.errors
    )


def test_additive_endorsement_multi_coverage_is_registry_refused() -> None:
    """r3 endorsement_additive_multi_coverage — a once-per-policy
    amount has no single tower on a multi-coverage plan."""
    from openpyxl import load_workbook

    wb = load_workbook(MERIDIAN)
    wb["endorsements"].append(
        ["flat_fee", "additive", "MS 30 01", "Flat fee", "", 125, "", "",
         "", "cite", "p.1"]
    )
    result = check_workbook(to_bytes(wb))
    assert not result.ok
    assert any(
        e.rule == "R-190" and "additive endorsement" in e.message.lower()
        for e in result.errors
    )


# ---------------------------------------------------------------------------
# r4 predicate_beyond_equality — check = build's contract on predicate
# operators (2026-07-15 filing-digitization review: the check accepted
# the full §4.6.1 grammar on chains/loadings while the domain's
# FactorPredicate is equality-only, so the build refused what the
# check had passed).
# ---------------------------------------------------------------------------

_BEYOND_EQUALITY = [
    ("!=", "form_input.building_age != 5"),
    ("<", "form_input.building_age < 5"),
    ("<=", "form_input.building_age <= 5"),
    (">", "form_input.building_age > 5"),
    (">=", "form_input.building_age >= 5"),
    ("in", "form_input.construction_class in [frame, fire_resistive]"),
    ("not-in", "form_input.construction_class not-in [frame]"),
]
_IDS = [op for op, _ in _BEYOND_EQUALITY]


@pytest.mark.parametrize(("op", "pred"), _BEYOND_EQUALITY, ids=_IDS)
def test_chain_predicate_beyond_equality_is_registry_refused(
    op: str, pred: str
) -> None:
    """Every §4.6.1 operator except `==` on chains.predicate fails the
    CHECK — the same workbooks used to pass check and 500 on build."""
    wb = build_mini()
    wb["chains"].cell(row=3, column=10, value=pred)
    result = run_check(wb)
    assert not result.ok
    assert any(
        e.rule == "R-190" and f"'{op}'" in e.message and "equality" in e.message
        for e in result.errors
    ), [(e.rule, e.message) for e in result.errors]


@pytest.mark.parametrize(("op", "pred"), _BEYOND_EQUALITY, ids=_IDS)
def test_loading_predicate_beyond_equality_is_registry_refused(
    op: str, pred: str
) -> None:
    """loadings.predicate builds into the same equality-only
    FactorPredicate — the refusal covers it too."""
    wb = build_mini()
    _add_valid_loadings(wb)
    wb["loadings"].cell(row=2, column=6, value=pred)
    result = run_check(wb)
    assert not result.ok
    assert any(
        e.rule == "R-190" and f"'{op}'" in e.message for e in result.errors
    ), [(e.rule, e.message) for e in result.errors]


def test_equality_predicates_check_clean() -> None:
    """`==` stays expressible on chains AND loadings."""
    wb = build_mini()
    _add_valid_loadings(wb)
    wb["chains"].cell(row=3, column=10, value="form_input.construction_class == frame")
    wb["loadings"].cell(row=2, column=6, value="form_input.building_age == 10")
    result = run_check(wb)
    assert result.ok, [(i.rule, i.message) for i in result.errors]


@pytest.mark.parametrize(
    "trigger",
    [
        "form_input.building_age == 5",
        "form_input.building_age != 5",
        "form_input.building_age < 5",
        "form_input.building_age <= 5",
        "form_input.building_age > 5",
        "form_input.building_age >= 5",
        "form_input.construction_class in [frame, fire_resistive]",
        "form_input.construction_class not-in [frame]",
    ],
)
def test_endorsement_trigger_full_grammar_checks_clean(trigger: str) -> None:
    """endorsements.trigger executes all eight operators (the domain's
    EndorsementTriggerParams + the engine comparator) — the equality
    narrowing applies to chains/loadings ONLY."""
    wb = build_mini()
    _add_valid_endorsements(wb)
    wb["endorsements"].cell(row=2, column=9, value=trigger)
    result = run_check(wb)
    assert result.ok, [(i.rule, i.message) for i in result.errors]


# ---------------------------------------------------------------------------
# R-146 coverage:total needs its producing round row — the last
# check=build parity gap (the builder always refused this; the check
# used to accept coverage:total unconditionally and let the build 422).
# ---------------------------------------------------------------------------

def _coverage_total_output(wb: Workbook) -> None:
    wb["outputs"].append(
        ["out_total", "total_premium", "Total premium", "coverage:total"]
    )


def test_coverage_total_with_clamp_only_adjustments_is_refused() -> None:
    """A final_adjustments sheet with no round row doesn't satisfy
    coverage:total — the total's producer is the plan-tail round,
    not just any adjustment."""
    wb = build_mini()
    _add_valid_adjustments(wb)  # clamp only
    _coverage_total_output(wb)
    result = run_check(wb)
    assert not result.ok
    assert any(
        e.rule == "R-146" and "round" in e.message for e in result.errors
    ), [(e.rule, e.message) for e in result.errors]


def test_coverage_total_with_round_row_checks_clean() -> None:
    """With a package-level round row, coverage:total resolves — the
    same construct the builder accepts (parity in both directions)."""
    wb = build_mini()
    ws = wb.create_sheet("final_adjustments")
    ws.append(["adjustment_id", "kind", "order", "applies_to", "min_value",
               "max_value", "round_increment", "round_min"])
    ws.append(["package_round", "round", 1, "", "", "", 1, ""])
    _coverage_total_output(wb)
    result = run_check(wb)
    assert result.ok, [(i.rule, i.message) for i in result.errors]


# ---------------------------------------------------------------------------
# 4. R-191 umbrella (Brief 94.1 T1) — every machine-detectable `partial`
#    registry construct warns with the registry's message.
# ---------------------------------------------------------------------------

# How the mini workbook declares each detectable partial construct. A NEW
# partial construct that is machine-detectable MUST get an entry here (and a
# warning path in the checker) — this test failing is the spec §8 R-191
# umbrella doing its job. (Empty since Brief 95 C5 moved
# `linear_interpolation` to supported — the umbrella stays armed for the
# next partial construct.)
PARTIAL_SYNTHESIZERS: dict[str, Mutator] = {}

_NOT_DETECTABLE = ("not detectable", "not machine-detectable", "policy-level")


def test_every_detectable_partial_construct_warns_with_registry_message() -> None:
    """Spec §8 R-191 (umbrella): a `partial` construct passes the check
    with the registry's warning — via its construct-specific rule id
    where one exists (R-111 carries `linear_interpolation`). A partial
    construct that is workbook-detectable but silent would be a truth
    gap; this test makes that unshippable."""
    registry = json.loads(DOCS_REGISTRY.read_text(encoding="utf-8"))
    partials = [c for c in registry["constructs"] if c["status"] == "partial"]
    detectable = [
        c
        for c in partials
        if not any(
            token in (c.get("detected_by") or "") for token in _NOT_DETECTABLE
        )
    ]
    checked = 0
    for construct in detectable:
        cid = construct["id"]
        assert cid in PARTIAL_SYNTHESIZERS, (
            f"registry construct '{cid}' is partial and workbook-detectable "
            "but has no synthesizer in PARTIAL_SYNTHESIZERS — add one (and a "
            "checker warning path) so the R-191 umbrella holds."
        )
        wb = build_mini()
        PARTIAL_SYNTHESIZERS[cid](wb)
        result = run_check(wb)
        assert result.ok, [(i.rule, i.message) for i in result.errors]
        assert any(construct["message"] in w.message for w in result.warnings), (
            cid,
            [(w.rule, w.message) for w in result.warnings],
        )
        checked += 1
    # Every workbook-detectable partial construct warned — vacuously true
    # when none exist (all partial constructs graduated or are policy-
    # level), which is the umbrella at rest, not a gap.
    assert checked == len(detectable)


# ---------------------------------------------------------------------------
# 5. The starter kit (Brief 94 §2, register U1) — packaged assets.
# ---------------------------------------------------------------------------

TEMPLATE_DOCS = REPO_ROOT / "docs" / "specs" / "examples" / "template" / "openrater_workbook_template.xlsx"
SPEC_DOCS = REPO_ROOT / "docs" / "specs" / "filing-transcription-spec.md"

STARTER_KIT_FILES = {
    "spec": ("filing-transcription-spec.md", SPEC_DOCS),
    "template": ("openrater_workbook_template.xlsx", TEMPLATE_DOCS),
    "example": ("nonprofit_do_gl.workbook.xlsx", CANONICAL),
}


def _packaged_asset(filename: str) -> bytes:
    return (
        resources.files("openrater.rates.ingest")
        .joinpath("assets")
        .joinpath(filename)
        .read_bytes()
    )


def test_starter_kit_assets_match_docs_copies() -> None:
    """Every packaged starter-kit artifact is byte-identical to its
    docs/ source — the same no-drift guard the capability registry has.
    Regenerating a source without re-copying it here fails CI."""
    for kind, (filename, docs_path) in STARTER_KIT_FILES.items():
        assert _packaged_asset(filename) == docs_path.read_bytes(), (
            f"packaged {kind} asset ({filename}) drifted from {docs_path}"
        )


def test_starter_kit_endpoint_serves_assets(client) -> None:  # noqa: ANN001
    for kind, (filename, _docs_path) in STARTER_KIT_FILES.items():
        resp = client.get(f"/api/v1/plans/ingest/assets/{kind}")
        assert resp.status_code == 200, (kind, resp.status_code)
        assert filename in resp.headers["content-disposition"]
        assert resp.content == _packaged_asset(filename)
    if_missing = client.get("/api/v1/plans/ingest/assets/poster")
    assert if_missing.status_code == 404
    assert if_missing.json()["error"]["code"] == "ingest_asset_not_found"


def test_template_is_alive_check_side() -> None:
    """Brief 94 CT-2, check half: the downloadable template checks
    clean — zero errors, zero warnings; its only notice is the
    spec-blessed README sheet (R-203). The teaching `notes` columns
    must not trip R-202."""
    result = check_workbook(
        _packaged_asset("openrater_workbook_template.xlsx"),
        filename="openrater_workbook_template.xlsx",
    )
    assert result.ok, [(i.rule, i.message) for i in result.errors]
    assert result.warnings == [], [(w.rule, w.message) for w in result.warnings]
    assert [n.rule for n in result.notices] == ["R-203"]
    counts = result.manifest.counts
    assert counts.factor_tables == 2
    assert counts.test_cases == 1


# ---------------------------------------------------------------------------
# 6. Citation coverage (Brief 94 U5) — the manifest's factor-cell counter
#    mirrors R-201's rule.
# ---------------------------------------------------------------------------

def test_citation_coverage_counts_mirror_r201() -> None:
    """The mini workbook cites everything: 6 of 6 factor cells (2 1-D
    rows + 4 matrix cells). Stripping the 1-D table's table-level
    citation AND one row's own citation drops coverage to 5 AND fires
    R-201 — the counter and the warning can never disagree."""
    wb = build_mini()
    result = run_check(wb)
    c = result.manifest.counts
    assert (c.factor_cells, c.factor_cells_cited) == (6, 6)
    assert not any(w.rule == "R-201" for w in result.warnings)

    ws = wb["ft.construction_class"]
    # NB: cell(..., value=None) is a READ in openpyxl — assign explicitly.
    meta_row = find_row(ws, 1, "citation_rule")
    ws.cell(row=meta_row, column=2).value = None
    frame_row = find_row(ws, 1, "frame")
    ws.cell(row=frame_row, column=3).value = None

    result2 = run_check(wb)
    c2 = result2.manifest.counts
    assert (c2.factor_cells, c2.factor_cells_cited) == (6, 5)
    assert any(w.rule == "R-201" for w in result2.warnings)


# ---------------------------------------------------------------------------
# 7. Brief 94.5 hardening — R-002 merged cells + the geo/duplicate
#    firing tests the sweep never had (T3 · T4).
# ---------------------------------------------------------------------------

def test_merged_cells_refused_with_range_cited() -> None:
    """R-002: a merged range in a DATA sheet is an error citing
    sheet!range; a merged range on an ignored prose sheet is fine
    (R-203 skips it before the scan)."""
    wb = build_mini()
    wb["inputs"].merge_cells("A2:B2")
    result = run_check(wb)
    assert not result.ok
    hits = [e for e in result.errors if e.rule == "R-002"]
    assert hits and hits[0].sheet == "inputs" and hits[0].cell == "A2:B2", [
        (e.rule, e.sheet, e.cell) for e in result.errors
    ]

    prose = build_mini()
    readme = prose.create_sheet("README")
    readme.append(["OpenRater rating workbook"])
    readme.merge_cells("A1:C1")
    result2 = run_check(prose)
    assert result2.ok, [(i.rule, i.message) for i in result2.errors]
    assert not any(i.rule == "R-002" for i in result2.warnings + result2.notices)


def _add_geo_setup(wb: Workbook) -> None:
    """A minimal, valid geographic dimension (one territory) the geo
    negative tests mutate around."""
    ws = wb["dimensions"]
    ws.append(["territory", "Territory", "geographic", "rating-input", "string",
               "geographic", "zip", "subset:IL", ""])
    ws = wb["dimension_levels"]
    ws.append(["territory", "geographic", "t1", "Territory 1", "", "", "", "T1"])


def test_geo_sheet_without_geographic_dimension_is_r080() -> None:
    wb = build_mini()
    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code"])
    ws.append(["60601", "T1"])
    result = run_check(wb)
    assert not result.ok
    assert any(e.rule == "R-080" for e in result.errors), [
        (e.rule, e.message) for e in result.errors
    ]


def test_duplicate_geo_zip_is_r081() -> None:
    wb = build_mini()
    _add_geo_setup(wb)
    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code"])
    ws.append(["60601", "T1"])
    ws.append(["60601", "T1"])
    result = run_check(wb)
    assert not result.ok
    assert any(e.rule == "R-081" for e in result.errors), [
        (e.rule, e.message) for e in result.errors
    ]


def test_unmatched_territory_code_is_r082() -> None:
    wb = build_mini()
    _add_geo_setup(wb)
    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code"])
    ws.append(["60601", "T9"])
    result = run_check(wb)
    assert not result.ok
    assert any(e.rule == "R-082" for e in result.errors), [
        (e.rule, e.message) for e in result.errors
    ]


# ---------------------------------------------------------------------------
# 8. Geo-universe checks (R-083…R-086) — coverage holes surface at
#    validate time, and a bare territory list is categorical.
# ---------------------------------------------------------------------------

def test_scope_coverage_hole_is_r083_notice() -> None:
    """A subset:IL zip sheet mapping one ZCTA leaves the rest of the
    state's universe unmapped — a NOTICE naming the counts (a
    consequence, not a defect: a program may write only part of its
    state, and unmapped keys refuse honestly at rating time)."""
    wb = build_mini()
    _add_geo_setup(wb)
    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code"])
    ws.append(["60601", "T1"])
    result = run_check(wb)
    hits = [n for n in result.notices if n.rule == "R-083"]
    assert hits and "unmapped" in hits[0].message and "subset:IL" in hits[0].message, [
        (i.rule, i.message) for i in result.notices
    ]
    # Coverage is advice, not a gate.
    assert result.ok, [(e.rule, e.message) for e in result.errors]


def test_out_of_scope_key_is_r084_warning() -> None:
    """A Nebraska ZCTA inside a subset:IL sheet is a typo or a wrong
    scope — warned with the key's actual state."""
    wb = build_mini()
    _add_geo_setup(wb)
    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code"])
    ws.append(["60601", "T1"])
    ws.append(["68502", "T1"])
    result = run_check(wb)
    hits = [w for w in result.warnings if w.rule == "R-084"]
    assert hits and "68502 (NE)" in hits[0].message, [
        (i.rule, i.message) for i in result.warnings
    ]


def test_unknown_zcta_is_r085_notice() -> None:
    """A key the Census universe doesn't know — PO-box ZIPs have no
    ZCTA, and typos look exactly like this — is a notice."""
    wb = build_mini()
    _add_geo_setup(wb)
    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code"])
    ws.append(["60601", "T1"])
    ws.append(["99999", "T1"])
    result = run_check(wb)
    hits = [n for n in result.notices if n.rule == "R-085"]
    assert hits and "99999" in hits[0].message, [
        (i.rule, i.message) for i in result.notices
    ]


def test_geographic_without_indicator_is_r086_warning() -> None:
    """shape=geographic with zip granularity but NO geo sheet and NO
    territory_members: a bare territory list — the filing's geography
    never materialized, so the dim should be plain categorical."""
    wb = build_mini()
    _add_geo_setup(wb)  # declares the dim; adds no geo sheet
    result = run_check(wb)
    hits = [w for w in result.warnings if w.rule == "R-086"]
    assert hits and "categorical" in hits[0].message, [
        (i.rule, i.message) for i in result.warnings
    ]


def test_geo_sheet_presence_clears_r086() -> None:
    wb = build_mini()
    _add_geo_setup(wb)
    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code"])
    ws.append(["60601", "T1"])
    result = run_check(wb)
    assert not any(w.rule == "R-086" for w in result.warnings), [
        (i.rule, i.message) for i in result.warnings
    ]


def test_duplicate_loading_id_is_r007() -> None:
    wb = build_mini()
    ws = wb.create_sheet("loadings")
    ws.append(["loading_id", "factor_kind", "display_name", "factor",
               "applies_to", "predicate"])
    ws.append(["expense", "expense", "Expense load", 1.18, "", ""])
    ws.append(["expense", "profit", "Profit load", 1.05, "", ""])
    result = run_check(wb)
    assert not result.ok
    assert any(e.rule == "R-007" for e in result.errors), [
        (e.rule, e.message) for e in result.errors
    ]
