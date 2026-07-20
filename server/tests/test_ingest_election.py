# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""Brief 95 phase 95.5 — coverage election (C4, spec §4.1 `?` markers).

The check side (R-048: markers stripped for every name consumer, at
least one required coverage, electable needs an exposure stage) and
the build side (the chain spec carries `elective: true`, which the
projector turns into the coverage.election head)."""

from __future__ import annotations

import json

import pytest
from openpyxl import Workbook

from tests.test_ingest_build import _fake_score, tmp_db  # noqa: F401 — fixture
from tests.test_ingest_check import build_mini, find_row, run_check, to_bytes


def _set_coverages(wb: Workbook, value: str) -> None:
    ws = wb["plan"]
    ws.cell(row=find_row(ws, 1, "coverages"), column=2, value=value)


def build_electable() -> Workbook:
    """The mini workbook with a second (liability) coverage so building
    can be marked electable while liability stays required."""
    wb = build_mini()
    _set_coverages(wb, "building?,liability")
    ws = wb["chains"]
    ws.append(["liability", 0, "base", "liab_base", "", "", "literal:2.0",
               2.0, "", ""])
    ws.append(["liability", 1, "exposure", "liab_exposure", "", "",
               "form_input.units", "", 1, ""])
    ws.append(["liability", 2, "lcm", "liab_lcm", "", "", "", 1.30, "", ""])
    ws = wb["inputs"]
    ws.append(["units", "Rating units", "number", True, "", "", ""])
    ws = wb["outputs"]
    ws.append(["out_liability", "liability_premium", "Liability premium",
               "liab_exposure"])
    ws = wb["test_cases"]
    ws.cell(row=1, column=7, value="units")
    ws.cell(row=2, column=7, value=800)
    return wb


def _errors_for(result, rule: str) -> list[str]:
    return [i.message for i in result.errors if i.rule == rule]


# ---------------------------------------------------------------------------
# R-048 + the stripped-marker contract.
# ---------------------------------------------------------------------------

def test_election_markers_check_clean() -> None:
    result = run_check(build_electable())
    assert result.errors == [], [(i.rule, i.message) for i in result.errors]
    # The manifest sees CLEAN coverage names — markers never leak.
    assert result.manifest is not None
    assert "building?" not in json.dumps(result.manifest.model_dump())


def test_r048_all_elective_refused() -> None:
    wb = build_mini()
    _set_coverages(wb, "building?")
    result = run_check(wb)
    assert any(
        "at least one" in m for m in _errors_for(result, "R-048")
    ), [(i.rule, i.message) for i in result.errors]


def test_r048_elective_without_exposure_stage_refused() -> None:
    wb = build_electable()
    # Mark liability electable too — but give the marker to a block
    # with no exposure row by deleting liability's exposure line.
    _set_coverages(wb, "building,liability?")
    ws = wb["chains"]
    row = find_row(ws, 4, "liab_exposure")
    ws.delete_rows(row)
    # The outputs row pointed at the deleted stage — repoint at base.
    out = wb["outputs"]
    out.cell(row=find_row(out, 1, "out_liability"), column=4, value="liab_lcm")
    result = run_check(wb)
    assert any(
        "no exposure stage" in m for m in _errors_for(result, "R-048")
    ), [(i.rule, i.message) for i in result.errors]


def test_r048_elective_without_chain_refused() -> None:
    wb = build_mini()
    _set_coverages(wb, "building,extra?")
    result = run_check(wb)
    assert any(
        "no chain block" in m for m in _errors_for(result, "R-048")
    ), [(i.rule, i.message) for i in result.errors]


def test_marker_on_unknown_slug_still_slug_checked() -> None:
    wb = build_mini()
    _set_coverages(wb, "building,Bad Name?")
    result = run_check(wb)
    # The stripped name goes through the ordinary slug check.
    assert any("Bad Name" in i.message for i in result.errors)


# ---------------------------------------------------------------------------
# The build side — the chain spec carries the flag.
# ---------------------------------------------------------------------------

def test_build_emits_elective_flag(
    tmp_db, monkeypatch: pytest.MonkeyPatch  # noqa: ANN001
) -> None:
    from openrater.rates.ingest.service import build_workbook

    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0, "liability_premium": 2080.0}),
    )
    outcome = build_workbook(
        db=tmp_db, data=to_bytes(build_electable()), filename="elect.xlsx"
    )
    conn = tmp_db.connection()
    chain_cfg = json.loads(
        conn.execute(
            "SELECT config_json FROM rating_plan_stages "
            "WHERE rating_plan_id = ? AND stage_kind = 'multiplicative_chain'",
            (outcome.rating_plan_id,),
        ).fetchone()["config_json"]
    )
    conn.close()
    by_cov = {spec["coverage_value"]: spec for spec in chain_cfg["chains"]}
    assert by_cov["building"].get("elective") is True
    assert "elective" not in by_cov["liability"]
