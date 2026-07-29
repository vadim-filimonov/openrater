# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""Brief 92.R phase 92R.2 — the diff engine + the revision check.

Three layers:

  1. `diff_workbooks` over a scripted revision of the mini workbook —
     every state (added/changed/removed/unchanged), cell-grain factor
     moves with % chips, blast radius on removed levels, the filing's
     revised expectations, and byte-identical determinism.
  2. `POST /plans/{id}/reingest/check` — the happy diff, the dirty
     workbook (200 + cell-addressed report, no diff), the identity
     mismatch refusal, the no-history refusal, the missing plan 404,
     the pre-92R blob-less caveat, and the hand-edit flag.
  3. The CLI `diff` twin (exit codes + human output).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
from test_ingest_build import _fake_score  # noqa: E402
from test_ingest_check import build_mini, find_row, to_bytes  # noqa: E402

from openrater.rates.ingest.__main__ import main as cli_main  # noqa: E402
from openrater.rates.ingest.diff import diff_workbooks  # noqa: E402
from openrater.rates.ingest.parser import parse_workbook  # noqa: E402

# ---------------------------------------------------------------------------
# The scripted revision: the mini workbook, re-issued.
# ---------------------------------------------------------------------------


def build_base() -> Workbook:
    """The mini workbook + an endorsement (so the revision can remove
    one) — the base a plan was 'built' from."""
    wb = build_mini()
    ws = wb.create_sheet("endorsements")
    ws.append(["endorsement_id", "kind", "form_number", "display_name",
               "factor", "amount", "coverage", "sublimit", "trigger"])
    ws.append(["equip_breakdown", "factor", "BP 04 17", "Equipment breakdown",
               1.06, "", "", "", ""])
    return wb


def build_revision() -> Workbook:
    """v1.1.0: one factor +5%, a new level + its cells, the endorsement
    removed, one expectation revised, one 2-D cell moved."""
    wb = build_base()
    wb.remove(wb["endorsements"])

    ws = wb["plan"]
    ws.cell(row=find_row(ws, 1, "version"), column=2).value = "1.1.0"

    ft = wb["ft.construction_class"]
    ft.cell(row=find_row(ft, 1, "frame"), column=2).value = 1.05

    lv = wb["dimension_levels"]
    lv.append(["construction_class", "categorical", "masonry", "Masonry (ISO 2)",
               "iso2", "", "", ""])
    ft.append(["masonry", 0.92, "Table 5.A", "p.51"])
    fx = wb["ft.constr_x_age"]
    fx.append(["masonry", 0.97, 1.08])

    tc = wb["test_cases"]
    tc.cell(row=find_row(tc, 1, "tc_1"), column=6).value = 409.50

    fx.cell(row=find_row(fx, 1, "fire_resistive"), column=3).value = 1.10
    return wb


def _diff():
    base, _ = parse_workbook(to_bytes(build_base()))
    new, _ = parse_workbook(to_bytes(build_revision()))
    return diff_workbooks(base, new)


def test_diff_states_and_cell_grain() -> None:
    d = _diff()
    # The masonry LEVEL is the added construct; its two new table cells
    # ride as field-changes inside the changed tables (D3 — cells
    # belong to their table's story).
    assert (d.totals.added, d.totals.changed, d.totals.removed) == (1, 4, 1)
    by_section = {s.section: s for s in d.sections}

    # The +5% factor move, cell-grain with the % chip.
    ftab = by_section["factor_tables"]
    cc = next(i for i in ftab.items if i.key == "ft.construction_class")
    frame = next(c for c in cc.changes if c.field == "frame")
    assert (frame.from_, frame.to, frame.pct) == (1.0, 1.05, 5.0)
    # The added level's cell rides the same table as an added field.
    masonry = next(c for c in cc.changes if c.field == "masonry")
    assert masonry.from_ is None and masonry.to == 0.92

    # The removed endorsement names its form + trigger posture.
    endo = by_section["endorsements"]
    assert endo.removed == 1
    assert "BP 04 17" in endo.items[0].summary
    assert "always attached" in endo.items[0].summary

    # The filing's revised expectation speaks in those words, with %.
    tcs = by_section["test_cases"]
    revised = next(i for i in tcs.items if i.state == "changed")
    assert "The filing revised tc_1" in revised.summary
    assert "(+5.0%)" in revised.summary

    # The added level is an added item; unchanged rows are counted.
    dims = by_section["dimensions"]
    assert any(i.state == "added" and "masonry" in i.key for i in dims.items)
    assert dims.unchanged > 0

    # The plan sheet records the version bump.
    plan = by_section["plan"]
    ver = next(c for c in plan.items[0].changes if c.field == "version")
    assert (ver.from_, ver.to) == ("1.0.0", "1.1.0")


def test_removed_level_blast_radius() -> None:
    base, _ = parse_workbook(to_bytes(build_base()))
    # Remove the fire_resistive level + its cells everywhere.
    pruned = build_base()
    lv = pruned["dimension_levels"]
    for r in range(lv.max_row, 1, -1):
        if lv.cell(row=r, column=3).value == "fire_resistive":
            lv.delete_rows(r)
    ft = pruned["ft.construction_class"]
    for r in range(ft.max_row, 1, -1):
        if ft.cell(row=r, column=1).value == "fire_resistive":
            ft.delete_rows(r)
    fx = pruned["ft.constr_x_age"]
    for r in range(fx.max_row, 1, -1):
        if fx.cell(row=r, column=1).value == "fire_resistive":
            fx.delete_rows(r)
    new, _ = parse_workbook(to_bytes(pruned))
    d = diff_workbooks(base, new)
    dims = next(s for s in d.sections if s.section == "dimensions")
    removed = next(i for i in dims.items if i.state == "removed")
    # 1 one-D row + 2 matrix cells = 3 cells of blast radius.
    assert "3 factor cells" in removed.summary


def test_diff_is_deterministic() -> None:
    a = _diff().model_dump_json(by_alias=True)
    b = _diff().model_dump_json(by_alias=True)
    assert a == b


def test_gate_diff_speaks_variables_and_pct_stays_rate_grammar() -> None:
    """MVP-022 — a gate move names its VARIABLE ("years_in_business
    threshold 3 → 5"), never the sheet column ("value_2 changed"); and
    pct is rate grammar: absent on thresholds, present on factor cells
    and revised expectations (pinned above)."""

    def with_gates(value: int, value_2: int) -> Workbook:
        wb = build_base()
        ws = wb.create_sheet("gates")
        ws.append(["order", "rule_id", "variable", "op", "value",
                   "variable_2", "op_2", "value_2", "tier", "reasoning"])
        ws.append([1, "decline_big_young", "tiv", "gt", value,
                   "years_in_business", "lt", value_2, "decline",
                   "Large + young"])
        ws.append([99, "__default__", "", "", "", "", "", "",
                   "standard", "Standard appetite"])
        return wb

    base, _ = parse_workbook(to_bytes(with_gates(5_000_000, 3)))
    new, _ = parse_workbook(to_bytes(with_gates(6_000_000, 5)))
    gates = next(
        s for s in diff_workbooks(base, new).sections if s.section == "gates"
    )
    item = next(i for i in gates.items if i.state == "changed")
    assert item.summary == (
        "Rule decline_big_young: tiv threshold 5000000 → 6000000; "
        "years_in_business threshold 3 → 5."
    )
    assert "value_2" not in item.summary
    # A year-count moving 3 → 5 is not "+66.7%".
    assert all(c.pct is None for c in item.changes)


# ---------------------------------------------------------------------------
# The endpoint.
# ---------------------------------------------------------------------------


def _build_plan(client, monkeypatch) -> str:  # noqa: ANN001
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    resp = client.post(
        "/api/v1/plans/ingest?filename=base.xlsx", content=to_bytes(build_base())
    )
    assert resp.status_code == 200, resp.text
    return resp.json()["rating_plan_id"]


def test_reingest_check_happy_diff(client, monkeypatch) -> None:  # noqa: ANN001
    plan_id = _build_plan(client, monkeypatch)
    resp = client.post(
        f"/api/v1/plans/{plan_id}/reingest/check?filename=rev.xlsx",
        content=to_bytes(build_revision()),
    )
    assert resp.status_code == 200, resp.text
    payload = resp.json()
    assert payload["check"]["ok"] is True
    assert payload["base"]["workbook_version"] == "1.0.0"
    assert payload["base_missing_reason"] is None
    assert payload["hand_edited_since_build"] is False
    assert payload["plan_content_hash"]
    totals = payload["diff"]["totals"]
    assert totals["removed"] == 1 and totals["added"] == 1
    # The wire uses ADR-0017's field-change alias.
    ftab = next(
        s for s in payload["diff"]["sections"] if s["section"] == "factor_tables"
    )
    frame = next(
        c
        for i in ftab["items"]
        for c in i["changes"]
        if c["field"] == "frame"
    )
    assert frame["from"] == 1.0 and frame["to"] == 1.05 and frame["pct"] == 5.0


def test_reingest_check_dirty_workbook_reports_no_diff(
    client, monkeypatch  # noqa: ANN001
) -> None:
    plan_id = _build_plan(client, monkeypatch)
    dirty = build_revision()
    dirty.remove(dirty["inputs"])
    resp = client.post(
        f"/api/v1/plans/{plan_id}/reingest/check", content=to_bytes(dirty)
    )
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["check"]["ok"] is False
    assert "R-020" in {e["rule"] for e in payload["check"]["errors"]}
    assert payload["diff"] is None


def test_reingest_check_identity_mismatch_refuses(
    client, monkeypatch  # noqa: ANN001
) -> None:
    plan_id = _build_plan(client, monkeypatch)
    stranger = build_revision()
    ws = stranger["plan"]
    ws.cell(row=find_row(ws, 1, "rating_plan_id"), column=2).value = "another-plan-2026"
    resp = client.post(
        f"/api/v1/plans/{plan_id}/reingest/check", content=to_bytes(stranger)
    )
    assert resp.status_code == 422
    err = resp.json()["error"]
    assert err["code"] == "reingest_identity_mismatch"
    assert "another-plan-2026" in err["message"]
    assert plan_id in err["message"]


def test_reingest_check_refusals_and_fallbacks(client, monkeypatch) -> None:  # noqa: ANN001
    # Missing plan → 404.
    resp = client.post(
        "/api/v1/plans/ghost_plan/reingest/check", content=to_bytes(build_revision())
    )
    assert resp.status_code == 404

    # A hand-built plan (no build history) → named refusal.
    created = client.post(
        "/api/v1/plans",
        json={"display_name": "Hand-built", "product": "bop", "effective_date": "2026-01-01"},
    )
    assert created.status_code in (200, 201), created.text
    hand_id = created.json()["rating_plan_id"]
    resp = client.post(
        f"/api/v1/plans/{hand_id}/reingest/check", content=to_bytes(build_revision())
    )
    assert resp.status_code == 422
    assert resp.json()["error"]["code"] == "reingest_no_build_history"

    # Pre-92R rows (no stored bytes) → the honest caveat, no diff.
    plan_id = _build_plan(client, monkeypatch)
    db = client.app.state.db
    conn = db.connection()
    conn.execute(
        "UPDATE plan_build_reports SET workbook_blob = NULL WHERE rating_plan_id = ?",
        (plan_id,),
    )
    conn.commit()
    resp = client.post(
        f"/api/v1/plans/{plan_id}/reingest/check", content=to_bytes(build_revision())
    )
    assert resp.status_code == 200
    payload = resp.json()
    assert payload["diff"] is None
    assert "before revisions stored the workbook bytes" in payload["base_missing_reason"]

    # Hand-edited-since-build flag (CT-4's data half).
    conn = db.connection()
    # A REAL edit (drift-honesty: with the as-built hash stored, a bare
    # timestamp touch that changed no content is no longer an edit).
    resp_edit = client.put(
        f"/api/v1/plans/{plan_id}/factor-tables/construction_class/cells",
        json={"cells": {"frame": 0.9, "jm": 1.15}},
    )
    assert resp_edit.status_code == 200, resp_edit.text
    resp = client.post(
        f"/api/v1/plans/{plan_id}/reingest/check", content=to_bytes(build_revision())
    )
    assert resp.json()["hand_edited_since_build"] is True


# ---------------------------------------------------------------------------
# The CLI twin.
# ---------------------------------------------------------------------------


def test_cli_diff(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    base = tmp_path / "base.xlsx"
    base.write_bytes(to_bytes(build_base()))
    rev = tmp_path / "rev.xlsx"
    rev.write_bytes(to_bytes(build_revision()))

    assert cli_main(["diff", str(base), str(rev)]) == 0
    out = capsys.readouterr().out
    assert "1 added · 4 changed · 1 removed" in out
    assert "frame: 1 → 1.05  (+5.0%)" in out

    assert cli_main(["diff", str(base), str(rev), "--json"]) == 0
    payload = json.loads(capsys.readouterr().out)
    assert payload["totals"]["removed"] == 1

    assert cli_main(["diff", str(base), str(tmp_path / "missing.xlsx")]) == 2
