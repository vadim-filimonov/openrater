# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""The eval harness's own regression net (Brief 2 P2).

Two invariants keep the scorer honest:
  · IDENTITY — the golden scored against itself is a perfect PASS
    (otherwise every real attempt is graded against a broken ruler);
  · SENSITIVITY — a mutant golden with one wrong factor cell, one
    deleted gap row, and one altered worked-example total FAILS with
    exactly those findings (otherwise the scorer is a rubber stamp).
"""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import openpyxl
import pytest

_REPO = Path(__file__).resolve().parents[2]
_SCORER = _REPO / "evals" / "score_transcription.py"
_GOLDEN = (
    _REPO
    / "docs/specs/examples/meridian-shopfront-bop/meridian_shopfront_bop.workbook.xlsx"
)


@pytest.fixture(scope="module")
def scorer():
    spec = importlib.util.spec_from_file_location("score_transcription", _SCORER)
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules["score_transcription"] = mod
    spec.loader.exec_module(mod)
    return mod


def _score(scorer, attempt: Path) -> dict:
    g_result, g_model = scorer.parse(_GOLDEN)
    a_result, a_model = scorer.parse(attempt)
    report = {
        "check": scorer.score_check(a_result),
        "cells": scorer.score_cells(g_model, a_model),
        "examples": scorer.score_examples(g_model, a_model),
        "gaps": scorer.score_gaps(a_model),
    }
    passed, failed = scorer.verdict(report)
    report["verdict"] = {"pass": passed, "failed_bars": failed}
    return report


def test_identity_is_a_perfect_pass(scorer):
    report = _score(scorer, _GOLDEN)
    assert report["verdict"]["pass"], report["verdict"]["failed_bars"]
    assert report["check"]["ok"]
    assert report["cells"]["accuracy"] == 1.0
    assert report["cells"]["golden_cells"] == 151  # 115 factor cells + 36 geo rows
    assert report["examples"]["totals_match"] == 8
    assert report["examples"]["tiers_match"] == 8
    assert report["gaps"]["missing_topics"] == []


def test_mutant_fails_on_exactly_the_injected_defects(scorer, tmp_path):
    wb = openpyxl.load_workbook(_GOLDEN)

    # Defect 1 — one wrong factor cell: the sprinklered credit 0.92 → 0.95.
    ws = wb["ft.sprinkler_prop"]
    hits = [c for row in ws.iter_rows() for c in row if c.value == 0.92]
    assert len(hits) == 1, "expected exactly one 0.92 cell in the sprinkler table"
    hits[0].value = 0.95

    # Defect 2 — a genuinely DROPPED convention: delete the
    # years-in-business gap row AND clear the input's default_value.
    # (Either one alone is an honest capture — the axis accepts a gaps
    # row or the filing's default on the named input; only losing both
    # is the dishonesty it exists to catch.)
    ws = wb["gaps_and_assumptions"]
    target = next(
        r[0].row
        for r in ws.iter_rows()
        if any("Years-in-business" in str(c.value or "") for c in r)
    )
    ws.delete_rows(target)
    ws = wb["inputs"]
    header = {c.value: c.column for c in ws[1]}
    name_col, default_col = header["name"], header["default_value"]
    yib = next(
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=name_col).value == "years_in_business"
    )
    assert ws.cell(row=yib, column=default_col).value == 5
    ws.cell(row=yib, column=default_col).value = None

    # Defect 3 — a mis-transcribed worked example: mv_01's total 1898 → 1900.
    ws = wb["test_cases"]
    header = {c.value: c.column for c in ws[1]}
    col = header["expected_total_premium"]
    row = next(
        r[0].row for r in ws.iter_rows(min_row=2) if r[0].value == "mv_01"
    )
    assert ws.cell(row=row, column=col).value == 1898
    ws.cell(row=row, column=col).value = 1900

    mutant = tmp_path / "mutant.xlsx"
    wb.save(mutant)

    report = _score(scorer, mutant)
    assert not report["verdict"]["pass"]

    cells = report["cells"]
    assert len(cells["wrong_value"]) == 1
    assert cells["wrong_value"][0]["table"] == "sprinkler_prop"
    assert cells["wrong_value"][0]["attempt"] == 0.95
    assert cells["missing"] == [] and cells["extra"] == []

    gaps = report["gaps"]
    assert gaps["missing_topics"] == ["years_in_business_default"]
    assert "sprinkler_default" in gaps["matched_topics"]

    ex = report["examples"]
    assert ex["totals_match"] == 7
    assert any("mv_01" in f and "1900" in f for f in ex["findings"])
    # The tier and input columns were untouched — no collateral findings.
    assert ex["tiers_match"] == 8
    assert ex["inputs_faithful"] == 8

    # All three injected defects surface as distinct failed bars.
    assert set(report["verdict"]["failed_bars"]) == {
        "every factor cell matches the filing",
        "all worked examples transcribed with the filing's totals + tiers",
        "both Rule A.4 conventions recorded as gaps",
    }


def test_scorer_cli_exit_codes(scorer, tmp_path):
    """The CLI is the eval interface — 0 on PASS, 1 on FAIL, and the
    JSON report lands where asked."""
    out = tmp_path / "report.json"
    rc = scorer.main([str(_GOLDEN), "--json", str(out)])
    assert rc == 0
    assert out.is_file()

    wb = openpyxl.load_workbook(_GOLDEN)
    ws = wb["ft.sprinkler_prop"]
    for row in ws.iter_rows():
        for c in row:
            if c.value == 0.92:
                c.value = 0.90
    mutant = tmp_path / "mutant.xlsx"
    wb.save(mutant)
    assert scorer.main([str(mutant)]) == 1
