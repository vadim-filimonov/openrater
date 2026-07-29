# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""FCA #16 follow-up — the current-state export: repairs physically
travel.

The default export stays byte-exact (owner O1: hash-identical round
trips answer `ingest_already_built`). `?current=true` serves the SAME
container with the live plan state written into the two tracked edit
classes — factor-table cells (1-D by level_id row, grids by row::col)
and gates!value cells — surgically: only the touched sheets' zip
entries differ; every other entry is byte-identical. What cannot be
placed is NAMED in `X-Current-Unapplied`, never silently dropped.
"""

from __future__ import annotations

import hashlib
import io
import sys
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent))
from test_ingest_build import _fake_score  # noqa: E402
from test_ingest_check import build_mini, find_row, to_bytes  # noqa: E402


def _gated_mini() -> Workbook:
    """The mini workbook + a gates sheet (a numeric threshold rule, a
    string-valued rule, the required __default__ row) — and one 2-D
    grid cell left BLANK (spec: blank = refuse, never zero), so the
    app can fill it in and the export must INSERT a cell the sheet
    never serialized."""
    wb = build_mini()
    fx = wb["ft.constr_x_age"]
    fx.cell(row=find_row(fx, 1, "fire_resistive"), column=3).value = None
    ws = wb.create_sheet("gates")
    ws.append(
        ["order", "rule_id", "variable", "op", "value", "tier", "reasoning"]
    )
    ws.append([1, "decline_large", "tiv", "gt", 5000000, "decline", "Too large"])
    ws.append(
        [2, "class_gate", "construction_class", "eq", "frame", "submit",
         "Frame referral"]
    )
    ws.append([99, "__default__", "", "", "", "standard", "Standard appetite"])
    return wb


def _build_gated_plan(client, monkeypatch) -> tuple[str, bytes]:  # noqa: ANN001
    """Build the gated mini via the endpoint; return (plan_id, posted bytes)."""
    monkeypatch.setattr(
        "openrater.rates.ingest.vectors.score_once",
        _fake_score({"building_premium": 390.0}),
    )
    data = to_bytes(_gated_mini())
    resp = client.post("/api/v1/plans/ingest?filename=mini.xlsx", content=data)
    assert resp.status_code == 200, resp.text
    return resp.json()["rating_plan_id"], data


def _patch_gate_value(client, plan_id: str, value: int) -> None:  # noqa: ANN001
    """The in-app gate repair: read the live eligibility config, move
    the rule's threshold, and save it back — the same loop the
    Eligibility tab drives."""
    cfg = client.get(
        f"/api/v1/plans/{plan_id}/stages/eligibility_gate/config"
    ).json()["config_json"]
    assert cfg["rules"][0]["rule_id"] == "decline_large"
    cfg["rules"][0]["value"] = value
    resp = client.patch(
        f"/api/v1/drafts/{plan_id}",
        json={
            "stage_patches": [
                {"stage_id": "eligibility_gate", "config_json": cfg}
            ],
            "note": "raise the TIV gate",
        },
    )
    assert resp.status_code == 200, resp.text


def test_current_export_carries_cell_and_gate_edits_surgically(
    client, monkeypatch  # noqa: ANN001
) -> None:
    plan_id, posted = _build_gated_plan(client, monkeypatch)

    # The in-app repairs: a 1-D factor cell, a 2-D grid cell, a fill
    # of the grid cell the workbook left blank, a numeric gate
    # threshold, and a string-valued gate literal.
    resp = client.put(
        f"/api/v1/plans/{plan_id}/factor-tables/construction_class/cells",
        json={"cells": {"frame": 0.85, "fire_resistive": 0.78}},
    )
    assert resp.status_code == 200, resp.text
    resp = client.put(
        f"/api/v1/plans/{plan_id}/factor-tables/constr_x_age/cells",
        json={
            "cells": {
                "frame::age_0_25": 1.11,
                "frame::age_25_plus": 1.20,
                "fire_resistive::age_0_25": 0.95,
                "fire_resistive::age_25_plus": 1.05,
            }
        },
    )
    assert resp.status_code == 200, resp.text
    _patch_gate_value(client, plan_id, 7500000)
    cfg = client.get(
        f"/api/v1/plans/{plan_id}/stages/eligibility_gate/config"
    ).json()["config_json"]
    assert cfg["rules"][1]["rule_id"] == "class_gate"
    cfg["rules"][1]["value"] = "fire_resistive"
    resp = client.patch(
        f"/api/v1/drafts/{plan_id}",
        json={
            "stage_patches": [
                {"stage_id": "eligibility_gate", "config_json": cfg}
            ],
            "note": "gate the other class",
        },
    )
    assert resp.status_code == 200, resp.text

    # The DEFAULT export stays the build's exact bytes (owner O1),
    # stamped as diverged.
    build = client.get(f"/api/v1/plans/{plan_id}/workbook")
    assert build.status_code == 200
    assert build.content == posted
    assert build.headers["x-workbook-state"] == "build"
    assert build.headers["x-workbook-hash"]
    assert build.headers["x-edited-since-build"] == "true"

    # The CURRENT export: -current filename, no build-identity claim,
    # a fresh content hash, and every tracked edit written in.
    cur = client.get(f"/api/v1/plans/{plan_id}/workbook?current=true")
    assert cur.status_code == 200
    assert 'filename="mini-current.xlsx"' in cur.headers["content-disposition"]
    assert cur.headers["x-workbook-state"] == "current"
    assert "x-workbook-hash" not in cur.headers
    assert (
        cur.headers["x-workbook-sha256"]
        == hashlib.sha256(cur.content).hexdigest()
    )
    assert int(cur.headers["x-current-rewrite-count"]) == 5
    assert int(cur.headers["x-current-unapplied-count"]) == 0
    assert cur.headers["x-edited-since-build"] == "true"

    # Re-parse: both edit classes physically landed — including the
    # cell INSERTED into a slot the sheet never serialized, and the
    # string gate literal (an inline string; sharedStrings untouched).
    wb = load_workbook(io.BytesIO(cur.content))
    ft = wb["ft.construction_class"]
    assert ft.cell(row=find_row(ft, 1, "frame"), column=2).value == 0.85
    fx = wb["ft.constr_x_age"]
    assert fx.cell(row=find_row(fx, 1, "frame"), column=2).value == 1.11
    assert fx.cell(row=find_row(fx, 1, "fire_resistive"), column=3).value == 1.05
    g = wb["gates"]
    assert g.cell(row=find_row(g, 2, "decline_large"), column=5).value == 7500000
    assert (
        g.cell(row=find_row(g, 2, "class_gate"), column=5).value
        == "fire_resistive"
    )

    # Surgical: ONLY the touched sheets' entries differ; every other
    # zip entry — plan, inputs, chains, styles, doc properties — is
    # byte-identical to the build container.
    from openrater.rates.ingest.current import _sheet_paths

    src = zipfile.ZipFile(io.BytesIO(posted))
    out = zipfile.ZipFile(io.BytesIO(cur.content))
    assert src.namelist() == out.namelist()
    paths = _sheet_paths(src)
    touched = {
        paths["ft.construction_class"],
        paths["ft.constr_x_age"],
        paths["gates"],
    }
    for name in src.namelist():
        if name in touched:
            assert src.read(name) != out.read(name), (
                f"{name} carries a rewrite — its entry should differ"
            )
        else:
            assert src.read(name) == out.read(name), (
                f"{name} is untouched — its entry must stay byte-stable"
            )

    # The current bytes are a WORKING workbook: they check clean, and
    # they register as a revision of this plan — never as the build
    # identity (`already_built` is the byte-exact export's answer).
    check = client.post("/api/v1/plans/ingest/check", content=cur.content).json()
    assert check["ok"] is True, check
    assert check["already_built"] is None
    assert check["revises"]["rating_plan_id"] == plan_id

    # The closing of the loop: re-ingesting the current export lands
    # the live state as the new build — the plan comes out PRISTINE
    # (no edits-since-build), because the repairs travelled in the
    # file instead of being replaced by it.
    applied = client.post(
        f"/api/v1/plans/{plan_id}/reingest?replace_edits=true",
        content=cur.content,
    )
    assert applied.status_code == 200, applied.text
    after = client.get(f"/api/v1/plans/{plan_id}/edits-since-build").json()
    assert after["edited"] is False
    assert after["changes"] == []


def test_current_export_without_edits_is_the_build_bytes(
    client, monkeypatch  # noqa: ANN001
) -> None:
    """Zero divergence: the build bytes ARE current, so the identity
    claim honestly rides along — under the -current name and state."""
    plan_id, posted = _build_gated_plan(client, monkeypatch)
    cur = client.get(f"/api/v1/plans/{plan_id}/workbook?current=true")
    assert cur.status_code == 200
    assert cur.content == posted
    assert cur.headers["x-workbook-state"] == "current"
    assert cur.headers["x-workbook-hash"]  # bytes ARE the build bytes
    assert int(cur.headers["x-current-rewrite-count"]) == 0
    assert 'filename="mini-current.xlsx"' in cur.headers["content-disposition"]


def test_current_export_names_what_it_cannot_place(
    client, monkeypatch  # noqa: ANN001
) -> None:
    """Honest degrade: a comparator change (gt -> ge) has no value-cell
    rewrite that wouldn't lie — the export refuses it BY NAME instead
    of writing a value against the wrong op."""
    plan_id, posted = _build_gated_plan(client, monkeypatch)
    cfg = client.get(
        f"/api/v1/plans/{plan_id}/stages/eligibility_gate/config"
    ).json()["config_json"]
    cfg["rules"][0]["op"] = "ge"
    cfg["rules"][0]["value"] = 7500000
    resp = client.patch(
        f"/api/v1/drafts/{plan_id}",
        json={
            "stage_patches": [
                {"stage_id": "eligibility_gate", "config_json": cfg}
            ],
            "note": "gate now ge",
        },
    )
    assert resp.status_code == 200, resp.text

    cur = client.get(f"/api/v1/plans/{plan_id}/workbook?current=true")
    assert cur.status_code == 200
    assert int(cur.headers["x-current-rewrite-count"]) == 0
    assert int(cur.headers["x-current-unapplied-count"]) == 1
    assert "decline_large" in cur.headers["x-current-unapplied"]
    # Nothing travelled, so the served bytes are still the build bytes
    # (with the identity claim), and the gates row keeps the as-built
    # comparator AND value — never a half-rewritten lie.
    assert cur.content == posted
    g = load_workbook(io.BytesIO(cur.content))["gates"]
    row = find_row(g, 2, "decline_large")
    assert g.cell(row=row, column=4).value == "gt"
    assert g.cell(row=row, column=5).value == 5000000
