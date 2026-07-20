# Copyright 2026 Vadim Filimonov and the OpenRater contributors
# SPDX-License-Identifier: Apache-2.0
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
"""Workbook reader — bytes → ParsedWorkbook + structural issues.

Structure-level rules live here (they're discovered while reading):
R-001 unreadable · R-003 header/required columns · R-004 formulas ·
R-020..R-026 required sheets · R-100 ft.* metadata block ·
R-202 unknown trailing columns · R-203 unrecognized sheets.
Everything content-shaped lives in `lint.py`.

The parser is deliberately lenient past structure: it loads whatever
it can so lint can report the maximum number of problems in one pass
(the fix-and-re-drop loop should converge fast, not one error at a
time).
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from openrater.rates.ingest.model import (
    MISSING_SHEET_RULE,
    OPTIONAL_SHEETS,
    REQUIRED_SHEETS,
    SHEET_PREFIXES,
    Cellv,
    CheckIssue,
    FactorGridCell,
    FactorTable,
    GeoSheet,
    ParsedWorkbook,
    Row,
    Table,
    issue,
)

# Canonical column sets per sheet (spec §4). `notes` is always allowed.
SHEET_COLUMNS: dict[str, tuple[tuple[str, ...], tuple[str, ...]]] = {
    # sheet: (required columns, optional columns)
    "inputs": (
        ("name", "label", "data_type", "required"),
        (
            "allowed_values",
            "default_value",
            "unit",
            "min",
            "max",
            "maps_to_dimension",
            "derived_from",
            "description",
            "citation_rule",
            "citation_page",
        ),
    ),
    "dimensions": (
        ("slug", "display_name", "shape", "role", "data_type"),
        (
            "dimension_type",
            "geo_granularity",
            "geo_scope",
            "axes",
            "class_library_id",
            "description",
            "source_page",
        ),
    ),
    "dimension_levels": (
        ("dimension_slug", "kind", "level_id", "label"),
        (
            "aliases",
            "min",
            "max",
            "territory_ref",
            "territory_members",
            "citation_rule",
            "citation_page",
        ),
    ),
    "chains": (
        ("coverage", "order", "stage_kind", "stage_id"),
        (
            "coverage_value",
            "factor_table",
            "dimension",
            "input_binding",
            "value",
            "exposure_divisor",
            "predicate",
            "description",
            "citation_rule",
            "citation_page",
        ),
    ),
    "gates": (
        ("order", "rule_id", "variable", "op", "value", "tier", "reasoning"),
        (
            "variable_2",
            "op_2",
            "value_2",
            "variable_3",
            "op_3",
            "value_3",
            "citation_rule",
            "citation_page",
        ),
    ),
    "modifiers": (
        (
            "schedule_id",
            "schedule_name",
            "scope",
            "total_cap_pct",
            "category_id",
            "category_name",
            "range_pct",
        ),
        ("tier_filter", "citation_rule", "citation_page"),
    ),
    "endorsements": (
        ("endorsement_id", "kind", "form_number", "display_name"),
        (
            "factor",
            "amount",
            "coverage",
            "sublimit",
            "trigger",
            "citation_rule",
            "citation_page",
        ),
    ),
    "loadings": (
        ("loading_id", "factor_kind", "display_name", "factor"),
        (
            "factor_unit",
            "applies_to",
            "predicate",
            "citation_rule",
            "citation_page",
        ),
    ),
    "final_adjustments": (
        ("adjustment_id", "kind", "order"),
        (
            "applies_to",
            "min_value",
            "max_value",
            "round_increment",
            "round_min",
            "citation_rule",
            "citation_page",
        ),
    ),
    "outputs": (
        ("output_id", "field_name", "display_name", "source"),
        ("description",),
    ),
    "gaps_and_assumptions": (
        ("kind", "description", "impact"),
        ("citation_rule", "citation_page", "related"),
    ),
}

# Spec §4.15 — no `factor` column: per-geography factors live in ft.*
# tables keyed by the geographic dimension (Brief 95 C6; a stray factor
# column falls to the R-202 unknown-trailing-column warning, preserved).
GEO_COLUMNS = (("territory_code",), ("zip", "county", "citation_rule", "citation_page"))

FT_META_REQUIRED = ("table_id", "display_name", "dimensionality", "row_dimension", "lookup_method")
FT_META_OPTIONAL = (
    "col_dimension",
    "interpolation",
    "default_value",
    "citation_rule",
    "citation_page",
    "citation_note",
)


def _ref(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _scan_formulas(ws: Worksheet, sheet: str, issues: list[CheckIssue]) -> None:
    for row in ws.iter_rows():
        for c in row:
            if c.data_type == "f":
                issues.append(
                    issue(
                        "R-004",
                        "error",
                        "Formula found — write the computed literal value instead "
                        "(the reader never evaluates formulas).",
                        sheet=sheet,
                        cell=c.coordinate,
                    )
                )


def _scan_merged_cells(ws: Worksheet, sheet: str, issues: list[CheckIssue]) -> None:
    """R-002 (Brief 94.5) — a merged range reads as its top-left cell
    plus blanks, so a merged header silently drops a column and a
    merged data cell silently empties its neighbors. Refuse with the
    range cited; ignored prose sheets never reach here (R-203 skips
    them before the scan)."""
    for merged in ws.merged_cells.ranges:
        issues.append(
            issue(
                "R-002",
                "error",
                "Merged cells — unmerge the range; every cell but its "
                "top-left reads as blank (the reader never guesses what a "
                "merge meant).",
                sheet=sheet,
                cell=str(merged),
            )
        )


def _row_values(ws: Worksheet, row_idx: int) -> list[Any]:
    return [c.value for c in ws[row_idx]] if row_idx <= ws.max_row else []


def _is_blank_row(values: list[Any]) -> bool:
    return all(v is None or (isinstance(v, str) and v.strip() == "") for v in values)


def _read_table(
    ws: Worksheet,
    sheet: str,
    required: tuple[str, ...],
    optional: tuple[str, ...],
    issues: list[CheckIssue],
    *,
    dynamic_columns: bool = False,
    header_row: int = 1,
) -> Table:
    """Read a header-row data sheet into a Table.

    `dynamic_columns=True` (test_cases) accepts any column name; the
    canonical set is still checked for the required ones.
    """
    header_cells = ws[header_row] if ws.max_row >= header_row else []
    columns: dict[int, str] = {}
    seen: list[str] = []
    for c in header_cells:
        if c.value is None or str(c.value).strip() == "":
            continue
        name = str(c.value).strip()
        columns[c.column] = name
        seen.append(name)

    for req in required:
        if req not in seen:
            issues.append(
                issue(
                    "R-003",
                    "error",
                    f"Required column '{req}' is missing from the header row.",
                    sheet=sheet,
                    cell=f"A{header_row}",
                )
            )
    if not dynamic_columns:
        known = set(required) | set(optional) | {"notes"}
        extras = [n for n in seen if n not in known]
        if extras:
            issues.append(
                issue(
                    "R-202",
                    "warning",
                    f"Unknown column(s) {', '.join(sorted(extras))} — preserved but "
                    "not read by the platform.",
                    sheet=sheet,
                )
            )

    table = Table(sheet=sheet, columns=seen)
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = _row_values(ws, row_idx)
        if _is_blank_row(values):
            continue
        cells: dict[str, Cellv] = {}
        for col_idx, name in columns.items():
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            v2 = v.strip() if isinstance(v, str) else v
            cells[name] = Cellv(value=v2, cell=_ref(row_idx, col_idx))
        table.rows.append(Row(index=row_idx, cells=cells))
    return table


def _read_plan(ws: Worksheet, issues: list[CheckIssue]) -> dict[str, Cellv]:
    """The `plan` sheet is key-value: col A = field, col B = value.
    A header row (`field`/`value`) is allowed and skipped."""
    out: dict[str, Cellv] = {}
    for row_idx in range(1, ws.max_row + 1):
        key = ws.cell(row=row_idx, column=1).value
        if key is None or str(key).strip() == "":
            continue
        key_s = str(key).strip()
        if row_idx == 1 and key_s.lower() == "field":
            continue
        val = ws.cell(row=row_idx, column=2).value
        if isinstance(val, str):
            val = val.strip()
        out[key_s] = Cellv(value=val, cell=_ref(row_idx, 2))
    return out


def _read_factor_table(ws: Worksheet, sheet: str, issues: list[CheckIssue]) -> FactorTable:
    """ft.* layout (spec §4.5): key/value metadata block, one blank row,
    then the grid (1-D rows or a 2-D matrix)."""
    ft = FactorTable(sheet=sheet, slug=sheet[len("ft.") :])

    # 1) metadata block: rows until the first fully-blank row.
    row_idx = 1
    while row_idx <= ws.max_row:
        values = _row_values(ws, row_idx)
        if _is_blank_row(values):
            break
        key = ws.cell(row=row_idx, column=1).value
        if key is not None and str(key).strip() != "":
            val = ws.cell(row=row_idx, column=2).value
            if isinstance(val, str):
                val = val.strip()
            ft.meta[str(key).strip()] = Cellv(value=val, cell=_ref(row_idx, 2))
        row_idx += 1

    for req in FT_META_REQUIRED:
        if req not in ft.meta:
            issues.append(
                issue(
                    "R-100",
                    "error",
                    f"Factor-table metadata block is missing required key '{req}'.",
                    sheet=sheet,
                    cell="A1",
                )
            )

    # 2) skip blank separator rows.
    while row_idx <= ws.max_row and _is_blank_row(_row_values(ws, row_idx)):
        row_idx += 1
    if row_idx > ws.max_row:
        return ft  # empty grid — lint raises R-109.

    dimensionality = str(ft.meta_value("dimensionality") or "1d").strip().lower()

    if dimensionality == "2d":
        # Matrix: header row = col labels (first cell is the row\col signpost).
        header = ws[row_idx]
        for c in header[1:]:
            if c.value is None or (isinstance(c.value, str) and c.value.strip() == ""):
                continue
            v = c.value.strip() if isinstance(c.value, str) else c.value
            ft.col_labels.append(Cellv(value=v, cell=c.coordinate))
        for r in range(row_idx + 1, ws.max_row + 1):
            values = _row_values(ws, r)
            if _is_blank_row(values):
                continue
            row_label = ws.cell(row=r, column=1)
            if row_label.value is None:
                continue
            rl = row_label.value.strip() if isinstance(row_label.value, str) else row_label.value
            ft.row_labels.append(Cellv(value=rl, cell=row_label.coordinate))
            for offset, col_cv in enumerate(ft.col_labels, start=2):
                cell = ws.cell(row=r, column=offset)
                if cell.value is None or (
                    isinstance(cell.value, str) and cell.value.strip() == ""
                ):
                    continue  # blank = default / refuse; never zero.
                ft.grid.append(
                    FactorGridCell(
                        row_key=rl,
                        col_key=col_cv.value,
                        factor=cell.value,
                        cell=cell.coordinate,
                    )
                )
    else:
        # 1-D grid — reuse the table reader from the grid's header row.
        grid_table = _read_table(
            ws,
            sheet,
            ("level_id", "factor"),
            ("citation_rule", "citation_page"),
            issues,
            header_row=row_idx,
        )
        ft.rows_1d = grid_table.rows

    return ft


def parse_workbook(data: bytes) -> tuple[ParsedWorkbook, list[CheckIssue]]:
    """Bytes → (model, structural issues). Never raises on bad input:
    an unreadable file returns an empty model + R-001."""
    issues: list[CheckIssue] = []
    wb_model = ParsedWorkbook()

    try:
        wb = load_workbook(io.BytesIO(data), data_only=False, read_only=False)
    except Exception as exc:  # noqa: BLE001 — any load failure is R-001.
        issues.append(
            issue(
                "R-001",
                "error",
                f"Not a readable .xlsx workbook ({type(exc).__name__}). "
                "Export as OOXML .xlsx and drop it again.",
            )
        )
        return wb_model, issues

    wb_model.sheet_names = list(wb.sheetnames)
    present = set(wb.sheetnames)

    for required_sheet in REQUIRED_SHEETS:
        if required_sheet not in present:
            issues.append(
                issue(
                    MISSING_SHEET_RULE[required_sheet],
                    "error",
                    f"Required sheet '{required_sheet}' is missing.",
                )
            )

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        recognized = (
            sheet in REQUIRED_SHEETS
            or sheet in OPTIONAL_SHEETS
            or sheet.startswith(SHEET_PREFIXES)
        )
        if not recognized:
            wb_model.unknown_sheets.append(sheet)
            issues.append(
                issue(
                    "R-203",
                    "notice",
                    f"Sheet '{sheet}' is not a data sheet — ignored (README-style "
                    "prose sheets are fine).",
                    sheet=sheet,
                )
            )
            continue

        _scan_formulas(ws, sheet, issues)
        _scan_merged_cells(ws, sheet, issues)

        if sheet == "plan":
            wb_model.plan = _read_plan(ws, issues)
            wb_model.plan_sheet_present = True
        elif sheet == "test_cases":
            wb_model.test_cases = _read_table(
                ws, sheet, ("case_id",), (), issues, dynamic_columns=True
            )
        elif sheet.startswith("ft."):
            wb_model.factor_tables.append(_read_factor_table(ws, sheet, issues))
        elif sheet.startswith("geo."):
            geo = GeoSheet(sheet=sheet, slug=sheet[len("geo.") :])
            geo.table = _read_table(ws, sheet, *GEO_COLUMNS, issues)
            wb_model.geo_sheets.append(geo)
        else:
            required, optional = SHEET_COLUMNS[sheet]
            table = _read_table(ws, sheet, required, optional, issues)
            setattr(
                wb_model,
                {
                    "inputs": "inputs",
                    "dimensions": "dimensions",
                    "dimension_levels": "dimension_levels",
                    "chains": "chains",
                    "gates": "gates",
                    "modifiers": "modifiers",
                    "endorsements": "endorsements",
                    "loadings": "loadings",
                    "final_adjustments": "final_adjustments",
                    "outputs": "outputs",
                    "gaps_and_assumptions": "gaps",
                }[sheet],
                table,
            )

    return wb_model, issues
