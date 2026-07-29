"""Generate the ALL-CONSTRUCTS example bundle (Brief 92 Phase 92.5).

"Meridian Mutual — Shopfront BOP — Nebraska": a FICTIONAL carrier and
program with invented factors (the geography is real Nebraska ZIP
codes — public facts; everything priced is made up). Its job is to
exercise every spec-v1.0 construct the nonprofit bundle can't:

  · exposure-rated coverage towers (per $100 of limit / per $1,000 of
    sales) with the engine's own tower rounding (rate→3dp, premium→$1)
  · a 2-D factor table (construction × protection; "row::col" cells)
  · a geographic dimension at ZIP granularity with a `geo.*` detail
    sheet (18 ZIPs → 3 territories)
  · an interpolation=linear table (supported since registry r9 /
    Brief 95 C5: the 1-D curve INTERPOLATES between band lower
    bounds, clamped ends — the verifier interpolates too)
  · endorsements (a factor that FIRES on unsprinklered risks; an
    additive whose trigger never fires in the vectors)
  · a modifier schedule (filed structure; neutral without per-risk
    applications — exactly how the vectors run)
  · loadings via `applies_to` (×1.06 on every tower, pre-round)
  · a per-coverage CLAMP (liability minimum $100) + the package round
    with a $500 floor
  · compound gates (2- and 3-condition AND rows)

The generator is SELF-VERIFYING in the strict sense: it computes every
expected premium by MIRRORING THE ENGINE'S documented semantics
(stagesToRuntimePlan: exposure mode holds the LCM out of the chain
product, rounds the rate to 3dp and the tower premium to $1;
endorsement layering multiplies each tower tip; the sidecar sweep
applies loadings then clamps in stage order; the tail round sums,
floors, rounds once) — and asserts that every intended behavior
actually fires across the 8 vectors (the clamp, the endorsement, the
floor, all four tiers). If the engine and this file ever disagree,
the live golden test catches it loudly.

Run from api-lab/backend:

    uv run --with openpyxl python \
        ../../docs/specs/examples/meridian-shopfront-bop/generate_workbook.py
"""

from __future__ import annotations

import json
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent
OUT = HERE / "meridian_shopfront_bop.workbook.xlsx"
# Brief 2 P2 — real citations: every workbook row cites the rule + PAGE
# of the committed reference filing (generate_filing.py writes the
# sidecar; regenerate the filing BEFORE the workbook).
_FILING = json.loads((HERE / "meridian_filing_pages.json").read_text())
FILING_DOC = _FILING["document"]


def cite(key: str) -> tuple[str, str]:
    sec = _FILING["sections"][key]
    return (f"{FILING_DOC} Rule {sec['rule']}", str(sec["page"]))


CITE = FILING_DOC  # straggler fallback (should not appear in rows)

# ---------------------------------------------------------------------------
# The filed data — ALL INVENTED.
# ---------------------------------------------------------------------------

CLASSES = [
    # (id, label, prop_factor, liab_factor)
    ("c101", "Retail — general merchandise", 1.00, 1.00),
    ("c102", "Retail — apparel", 0.92, 0.95),
    ("c103", "Office — professional", 0.78, 0.85),
    ("c104", "Restaurant — limited cooking", 1.45, 1.30),
    ("c105", "Bakery", 1.32, 1.18),
    ("c106", "Barber / beauty shop", 0.88, 1.05),
    ("c107", "Florist", 0.95, 0.90),
    ("c108", "Hardware store", 1.12, 1.08),
    ("c109", "Pharmacy", 0.85, 1.02),
    ("c110", "Print shop", 1.18, 1.06),
    ("c111", "Woodworking — light", 1.85, 1.55),
    ("c112", "Welding supply", 2.10, 1.72),
    # Brief 2 P2 — the fuller mainstreet catalog (filing realism). All
    # invented; the worked examples reference only c101–c112, so adding
    # classes never moves an expected premium.
    ("c113", "Grocery — neighborhood", 1.08, 1.04),
    ("c114", "Delicatessen", 1.22, 1.12),
    ("c115", "Coffee shop — no frying", 1.15, 1.10),
    ("c116", "Ice cream shop", 1.05, 0.98),
    ("c117", "Bookstore", 0.82, 0.85),
    ("c118", "Gift shop", 0.90, 0.88),
    ("c119", "Jewelry store", 1.02, 0.92),
    ("c120", "Shoe store", 0.89, 0.90),
    ("c121", "Sporting goods", 1.04, 1.06),
    ("c122", "Toy store", 0.93, 0.95),
    ("c123", "Pet shop", 1.06, 1.08),
    ("c124", "Bicycle shop — sales & repair", 1.10, 1.09),
    ("c125", "Camera / electronics repair", 1.08, 1.00),
    ("c126", "Tailor / dressmaker", 0.84, 0.82),
    ("c127", "Dry cleaner — drop store only", 0.98, 0.96),
    ("c128", "Laundromat — self service", 1.24, 1.15),
    ("c129", "Picture framing", 0.94, 0.88),
    ("c130", "Music store — instruments", 0.91, 0.87),
    ("c131", "Art gallery", 0.86, 0.84),
    ("c132", "Antique store", 1.00, 0.93),
    ("c133", "Office — real estate", 0.76, 0.86),
    ("c134", "Office — insurance agency", 0.75, 0.84),
    ("c135", "Office — accounting", 0.74, 0.83),
    ("c136", "Medical office — no surgery", 0.96, 1.14),
    ("c137", "Dental office", 0.98, 1.12),
    ("c138", "Veterinary office — small animal", 1.14, 1.20),
    ("c139", "Photography studio", 0.88, 0.90),
    ("c140", "Mailbox / shipping store", 0.92, 0.94),
]

CONSTRUCTION = [("frame", "Frame"), ("jm", "Joisted masonry"),
                ("masonry", "Masonry"), ("fr", "Fire-resistive")]
PROTECTION = [("p1_4", "Protection 1–4"), ("p5_8", "Protection 5–8"),
              ("p9_10", "Protection 9–10")]

# 2-D: construction × protection (rows × cols).
CONSTR_X_PROT = {
    ("frame", "p1_4"): 1.05, ("frame", "p5_8"): 1.18, ("frame", "p9_10"): 1.42,
    ("jm", "p1_4"): 0.96, ("jm", "p5_8"): 1.07, ("jm", "p9_10"): 1.28,
    ("masonry", "p1_4"): 0.88, ("masonry", "p5_8"): 0.97, ("masonry", "p9_10"): 1.15,
    ("fr", "p1_4"): 0.72, ("fr", "p5_8"): 0.81, ("fr", "p9_10"): 0.94,
}

BUILDING_BANDS = [  # (id, label, lo, hi, ilf) — interpolation=linear (engine interpolates, r9)
    ("bl_0_100k", "$0–$100K", 0, 100_000, 1.00),
    ("bl_100_250k", "$100K–$250K", 100_000, 250_000, 0.93),
    ("bl_250_500k", "$250K–$500K", 250_000, 500_000, 0.87),
    ("bl_500k_1m", "$500K–$1M", 500_000, 1_000_000, 0.82),
    ("bl_1m_plus", "$1M+", 1_000_000, float("inf"), 0.78),
]

SALES_BANDS = [
    ("sb_0_250k", "$0–$250K", 0, 250_000, 0.90),
    ("sb_250k_1m", "$250K–$1M", 250_000, 1_000_000, 1.00),
    ("sb_1m_3m", "$1M–$3M", 1_000_000, 3_000_000, 1.12),
    ("sb_3m_plus", "$3M+", 3_000_000, float("inf"), 1.27),
]

SPRINKLER = [("true", "Sprinklered", 0.92), ("false", "Not sprinklered", 1.00)]

# Real Nebraska ZIP codes (public geography); the grouping + factors invented.
TERRITORIES = {
    "t1": (["68001", "68002", "68003", "68015", "68018", "68064"], 0.94, 1.02),
    "t2": (["68102", "68104", "68105", "68106", "68107", "68110"], 1.00, 1.00),
    "t3": (["68502", "68503", "68504", "68505", "68506", "68510"], 1.12, 0.95),
    # Brief 2 P2 — statewide coverage for filing realism (real NE ZIPs,
    # invented factors; no worked example rates in t4–t6).
    "t4": (["68005", "68046", "68123", "68128", "68147", "68157"], 0.97, 1.01),
    "t5": (["68801", "68803", "68845", "68847", "68901", "68902"], 0.91, 0.93),
    "t6": (["68025", "68601", "68701", "68702", "68776", "68787"], 0.89, 0.97),
}

BASE = {"building": 0.18, "bpp": 0.32, "liability": 1.15}
LCM = 1.42
LOADING = 1.06
ENDORSEMENT_FACTOR = 1.08  # equip breakdown — fires when NOT sprinklered
LIAB_MIN = 100.0
TOTAL_FLOOR = 500.0


def r3(x: float) -> float:
    return round_half_up(x, 3)


def round_half_up(x: float, decimals: int = 0) -> float:
    q = Decimal(1).scaleb(-decimals)
    return float(Decimal(str(x)).quantize(q, rounding=ROUND_HALF_UP))


def band_of(value: float, bands) -> tuple[str, float]:
    for level_id, _label, lo, hi, factor in bands:
        if lo <= value < hi:
            return level_id, factor
    raise AssertionError(f"{value} misses every band")


def interp_of(value: float, bands) -> float:
    """Brief 95 C5 (registry r9) — a 1-D `interpolation=linear` curve
    interpolates between band LOWER bounds, clamped at the ends —
    mirroring the engine's interpolate kind exactly."""
    pts = sorted((lo, factor) for _id, _lbl, lo, _hi, factor in bands)
    if value <= pts[0][0]:
        return pts[0][1]
    if value >= pts[-1][0]:
        return pts[-1][1]
    for (x0, y0), (x1, y1) in zip(pts, pts[1:]):
        if x0 <= value <= x1:
            return y0 + (value - x0) / (x1 - x0) * (y1 - y0)
    raise AssertionError(f"{value} misses every segment")


def territory_of(zip_code: str) -> tuple[str, float, float]:
    for code, (zips, prop_f, liab_f) in TERRITORIES.items():
        if zip_code in zips:
            return code, prop_f, liab_f
    raise AssertionError(f"{zip_code} in no territory")


# ---------------------------------------------------------------------------
# The engine-mirrored verifier.
# ---------------------------------------------------------------------------

def price(v: dict) -> dict:
    """Mirror the engine: per-tower round(rate3 × exposure × LCM), the
    endorsement layer, the loading sweep, the liability clamp, the
    floored package round."""
    class_prop = next(c[2] for c in CLASSES if c[0] == v["class_code"])
    class_liab = next(c[3] for c in CLASSES if c[0] == v["class_code"])
    cxp = CONSTR_X_PROT[(v["construction_class"], v["protection_class"])]
    bl_ilf = interp_of(v["building_limit"], BUILDING_BANDS)
    _sb_band, sb_ilf = band_of(v["annual_gross_sales"], SALES_BANDS)
    spr = 0.92 if v["sprinklered"] else 1.00
    _t, terr_prop, terr_liab = territory_of(v["zip"])

    towers: dict[str, float] = {}
    towers["building"] = round_half_up(
        r3(BASE["building"] * class_prop * cxp * bl_ilf * spr * terr_prop)
        * (v["building_limit"] / 100.0)
        * LCM
    )
    towers["bpp"] = round_half_up(
        r3(BASE["bpp"] * class_prop * cxp * spr * terr_prop)
        * (v["bpp_limit"] / 100.0)
        * LCM
    )
    towers["liability"] = round_half_up(
        r3(BASE["liability"] * class_liab * sb_ilf * terr_liab)
        * (v["annual_gross_sales"] / 1000.0)
        * LCM
    )

    fired = {"endorsement": False, "clamp": False, "floor": False}
    # Endorsement layer (equip breakdown ×1.08 when not sprinklered) —
    # multiplies each tower tip; the additive endorsement's trigger
    # (years_in_business < 1) never fires in these vectors.
    if not v["sprinklered"]:
        fired["endorsement"] = True
        towers = {k: t * ENDORSEMENT_FACTOR for k, t in towers.items()}
    # Sidecar sweep, stage order: loading then clamp.
    towers = {k: t * LOADING for k, t in towers.items()}
    if towers["liability"] < LIAB_MIN:
        fired["clamp"] = True
        towers["liability"] = LIAB_MIN
    total_raw = sum(towers.values())
    if total_raw < TOTAL_FLOOR:
        fired["floor"] = True
    total = round_half_up(max(total_raw, TOTAL_FLOOR))
    return {"towers": towers, "total": total, "fired": fired}


def tier_of(v: dict) -> str:
    if v["annual_gross_sales"] > 5_000_000 and v["years_in_business"] < 3:
        return "decline"
    if v["class_code"] in ("c111", "c112"):
        return "submit"
    if v["sprinklered"] and v["protection_class"] == "p1_4":
        return "preferred"
    return "standard"


VECTORS = [
    # Sweep: territories, sprinkler on/off (the endorsement), the liability
    # clamp, the package floor, all four tiers, every band boundary style.
    dict(case_id="mv_01", name="Retail t2, sprinklered, preferred",
         class_code="c101", building_limit=250_000, bpp_limit=80_000,
         annual_gross_sales=600_000, construction_class="jm",
         protection_class="p1_4", zip="68102", sprinklered=True,
         years_in_business=8),
    dict(case_id="mv_02", name="Restaurant t3, unsprinklered (endorsement fires)",
         class_code="c104", building_limit=400_000, bpp_limit=150_000,
         annual_gross_sales=1_200_000, construction_class="frame",
         protection_class="p5_8", zip="68502", sprinklered=False,
         years_in_business=12),
    dict(case_id="mv_03", name="Office t1, tiny liability (clamp fires)",
         class_code="c103", building_limit=180_000, bpp_limit=40_000,
         annual_gross_sales=30_000, construction_class="fr",
         protection_class="p1_4", zip="68001", sprinklered=False,
         years_in_business=20),
    dict(case_id="mv_04", name="Micro florist (package floor fires)",
         class_code="c107", building_limit=40_000, bpp_limit=10_000,
         annual_gross_sales=25_000, construction_class="masonry",
         protection_class="p5_8", zip="68002", sprinklered=True,
         years_in_business=6),
    dict(case_id="mv_05", name="Woodworking (submit tier)",
         class_code="c111", building_limit=350_000, bpp_limit=120_000,
         annual_gross_sales=900_000, construction_class="jm",
         protection_class="p9_10", zip="68104", sprinklered=True,
         years_in_business=15),
    dict(case_id="mv_06", name="Big young welding supply (decline tier)",
         class_code="c112", building_limit=900_000, bpp_limit=300_000,
         annual_gross_sales=6_000_000, construction_class="masonry",
         protection_class="p5_8", zip="68505", sprinklered=False,
         years_in_business=2),
    dict(case_id="mv_07", name="Pharmacy at a band edge ($250K exactly)",
         class_code="c109", building_limit=250_000, bpp_limit=250_000,
         annual_gross_sales=250_000, construction_class="fr",
         protection_class="p9_10", zip="68015", sprinklered=True,
         years_in_business=9),
    dict(case_id="mv_08", name="Hardware t3, defaults exercised implicitly",
         class_code="c108", building_limit=1_250_000, bpp_limit=400_000,
         annual_gross_sales=2_400_000, construction_class="frame",
         protection_class="p9_10", zip="68510", sprinklered=False,
         years_in_business=30),
]


# ---------------------------------------------------------------------------
# Workbook assembly.
# ---------------------------------------------------------------------------

def sheet_rows(ws, rows):
    for r in rows:
        ws.append(list(r))


def add_ft(wb, slug, display, dims, method, *, interpolation=None,
           rows_1d=None, matrix=None, cite_key=None):
    c_rule, c_page = cite(cite_key) if cite_key else (CITE, "")
    ws = wb.create_sheet(f"ft.{slug}")
    meta = [
        ("table_id", slug),
        ("display_name", display),
        ("dimensionality", "2d" if matrix else "1d"),
        ("row_dimension", dims[0]),
    ]
    if matrix:
        meta.append(("col_dimension", dims[1]))
    meta.append(("lookup_method", method))
    if interpolation:
        meta.append(("interpolation", interpolation))
    meta.append(("citation_rule", c_rule))
    meta.append(("citation_page", c_page))
    sheet_rows(ws, meta)
    ws.append([])
    if matrix:
        cols = sorted({c for (_r, c) in matrix})
        rows = sorted({r for (r, _c) in matrix})
        ws.append(["row\\col", *cols])
        for r in rows:
            ws.append([r, *[matrix[(r, c)] for c in cols]])
    else:
        ws.append(["level_id", "factor", "citation_rule", "citation_page"])
        for level_id, factor in rows_1d:
            ws.append([level_id, factor, c_rule, c_page])


def build() -> None:
    priced = [price(v) for v in VECTORS]
    tiers = [tier_of(v) for v in VECTORS]
    assert any(p["fired"]["endorsement"] for p in priced)
    assert any(p["fired"]["clamp"] for p in priced), "no vector exercises the clamp"
    assert any(p["fired"]["floor"] for p in priced), "no vector exercises the floor"
    assert set(tiers) == {"preferred", "standard", "submit", "decline"}

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    sheet_rows(ws, [
        ("Meridian Shopfront BOP — the ALL-CONSTRUCTS example (fictional).",),
        ("Exercises: exposure towers, 2-D tables, geo ZIP detail, linear-"
         "interpolation flag (engine steps), endorsements, modifiers, "
         "loadings, per-coverage clamp, floored package round, compound gates.",),
        ("Every factor is invented. Regenerate via generate_workbook.py "
         "(self-verifying against engine-mirrored math).",),
    ])

    ws = wb.create_sheet("plan")
    sheet_rows(ws, [
        ("field", "value"),
        ("spec_version", "1.0"),
        ("rating_plan_id", "meridian-shopfront-bop-ne-2026"),
        ("display_name", "Meridian Shopfront BOP — Nebraska"),
        ("version", "1.0.0"),
        ("carrier", "Meridian Mutual (fictional)"),
        ("product", "bop"),
        ("jurisdiction_country", "US"),
        ("state", "NE"),
        ("effective_date", "2026-09-01"),
        ("coverages", "building,bpp,liability"),
        ("filing_type", "other"),
        ("source_documents", FILING_DOC),
        ("description", "Synthetic all-constructs program for the ingestion "
         "golden path. Fictional carrier; invented factors."),
    ])

    ws = wb.create_sheet("inputs")
    ws.append(["name", "label", "data_type", "required", "allowed_values",
               "default_value", "unit", "maps_to_dimension", "description",
               "citation_rule", "citation_page"])
    sheet_rows(ws, [
        ("class_code", "Class code", "string", True, "", "", "",
         "class_code", "", *cite("classes")),
        ("building_limit", "Building limit", "currency", True, "", "", "USD",
         "building_limit_band", "", *cite("building_bands")),
        ("bpp_limit", "BPP limit", "currency", True, "", "", "USD",
         "", "", *cite("coverages")),
        ("annual_gross_sales", "Annual gross sales", "currency", True, "", "",
         "USD", "sales_band", "", *cite("sales_bands")),
        ("construction_class", "Construction", "enum", True,
         ",".join(c[0] for c in CONSTRUCTION), "", "",
         "construction_class", "", *cite("construction")),
        ("protection_class", "Protection class", "enum", True,
         ",".join(p[0] for p in PROTECTION), "", "",
         "protection_class", "", *cite("construction")),
        ("zip", "Location ZIP", "string", True, "", "", "",
         "territory", "", *cite("territory_zips")),
        ("sprinklered", "Sprinklered", "boolean", False, "", "false", "",
         "sprinklered_level", "Absent -> not sprinklered (manual default).",
         *cite("conventions")),
        ("years_in_business", "Years in business", "number", False, "", 5,
         "years", "", "Absent -> 5 (manual default).", *cite("conventions")),
    ])

    ws = wb.create_sheet("dimensions")
    ws.append(["slug", "display_name", "shape", "role", "data_type",
               "dimension_type", "geo_granularity", "geo_scope", "axes"])
    sheet_rows(ws, [
        ("class_code", "Class code", "categorical", "rating-input", "string",
         "standard", "", "", ""),
        ("construction_class", "Construction", "categorical", "rating-input",
         "string", "standard", "", "", ""),
        ("protection_class", "Protection class", "categorical", "rating-input",
         "string", "standard", "", "", ""),
        ("building_limit_band", "Building limit band", "banded", "rating-input",
         "currency", "standard", "", "", ""),
        ("sales_band", "Sales band", "banded", "rating-input", "currency",
         "standard", "", "", ""),
        ("sprinklered_level", "Sprinklered", "categorical", "rating-input",
         "string", "standard", "", "", ""),
        ("territory", "Territory", "geographic", "rating-input", "string",
         "geographic", "zip", "subset:NE", ""),
    ])

    ws = wb.create_sheet("dimension_levels")
    ws.append(["dimension_slug", "kind", "level_id", "label", "aliases",
               "min", "max", "territory_ref", "citation_rule", "citation_page"])
    for cid, label, _p, _l in CLASSES:
        ws.append(["class_code", "categorical", cid, label, cid[1:], "", "",
                   "", *cite("classes")])
    # Book-intake §4 — the fixture DEMONSTRATES the alias vocabulary
    # (spec §4.4): a book that says "joisted masonry" or "Y"/"N"
    # rates out of the box, and the seeded plan teaches the column.
    construction_aliases = {"jm": "joisted masonry"}
    for cid, label in CONSTRUCTION:
        ws.append(["construction_class", "categorical", cid, label,
                   construction_aliases.get(cid, ""), "",
                   "", "", *cite("construction")])
    for pid, label in PROTECTION:
        ws.append(["protection_class", "categorical", pid, label, "", "",
                   "", "", *cite("construction")])
    for bid, label, lo, hi, _f in BUILDING_BANDS:
        ws.append(["building_limit_band", "banded", bid, label, "", lo,
                   "+inf" if hi == float("inf") else hi, "", *cite("building_bands")])
    for sid, label, lo, hi, _f in SALES_BANDS:
        ws.append(["sales_band", "banded", sid, label, "", lo,
                   "+inf" if hi == float("inf") else hi, "", *cite("sales_bands")])
    sprinkler_aliases = {"true": "Y,yes", "false": "N,no"}
    for sid, label, _f in SPRINKLER:
        ws.append(["sprinklered_level", "categorical", sid, label,
                   sprinkler_aliases.get(sid, ""), "",
                   "", "", *cite("sprinkler")])
    for code in TERRITORIES:
        ws.append(["territory", "geographic", code, f"Territory {code[1:]}",
                   "", "", "", code, *cite("territory_zips")])

    ws = wb.create_sheet("geo.territory")
    ws.append(["zip", "territory_code", "citation_rule", "citation_page"])
    for code, (zips, _p, _l) in TERRITORIES.items():
        for z in zips:
            ws.append([z, code, *cite("territory_zips")])

    add_ft(wb, "class_rate_prop", "Class factor (property)", ["class_code"],
           "direct", rows_1d=[(c[0], c[2]) for c in CLASSES], cite_key="classes")
    add_ft(wb, "class_rate_liab", "Class factor (liability)", ["class_code"],
           "direct", rows_1d=[(c[0], c[3]) for c in CLASSES], cite_key="classes")
    add_ft(wb, "constr_x_prot", "Construction × Protection",
           ["construction_class", "protection_class"], "direct",
           matrix=CONSTR_X_PROT, cite_key="construction")
    add_ft(wb, "building_ilf", "Building limit ILF", ["building_limit_band"],
           "binned", interpolation="linear",
           rows_1d=[(b[0], b[4]) for b in BUILDING_BANDS], cite_key="building_bands")
    add_ft(wb, "liab_ilf", "Liability sales ILF", ["sales_band"], "binned",
           rows_1d=[(s[0], s[4]) for s in SALES_BANDS], cite_key="sales_bands")
    add_ft(wb, "sprinkler_prop", "Sprinkler credit (property)",
           ["sprinklered_level"], "direct",
           rows_1d=[(s[0], s[2]) for s in SPRINKLER], cite_key="sprinkler")
    add_ft(wb, "territory_prop", "Territory factor (property)", ["territory"],
           "direct", rows_1d=[(c, t[1]) for c, t in TERRITORIES.items()], cite_key="territory_factors")
    add_ft(wb, "territory_liab", "Territory factor (liability)", ["territory"],
           "direct", rows_1d=[(c, t[2]) for c, t in TERRITORIES.items()], cite_key="territory_factors")

    ws = wb.create_sheet("chains")
    ws.append(["coverage", "order", "stage_kind", "stage_id", "factor_table",
               "dimension", "input_binding", "value", "exposure_divisor",
               "citation_rule", "citation_page"])
    def tower(cov, base, lookups, exposure_input, divisor):
        rows = [(cov, 0, "base", f"{cov}_base", "", "", f"literal:{base}",
                 base, "", *cite("base_rates"))]
        for i, (kind, table, dim) in enumerate(lookups, start=1):
            rows.append((cov, i, kind, f"{cov}_{table}", f"ft.{table}", dim,
                         "", "", "", *cite("rating_order")))
        rows.append((cov, len(lookups) + 1, "exposure", f"{cov}_exposure", "",
                     "", exposure_input, "", divisor, "", ""))
        rows.append((cov, len(lookups) + 2, "lcm", f"{cov}_lcm", "", "", "",
                     LCM, "", *cite("base_rates")))
        return rows

    sheet_rows(ws, tower("building", BASE["building"], [
        ("lookup.direct", "class_rate_prop", "class_code"),
        ("lookup.multi", "constr_x_prot", "construction_class"),
        ("lookup.range", "building_ilf", "building_limit_band"),
        ("lookup.direct", "sprinkler_prop", "sprinklered_level"),
        ("lookup.direct", "territory_prop", "territory"),
    ], "form_input.building_limit", 100))
    sheet_rows(ws, tower("bpp", BASE["bpp"], [
        ("lookup.direct", "class_rate_prop", "class_code"),
        ("lookup.multi", "constr_x_prot", "construction_class"),
        ("lookup.direct", "sprinkler_prop", "sprinklered_level"),
        ("lookup.direct", "territory_prop", "territory"),
    ], "form_input.bpp_limit", 100))
    sheet_rows(ws, tower("liability", BASE["liability"], [
        ("lookup.direct", "class_rate_liab", "class_code"),
        ("lookup.range", "liab_ilf", "sales_band"),
        ("lookup.direct", "territory_liab", "territory"),
    ], "form_input.annual_gross_sales", 1000))

    ws = wb.create_sheet("gates")
    ws.append(["order", "rule_id", "variable", "op", "value",
               "variable_2", "op_2", "value_2", "variable_3", "op_3", "value_3",
               "tier", "reasoning", "citation_rule", "citation_page"])
    sheet_rows(ws, [
        (1, "decline_big_young", "annual_gross_sales", "gt", 5_000_000,
         "years_in_business", "lt", 3, "", "", "",
         "decline", "Large + young — outside appetite.", *cite("eligibility")),
        (2, "submit_heavy_classes", "class_code", "in", "c111,c112",
         "", "", "", "", "", "",
         "submit", "Heavier classes route to an underwriter.", *cite("eligibility")),
        (3, "preferred_protected", "sprinklered", "eq", True,
         "protection_class", "in", "p1_4", "", "", "",
         "preferred", "Sprinklered + best protection.", *cite("eligibility")),
        (99, "__default__", "", "", "", "", "", "", "", "", "",
         "standard", "Standard appetite.", *cite("eligibility")),
    ])

    ws = wb.create_sheet("modifiers")
    ws.append(["schedule_id", "schedule_name", "scope", "total_cap_pct",
               "category_id", "category_name", "range_pct", "tier_filter",
               "citation_rule", "citation_page"])
    sheet_rows(ws, [
        ("shopfront_irpm", "Shopfront IRPM", "package", 25,
         "management", "Management experience", 10, "", *cite("modifiers")),
        ("shopfront_irpm", "Shopfront IRPM", "package", 25,
         "premises", "Premises condition", 15, "", *cite("modifiers")),
    ])

    ws = wb.create_sheet("endorsements")
    ws.append(["endorsement_id", "kind", "form_number", "display_name",
               "factor", "amount", "coverage", "sublimit", "trigger",
               "citation_rule", "citation_page"])
    sheet_rows(ws, [
        ("equip_breakdown", "factor", "MS 10 01", "Equipment breakdown",
         ENDORSEMENT_FACTOR, "", "", "",
         "form_input.sprinklered == false", *cite("endorsement_equip")),
        # A second FACTOR endorsement whose trigger never fires in the
        # vectors — exercises the attached=false path. (An additive here
        # would be registry-refused on this 3-coverage plan:
        # endorsement_additive_multi_coverage.)
        ("new_venture_surcharge", "factor", "MS 20 07", "New-venture surcharge",
         1.15, "", "", "", "form_input.years_in_business < 1",
         *cite("endorsement_venture")),
    ])

    ws = wb.create_sheet("loadings")
    ws.append(["loading_id", "factor_kind", "display_name", "factor",
               "applies_to", "predicate", "citation_rule", "citation_page"])
    sheet_rows(ws, [
        ("expense_load", "expense", "Expense loading", LOADING,
         "building,bpp,liability", "", *cite("base_rates")),
    ])

    ws = wb.create_sheet("final_adjustments")
    ws.append(["adjustment_id", "kind", "order", "applies_to", "min_value",
               "max_value", "round_increment", "round_min",
               "citation_rule", "citation_page"])
    sheet_rows(ws, [
        ("liab_min", "clamp", 1, "liability", LIAB_MIN, "", "", "",
         *cite("rounding")),
        ("round_total", "round", 2, "", "", "", 1, TOTAL_FLOOR,
         *cite("rounding")),
    ])

    ws = wb.create_sheet("outputs")
    ws.append(["output_id", "field_name", "display_name", "source",
               "description"])
    sheet_rows(ws, [
        ("out_building", "building_premium", "Building premium",
         "building_lcm", "Tower tip (loaded; unrounded past the tower)."),
        ("out_bpp", "bpp_premium", "BPP premium", "bpp_lcm", ""),
        ("out_liability", "liability_premium", "Liability premium",
         "liability_lcm", ""),
        ("out_total", "total_premium", "Total premium", "round_total",
         "Floored at $500, rounded to $1."),
    ])

    ws = wb.create_sheet("test_cases")
    ws.append(["case_id", "name", "class_code", "building_limit", "bpp_limit",
               "annual_gross_sales", "construction_class", "protection_class",
               "zip", "sprinklered", "years_in_business",
               "expected_building_premium", "expected_bpp_premium",
               "expected_liability_premium", "expected_total_premium",
               "expected_tier",
               "tolerance_building_premium", "tolerance_bpp_premium",
               "tolerance_liability_premium"])
    for v, p, tier in zip(VECTORS, priced, tiers):
        ws.append([
            v["case_id"], v["name"], v["class_code"], v["building_limit"],
            v["bpp_limit"], v["annual_gross_sales"], v["construction_class"],
            v["protection_class"], v["zip"], v["sprinklered"],
            v["years_in_business"],
            round(p["towers"]["building"], 4), round(p["towers"]["bpp"], 4),
            round(p["towers"]["liability"], 4), p["total"], tier,
            0.02, 0.02, 0.02,
        ])

    ws = wb.create_sheet("gaps_and_assumptions")
    ws.append(["kind", "description", "citation_rule", "citation_page",
               "impact", "related"])
    sheet_rows(ws, [
        ("assumption",
         "The manual's sprinkler question defaults to 'not sprinklered' when "
         "unanswered; transcribed as inputs.sprinklered default_value=false.",
         *cite("conventions"),
         "Risks without the answer rate unsprinklered (and the equipment-"
         "breakdown endorsement attaches).", "inputs!sprinklered"),
        ("assumption",
         "Years-in-business defaults to 5 when absent (manual convention).",
         *cite("conventions"),
         "Absent tenure neither declines (needs <3 AND >$5M sales) nor "
         "attaches the new-venture surcharge (<1).", "inputs!years_in_business"),
    ])

    wb.save(OUT)
    fired = {
        "endorsement": sum(p["fired"]["endorsement"] for p in priced),
        "clamp": sum(p["fired"]["clamp"] for p in priced),
        "floor": sum(p["fired"]["floor"] for p in priced),
    }
    print(f"verified {len(VECTORS)} vectors (tiers {sorted(set(tiers))}; "
          f"fired: {fired}) — wrote {OUT.name}")
    print(f"sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    build()
