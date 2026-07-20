"""Generate the starter-kit template workbook.

The template IS the spec §9 mini example — a complete, conformant,
*working* program (one categorical + one banded dimension, a 1-D and a
2-D factor table, a 5-stage chain, one verified test vector) — with a
teaching `notes` column on every data sheet and a README sheet that
explains the loop. It is not a form: it checks clean, builds, and its
vector verifies, so the user hollows out a living program instead of
filling in a dead one.

SELF-VERIFYING: before writing, the generator recomputes TC-1's premium
from the factors it is about to write (base × construction × constr×age
× TIV/100 × LCM) and asserts the spec §9 stated expectation ($390.00).
A mismatch aborts — the template is never silently wrong.

Run from the repository root:

    uv run --project server python \
        docs/specs/examples/template/generate_template.py

Output: docs/specs/examples/template/openrater_workbook_template.xlsx
(The packaged copy under
`server/src/openrater/rates/ingest/assets/` is pinned byte-identical
by CI; copy the regenerated file there too.)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent
OUT = HERE / "openrater_workbook_template.xlsx"

# ── The §9 mini program, verbatim ──────────────────────────────────────────

BASE = 0.150
CONSTRUCTION = {"frame": 1.00, "fire_resistive": 0.78}
CONSTR_X_AGE = {
    ("frame", "age_0_25"): 1.00,
    ("frame", "age_25_plus"): 1.20,
    ("fire_resistive", "age_0_25"): 0.95,
    ("fire_resistive", "age_25_plus"): 1.05,
}
EXPOSURE_DIVISOR = 100
LCM = 1.30

TC1 = {"construction_class": "frame", "building_age": 10, "tiv": 200_000}
TC1_EXPECTED = 390.00


def verify() -> None:
    age_band = "age_0_25" if TC1["building_age"] < 25 else "age_25_plus"
    rate = (
        BASE
        * CONSTRUCTION[TC1["construction_class"]]
        * CONSTR_X_AGE[(TC1["construction_class"], age_band)]
    )
    premium = rate * (TC1["tiv"] / EXPOSURE_DIVISOR) * LCM
    assert abs(premium - TC1_EXPECTED) < 0.005, (
        f"template math drifted: computed {premium:.2f}, spec §9 says "
        f"{TC1_EXPECTED:.2f} — fix the factors before writing anything"
    )


# ── Sheets ──────────────────────────────────────────────────────────────────

README_LINES = [
    "OpenRater — rating workbook TEMPLATE (filing-transcription-spec v1.0)",
    "",
    "This is a complete, working mini program — the spec's own §9 example —",
    "not an empty form. It checks clean, builds, and its one test case",
    "verifies at $390.00. Replace its contents with your filing's, sheet by",
    "sheet, keeping every header row exactly as-is.",
    "",
    "The loop:  transcribe → CHECK → fix what it names → re-check → build.",
    "Check without creating anything: the 'Build from a workbook' panel,",
    "POST /api/v1/plans/ingest/check, or",
    "`python -m openrater.rates.ingest check <file>`.",
    "",
    "Every `notes` column is teaching text — never read by the platform;",
    "delete or keep them freely. This README sheet is ignored too (R-203).",
    "",
    "The full grammar: docs/specs/filing-transcription-spec.md (also in the",
    "starter kit next to this file). Hand the spec to your AI with the",
    "filing; drop the workbook it produces.",
]


def build() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "README"
    for line in README_LINES:
        ws.append([line])

    ws = wb.create_sheet("plan")
    ws.append(["field", "value", "notes"])
    for field, value, note in (
        ("spec_version", "1.0", "always 1.0 for this spec"),
        ("rating_plan_id", "mini-bop-demo-il-2026", "slug: {carrier}-{product}-{state}-{year}"),
        ("display_name", "Mini BOP demo — Illinois — 2026", ""),
        ("version", "1.0.0", "bump when re-issuing a corrected workbook"),
        ("carrier", "Demo Mutual", "as filed"),
        ("product", "bop", "bop|cgl|do|eo|wc|auto|umbrella|excess|marine|inland_marine|other"),
        ("jurisdiction_country", "US", ""),
        ("state", "IL", "2-letter; empty = multistate"),
        ("effective_date", "2026-01-01", "YYYY-MM-DD"),
        ("coverages", "building", "CSV; each gets a chains block"),
        ("filing_type", "new", "new|revision|adoption|other"),
        ("source_documents", "mini_bop_manual.pdf", "every PDF consumed"),
    ):
        ws.append([field, value, note])

    ws = wb.create_sheet("inputs")
    ws.append(["name", "label", "data_type", "required", "allowed_values",
               "default_value", "unit", "citation_rule", "citation_page", "notes"])
    ws.append(["tiv", "Total insured value", "currency", True, "", "", "USD",
               "Rule 2.B", "p.9", "consumed by the exposure stage"])
    ws.append(["construction_class", "Construction class", "enum", True,
               "frame,fire_resistive", "", "", "Table 5.A", "p.51",
               "enum needs allowed_values"])
    ws.append(["building_age", "Building age", "number", True, "", "", "years",
               "Rule 2.C", "p.10",
               "a default_value here is the ONLY sanctioned missing-data softener"])

    ws = wb.create_sheet("dimensions")
    ws.append(["slug", "display_name", "shape", "role", "data_type",
               "dimension_type", "geo_granularity", "geo_scope", "axes", "notes"])
    ws.append(["construction_class", "Construction class", "categorical", "both",
               "enum", "standard", "", "", "", "levels live in dimension_levels"])
    ws.append(["building_age", "Building age (yrs)", "banded", "rating-input",
               "number", "standard", "", "", "",
               "banded = the curve form; bands in dimension_levels"])

    ws = wb.create_sheet("dimension_levels")
    ws.append(["dimension_slug", "kind", "level_id", "label", "aliases",
               "min", "max", "territory_ref", "notes"])
    ws.append(["construction_class", "categorical", "frame", "Frame (ISO 1)",
               "wood,iso1", "", "", "", "put the filing's own codes in aliases"])
    ws.append(["construction_class", "categorical", "fire_resistive",
               "Fire-resistive (ISO 6)", "fr,iso6", "", "", "", ""])
    ws.append(["building_age", "banded", "age_0_25", "0-25 yrs", "", 0, 25, "",
               "half-open [min, max); contiguous, gap-free"])
    ws.append(["building_age", "banded", "age_25_plus", "25+ yrs", "", 25,
               "+inf", "", "last band may close at +inf"])

    ws = wb.create_sheet("ft.construction_class")
    for row in (
        ("table_id", "construction_class"),
        ("display_name", "Construction factor"),
        ("dimensionality", "1d"),
        ("row_dimension", "construction_class"),
        ("lookup_method", "direct"),
        ("citation_rule", "Table 5.A"),
        ("citation_page", "p.51"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["level_id", "factor", "citation_rule", "citation_page", "notes"])
    ws.append(["frame", CONSTRUCTION["frame"], "Table 5.A", "p.51",
               "every key must be a declared level (or __default__)"])
    ws.append(["fire_resistive", CONSTRUCTION["fire_resistive"], "Table 5.A",
               "p.51", ""])

    ws = wb.create_sheet("ft.constr_x_age")
    for row in (
        ("table_id", "constr_x_age"),
        ("display_name", "Construction x Age factor"),
        ("dimensionality", "2d"),
        ("row_dimension", "construction_class"),
        ("col_dimension", "building_age"),
        ("lookup_method", "direct"),
        ("citation_rule", "Table 5.B"),
        ("citation_page", "p.52"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["row\\col", "age_0_25", "age_25_plus"])
    ws.append(["frame",
               CONSTR_X_AGE[("frame", "age_0_25")],
               CONSTR_X_AGE[("frame", "age_25_plus")]])
    ws.append(["fire_resistive",
               CONSTR_X_AGE[("fire_resistive", "age_0_25")],
               CONSTR_X_AGE[("fire_resistive", "age_25_plus")]])

    ws = wb.create_sheet("chains")
    ws.append(["coverage", "order", "stage_kind", "stage_id", "factor_table",
               "dimension", "input_binding", "value", "exposure_divisor",
               "predicate", "citation_rule", "citation_page", "notes"])
    ws.append(["building", 0, "base", "bld_base", "", "", f"literal:{BASE}",
               BASE, "", "", "Table 4.A", "p.40", "every block starts with base"])
    ws.append(["building", 1, "lookup.direct", "bld_constr",
               "ft.construction_class", "construction_class", "", "", "", "",
               "Table 5.A", "p.51", "multiplies"])
    ws.append(["building", 2, "lookup.multi", "bld_constr_age", "ft.constr_x_age",
               "construction_class", "", "", "", "", "Table 5.B", "p.52",
               "2-D lookup; dimension = the row dimension"])
    ws.append(["building", 3, "exposure", "bld_exposure", "", "",
               "form_input.tiv", "", EXPOSURE_DIVISOR, "", "", "",
               "divide by exposure units (per $100 here)"])
    ws.append(["building", 4, "lcm", "bld_lcm", "", "", "", LCM, "", "",
               "(carrier-set)", "", "loss-cost multiplier, last"])

    ws = wb.create_sheet("outputs")
    ws.append(["output_id", "field_name", "display_name", "source", "notes"])
    ws.append(["out_building", "building_premium", "Building premium",
               "bld_exposure", "source = the stage that feeds it"])

    ws = wb.create_sheet("test_cases")
    ws.append(["case_id", "name", "construction_class", "building_age", "tiv",
               "expected_building_premium", "notes"])
    ws.append(["tc_1", "Frame, 10 yrs, $200k (spec §9 worked example)",
               TC1["construction_class"], TC1["building_age"], TC1["tiv"],
               TC1_EXPECTED,
               "the filing's own worked examples go here, verbatim"])

    ws = wb.create_sheet("gaps_and_assumptions")
    ws.append(["kind", "description", "citation_rule", "citation_page",
               "impact", "related", "notes"])
    # Header-only: this mini program transcribes completely. The sheet must
    # exist even when empty — declaring "none" is a conscious act (§4.14).

    return wb


def main() -> None:
    verify()
    wb = build()
    wb.properties.creator = "OpenRater"
    wb.properties.title = "OpenRater rating workbook template (spec v1.0)"
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(HERE.parents[3])}")


if __name__ == "__main__":
    main()
