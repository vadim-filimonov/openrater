"""Generate the canonical filing-transcription-spec v1.0 example bundle.

Rebuilds the nonprofit D&O + GL rating program
(docs/rating-algorithms/nonprofit_990_do_gl_rating_v1.xlsx, a synthetic
pre-filing design) as a spec-conformant workbook — the worked example the
spec, the profiles, and (in Brief 92 Phase 92.3) the CI golden path all
point at.

The generator is SELF-VERIFYING: before writing anything it re-derives
every band and recomputes every expected premium + tier for all 20 test
cases from docs/rating-algorithms/nonprofit_990_test_cases.csv, and
asserts them against the source's prebinned CSV and stated expectations.
A single mismatch aborts the run — the canonical artifact is never
silently wrong.

Run from the repo root:

    cd api-lab/backend && uv run --with openpyxl python \
        ../../docs/specs/examples/nonprofit-do-gl/generate_workbook.py

Output: docs/specs/examples/nonprofit-do-gl/nonprofit_do_gl.workbook.xlsx
"""

from __future__ import annotations

import csv
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[3]
SRC = REPO / "docs" / "rating-algorithms"
OUT = HERE / "nonprofit_do_gl.workbook.xlsx"

# ---------------------------------------------------------------------------
# The filed data, verbatim from the source workbook (dump verified 2026-07-14:
# no formulas; every value literal).
# ---------------------------------------------------------------------------

NTEE = [
    # (level_id, label, do_factor, gl_factor, rationale)
    ("arts_culture", "Arts, Culture & Humanities", 1.0, 1.05, "Low governance risk; venue/event GL exposure"),
    ("education", "Education", 1.1, 1.15, "Board scrutiny; campus/premises exposure"),
    ("environment", "Environment", 1.0, 1.1, "Some field-work GL exposure"),
    ("animal", "Animal-Related", 0.95, 1.25, "Low D&O; bite/handling drives GL"),
    ("health_care", "Health Care", 1.3, 1.25, "High fiduciary + clinical premises exposure"),
    ("mental_health", "Mental Health & Crisis Intervention", 1.15, 1.15, "Vicarious + premises exposure"),
    ("diseases", "Diseases & Medical Disciplines", 1.05, 1.1, "Moderate fiduciary + outreach exposure"),
    ("medical_research", "Medical Research", 1.05, 1.05, "Moderate IRB + lab premises exposure"),
    ("crime_legal", "Crime & Legal-Related", 1.2, 1.05, "Litigation propensity high; minimal premises"),
    ("employment", "Employment", 1.15, 1.0, "EPLI/board exposure"),
    ("food_agriculture", "Food, Agriculture & Nutrition", 1.05, 1.2, "Foodservice handling drives GL"),
    ("housing_shelter", "Housing & Shelter", 1.15, 1.3, "Heavy premises + tenant exposure"),
    ("public_safety", "Public Safety & Disaster Relief", 1.05, 1.1, "Volunteer/field GL exposure"),
    ("recreation", "Recreation & Sports", 1.05, 1.4, "Participant-injury GL exposure dominates"),
    ("youth_development", "Youth Development", 1.2, 1.3, "Abuse-claim exposure flows to D&O AND GL"),
    ("human_services", "Human Services", 1.15, 1.2, "Service delivery + duty-of-care exposure"),
    ("international", "International / Foreign Affairs", 1.2, 1.1, "FCPA/sanctions + travel exposure"),
    ("civil_rights", "Civil Rights / Advocacy", 1.2, 1.0, "Political/defamation exposure"),
    ("community_improvement", "Community Improvement", 1.05, 1.05, "Low-moderate across both"),
    ("philanthropy", "Philanthropy & Foundations", 0.85, 0.8, "Well-governed; low operations"),
    ("science_tech", "Science & Technology", 0.95, 0.95, "Low-moderate across both"),
    ("social_science", "Social Science", 0.95, 0.9, "Low premises exposure"),
    ("public_societal", "Public & Societal Benefit", 1.0, 1.0, "Baseline"),
    ("religion", "Religion-Related", 1.2, 1.15, "Congregational disputes + abuse-claim exposure"),
    ("mutual_membership", "Mutual/Membership Benefit", 1.05, 1.05, "Moderate"),
    ("unknown_unclassified", "Unknown / Unclassified", 1.5, 1.5, "Refer-to-underwriter load"),
    ("unknown_no_ntee", "Unknown / no NTEE code", 1.5, 1.5, "Refer-to-underwriter load"),
]

REVENUE_BANDS = [
    # (level_id, label, lo, hi, do, gl)
    ("01_under_25k", "01. <$25K", 0, 25_000, 0.65, 0.5),
    ("02_25k_50k", "02. $25K-$50K", 25_000, 50_000, 0.75, 0.7),
    ("03_50k_100k", "03. $50K-$100K", 50_000, 100_000, 0.85, 0.85),
    ("04_100k_250k", "04. $100K-$250K", 100_000, 250_000, 1.0, 1.0),
    ("05_250k_500k", "05. $250K-$500K", 250_000, 500_000, 1.15, 1.2),
    ("06_500k_1m", "06. $500K-$1M", 500_000, 1_000_000, 1.35, 1.5),
    ("07_1m_5m", "07. $1M-$5M", 1_000_000, 5_000_000, 1.75, 2.1),
]

STATE_TIERS = {  # tier -> (do, gl)
    "T1": (0.85, 0.8),
    "T2": (0.95, 0.9),
    "T3": (1.0, 1.0),
    "T4": (1.15, 1.2),
    "T5": (1.3, 1.35),
}
STATES_BY_TIER = {
    "T1": ["AL", "AR", "IA", "ID", "KS", "KY", "ME", "MS", "MT", "ND", "NE", "NH", "OK", "SD", "VT", "WV", "WY"],
    "T2": ["IN", "MN", "MO", "NM", "NV", "OH", "SC", "TN", "UT", "WI"],
    "T3": ["AK", "AZ", "CO", "CT", "DC", "DE", "GA", "HI", "MA", "MD", "MI", "NC", "OR", "PA", "RI", "VA", "WA"],
    "T4": ["FL", "IL", "LA", "NJ", "TX"],
    "T5": ["CA", "NY"],
}
STATE_NAMES = {
    "AK": "Alaska", "AL": "Alabama", "AR": "Arkansas", "AZ": "Arizona", "CA": "California",
    "CO": "Colorado", "CT": "Connecticut", "DC": "District of Columbia", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "IA": "Iowa", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana",
    "MA": "Massachusetts", "MD": "Maryland", "ME": "Maine", "MI": "Michigan", "MN": "Minnesota",
    "MO": "Missouri", "MS": "Mississippi", "MT": "Montana", "NC": "North Carolina",
    "ND": "North Dakota", "NE": "Nebraska", "NH": "New Hampshire", "NJ": "New Jersey",
    "NM": "New Mexico", "NV": "Nevada", "NY": "New York", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VA": "Virginia",
    "VT": "Vermont", "WA": "Washington", "WI": "Wisconsin", "WV": "West Virginia", "WY": "Wyoming",
}
STATE_TIER = {code: tier for tier, codes in STATES_BY_TIER.items() for code in codes}
assert len(STATE_TIER) == 51

SUBSECTION = [
    ("501c3", "501(c)(3) charitable / religious / educational", 1.0, 1.0, "Base - 87% of sample"),
    ("501c4", "501(c)(4) social welfare", 1.15, 1.0, "Lobbying exposure for D&O"),
    ("501c6", "501(c)(6) business league / chamber", 1.05, 0.95, "Membership disputes; lighter premises"),
    ("501c7", "501(c)(7) social / recreational club", 0.9, 1.2, "Lower D&O, higher GL (premises)"),
    ("501c8", "501(c)(8) fraternal benefit society", 1.1, 1.05, "Member services"),
    ("501c_other", "501(c) other", 1.1, 1.05, "Refer if unknown"),
]

PAYROLL_BANDS = [  # (level_id, label, lo, hi, do) over employee_count; __default__ 1.1
    ("00_none", "0 employees", 0, 1, 0.85),
    ("01_micro", "1-5 employees", 1, 6, 0.95),
    ("02_small", "6-25 employees", 6, 26, 1.0),
    ("03_mid", "26-100 employees", 26, 101, 1.15),
    ("04_large", "100+ employees", 101, float("inf"), 1.4),
]
PAYROLL_DEFAULT = 1.1  # the filed "99_unknown" load

STRESS_BANDS = [  # over expense_ratio; __default__ 1.1
    ("01_under_85", "<0.85 (healthy surplus)", 0, 0.85, 0.9),
    ("02_85_100", "0.85-1.00", 0.85, 1.0, 1.0),
    ("03_100_115", "1.00-1.15 (deficit)", 1.0, 1.15, 1.2),
    ("04_over_115", ">1.15", 1.15, float("inf"), 1.5),
]
STRESS_DEFAULT = 1.1

OCCUPANCY_BANDS = [  # over occupancy_ratio; __default__ 1.15
    ("01_under_03", "<3%", 0, 0.03, 0.85),
    ("02_03_06", "3-6%", 0.03, 0.06, 1.0),
    ("03_06_10", "6-10%", 0.06, 0.10, 1.15),
    ("04_10_20", "10-20%", 0.10, 0.20, 1.35),
    ("05_over_20", ">20%", 0.20, float("inf"), 1.6),
]
OCCUPANCY_DEFAULT = 1.15

DO_BASE, GL_BASE, LCM = 600, 300, 1.35
CITE = "Nonprofit 990 D&O+GL rating v1 (source workbook)"


def band_of(value: float | None, bands) -> str | None:
    if value is None:
        return None
    for level_id, _label, lo, hi, *_ in bands:
        if lo <= value < hi:
            return level_id
    return None


def factor_of(level: str | None, bands, default: float) -> float:
    if level is None:
        return default
    for level_id, _label, _lo, _hi, f in bands:
        if level_id == level:
            return f
    return default


def round_half_up(x: float) -> int:
    return int(Decimal(str(x)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def derive_tier(ntee: str, revenue: float, ratio: float | None, _employees: float | None) -> str:
    """The gates sheet, as authored below, evaluated first-match-wins."""
    if ratio is not None and ratio > 1.15 and revenue < 50_000:
        return "Decline"
    if ntee in ("unknown_unclassified", "unknown_no_ntee"):
        return "Submit"
    if ratio is not None and ratio > 1.15:
        return "Submit"
    if revenue >= 5_000_000:
        return "Submit"
    if (
        ratio is not None
        and ratio < 0.85
        and ntee in ("philanthropy", "science_tech", "social_science")
        and revenue >= 250_000
    ):
        return "Preferred"
    return "Standard"


# ---------------------------------------------------------------------------
# Verification pass — abort on any mismatch.
# ---------------------------------------------------------------------------

def load_cases() -> list[dict]:
    with open(SRC / "nonprofit_990_test_cases.csv", newline="") as fh:
        raw = {r["acct_id"]: r for r in csv.DictReader(fh)}
    with open(SRC / "nonprofit_990_test_cases_prebinned.csv", newline="") as fh:
        pre = {r["acct_id"]: r for r in csv.DictReader(fh)}
    assert set(raw) == set(pre) and len(raw) == 20, "expected the 20 canonical cases"

    ntee_do = {i: d for i, _l, d, _g, _r in NTEE}
    ntee_gl = {i: g for i, _l, _d, g, _r in NTEE}
    sub_do = {i: d for i, _l, d, _g, _n in SUBSECTION}
    sub_gl = {i: g for i, _l, _d, g, _n in SUBSECTION}
    rev_do = {i: d for i, _l, _lo, _hi, d, _g in REVENUE_BANDS}
    rev_gl = {i: g for i, _l, _lo, _hi, _d, g in REVENUE_BANDS}

    cases = []
    total_round_divergence = 0
    for acct_id, row in sorted(raw.items()):
        p = pre[acct_id]
        revenue = float(row["revenue"])
        expenses = float(row["total_expenses"])
        occupancy = float(row["occupancy_expense"])
        employees = float(row["employee_count"]) if row["employee_count"] != "" else None
        ratio = expenses / revenue
        occ_ratio = occupancy / revenue

        # ratios as they will be WRITTEN to the sheet (8 dp) must bin identically
        ratio_w, occ_w = round(ratio, 8), round(occ_ratio, 8)
        rev_band = band_of(revenue, REVENUE_BANDS)
        pay_band = band_of(employees, PAYROLL_BANDS)
        stress_band = band_of(ratio_w, STRESS_BANDS)
        occ_band = band_of(occ_w, OCCUPANCY_BANDS)
        assert rev_band == p["revenue_band"], f"{acct_id} revenue band {rev_band} != {p['revenue_band']}"
        assert (pay_band or "99_unknown") == p["payroll_band"], f"{acct_id} payroll band"
        assert (stress_band or "99_unknown") == p["stress_band"], f"{acct_id} stress band"
        assert (occ_band or "99_unknown") == p["occupancy_intensity_band"], f"{acct_id} occupancy band"

        tier_code = STATE_TIER[row["state"]]
        st_do, st_gl = STATE_TIERS[tier_code]
        do_unrounded = (
            DO_BASE
            * ntee_do[row["ntee_major"]]
            * rev_do[rev_band]
            * st_do
            * sub_do[row["subsection_type"]]
            * factor_of(pay_band, PAYROLL_BANDS, PAYROLL_DEFAULT)
            * factor_of(stress_band, STRESS_BANDS, STRESS_DEFAULT)
            * LCM
        )
        gl_unrounded = (
            GL_BASE
            * ntee_gl[row["ntee_major"]]
            * rev_gl[rev_band]
            * st_gl
            * sub_gl[row["subsection_type"]]
            * factor_of(occ_band, OCCUPANCY_BANDS, OCCUPANCY_DEFAULT)
            * LCM
        )
        do = round_half_up(do_unrounded)
        gl = round_half_up(gl_unrounded)
        tier = derive_tier(row["ntee_major"], revenue, ratio_w, employees)
        assert do == int(row["expected_do_premium"]), f"{acct_id} D&O {do} != {row['expected_do_premium']}"
        assert gl == int(row["expected_gl_premium"]), f"{acct_id} GL {gl} != {row['expected_gl_premium']}"
        assert do + gl == int(row["expected_total_premium"]), f"{acct_id} total"
        assert tier == row["expected_tier"], f"{acct_id} tier {tier} != {row['expected_tier']}"
        # The platform rounds the PACKAGE total once (registry r2,
        # per_coverage_rounding); the source rounds each coverage first.
        # Track where the two disagree so the test_cases tolerances are
        # honest rather than hoped-for.
        engine_total = round_half_up(do_unrounded + gl_unrounded)
        if engine_total != do + gl:
            total_round_divergence += 1

        cases.append(
            {
                "case_id": acct_id.lower().replace("-", "_"),
                "name": row["name"],
                "ntee_major": row["ntee_major"],
                "revenue": revenue,
                "state": row["state"],
                "subsection_type": row["subsection_type"],
                "employee_count": employees,
                "expense_ratio": ratio_w,
                "occupancy_ratio": occ_w,
                "expected_do_premium": do,
                "expected_gl_premium": gl,
                "expected_total_premium": do + gl,
                "expected_tier": tier.lower(),
            }
        )
    print(
        f"total-rounding divergence (round(sum) vs sum(rounded)): "
        f"{total_round_divergence}/20 cases"
    )
    return cases, total_round_divergence


# ---------------------------------------------------------------------------
# Workbook assembly (spec v1.0 layouts).
# ---------------------------------------------------------------------------

def sheet_rows(ws, rows):
    for r in rows:
        ws.append(list(r))


def add_ft(wb, slug, display, row_dim, method, rows, default=None, note=None):
    """1-D factor table: metadata block, blank row, grid (spec §4.5.1)."""
    ws = wb.create_sheet(f"ft.{slug}")
    meta = [
        ("table_id", slug),
        ("display_name", display),
        ("dimensionality", "1d"),
        ("row_dimension", row_dim),
        ("lookup_method", method),
        ("citation_rule", CITE),
        ("citation_page", note or "source workbook"),
    ]
    sheet_rows(ws, meta)
    ws.append([])
    ws.append(["level_id", "factor", "citation_rule", "citation_page"])
    for level_id, factor in rows:
        ws.append([level_id, factor, CITE, note or "source workbook"])
    if default is not None:
        ws.append(["__default__", default, CITE, "filed unknown-value load"])


def build(cases: list[dict], divergence: int) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "README"
    sheet_rows(
        ws,
        [
            ("Canonical example bundle — filing-transcription-spec v1.0",),
            ("Transcribed from docs/rating-algorithms/nonprofit_990_do_gl_rating_v1.xlsx "
             "(synthetic pre-filing D&O + GL program for IRS-990 nonprofits).",),
            ("Regenerate: see generate_workbook.py next to this file. The generator "
             "re-verifies all 20 test cases (bands, premiums, tiers) before writing.",),
            ("Note: this program has no 2-D factor tables; see the spec §9 mini-example "
             "and the BOP profile for the matrix layout.",),
        ],
    )

    ws = wb.create_sheet("plan")
    sheet_rows(
        ws,
        [
            ("field", "value"),
            ("spec_version", "1.0"),
            ("rating_plan_id", "nonprofit-do-gl-multi-2026"),
            ("display_name", "Nonprofit 990 — D&O + GL — v1"),
            ("version", "1.0.0"),
            ("carrier", "Demo Mutual"),
            ("product", "do"),
            ("jurisdiction_country", "US"),
            ("state", ""),
            ("effective_date", "2026-05-25"),
            ("coverages", "do,gl"),
            ("filing_type", "other"),
            ("source_documents", "nonprofit_990_do_gl_rating_v1.xlsx"),
            ("description",
             "Synthetic pre-filing program (heuristic factors; no regulatory filing). "
             "D&O and GL rated as two towers; eligibility tier is a side verdict."),
        ],
    )

    ws = wb.create_sheet("inputs")
    ws.append(["name", "label", "data_type", "required", "allowed_values", "default_value",
               "unit", "maps_to_dimension", "description", "citation_rule", "citation_page"])
    sheet_rows(
        ws,
        [
            ("ntee_major", "NTEE major group", "string", False, "", "unknown_no_ntee", "",
             "ntee_major", "Missing NTEE maps to the filed unknown level (1.50x both LOBs).",
             CITE, "Plan Spine — edge cases"),
            ("revenue", "Total revenue", "currency", True, "", "", "USD",
             "revenue_band", "Binned into the 7 filed revenue bands.", CITE, "Dim - Revenue Band"),
            ("state", "State", "string", True, "", "", "",
             "state", "USPS 2-letter code; 50 states + DC.", CITE, "Dim - State"),
            ("subsection_type", "IRS 501(c) subsection", "string", False, "", "501c_other", "",
             "subsection_type", "Long-tail subsections fold into 501(c) other.",
             CITE, "Dim - Subsection Type"),
            ("employee_count", "Employee count", "number", False, "", "", "",
             "payroll_band", "Absent -> the filed unknown payroll load (table __default__ 1.10).",
             CITE, "Dim - Payroll Band"),
            ("expense_ratio", "Expense-to-revenue ratio", "number", False, "", "", "",
             "stress_band", "total_expenses / total_revenue, supplied pre-computed "
             "(see gaps_and_assumptions). Absent -> table __default__ 1.10.",
             CITE, "Plan Spine — derived"),
            ("occupancy_ratio", "Occupancy-expense ratio", "number", False, "", "", "",
             "occupancy_intensity", "occupancy_expense / total_revenue, supplied pre-computed. "
             "Absent -> table __default__ 1.15.", CITE, "Plan Spine — derived"),
        ],
    )

    ws = wb.create_sheet("dimensions")
    ws.append(["slug", "display_name", "shape", "role", "data_type", "dimension_type",
               "geo_granularity", "geo_scope", "axes"])
    sheet_rows(
        ws,
        [
            ("ntee_major", "NTEE major group", "categorical", "rating-input", "string", "standard", "", "", ""),
            ("revenue_band", "Revenue band", "banded", "rating-input", "currency", "standard", "", "", ""),
            ("state", "State", "geographic", "rating-input", "string", "geographic", "state", "national", ""),
            ("subsection_type", "IRS subsection", "categorical", "rating-input", "string", "standard", "", "", ""),
            ("payroll_band", "Payroll band", "banded", "rating-input", "number", "standard", "", "", ""),
            ("stress_band", "Financial stress band", "banded", "rating-input", "number", "standard", "", "", ""),
            ("occupancy_intensity", "Occupancy intensity band", "banded", "rating-input", "number", "standard", "", "", ""),
        ],
    )

    ws = wb.create_sheet("dimension_levels")
    ws.append(["dimension_slug", "kind", "level_id", "label", "aliases", "min", "max",
               "territory_ref", "citation_rule", "citation_page"])
    for level_id, label, *_ in NTEE:
        ws.append(["ntee_major", "categorical", level_id, label, "", "", "", "", CITE, "Dim - NTEE Major"])
    for level_id, label, lo, hi, *_ in REVENUE_BANDS:
        ws.append(["revenue_band", "banded", level_id, label, "", lo, hi, "", CITE, "Dim - Revenue Band"])
    for code in sorted(STATE_TIER):
        ws.append(["state", "geographic", code.lower(), STATE_NAMES[code], code, "", "", code,
                   CITE, "Dim - State"])
    for level_id, label, *_ in SUBSECTION:
        ws.append(["subsection_type", "categorical", level_id, label, "", "", "", "",
                   CITE, "Dim - Subsection Type"])
    for level_id, label, lo, hi, _f in PAYROLL_BANDS:
        ws.append(["payroll_band", "banded", level_id, label, "",
                   lo, "+inf" if hi == float("inf") else hi, "", CITE, "Dim - Payroll Band"])
    for level_id, label, lo, hi, _f in STRESS_BANDS:
        ws.append(["stress_band", "banded", level_id, label, "",
                   lo, "+inf" if hi == float("inf") else hi, "", CITE, "Dim - Stress Band"])
    for level_id, label, lo, hi, _f in OCCUPANCY_BANDS:
        ws.append(["occupancy_intensity", "banded", level_id, label, "",
                   lo, "+inf" if hi == float("inf") else hi, "", CITE, "Dim - Occupancy Intensity"])

    add_ft(wb, "ntee_do", "NTEE multiplier (D&O)", "ntee_major", "direct",
           [(i, d) for i, _l, d, _g, _r in NTEE], note="Dim - NTEE Major")
    add_ft(wb, "ntee_gl", "NTEE multiplier (GL)", "ntee_major", "direct",
           [(i, g) for i, _l, _d, g, _r in NTEE], note="Dim - NTEE Major")
    add_ft(wb, "revenue_do", "Revenue multiplier (D&O)", "revenue_band", "binned",
           [(i, d) for i, _l, _lo, _hi, d, _g in REVENUE_BANDS], note="Dim - Revenue Band")
    add_ft(wb, "revenue_gl", "Revenue multiplier (GL)", "revenue_band", "binned",
           [(i, g) for i, _l, _lo, _hi, _d, g in REVENUE_BANDS], note="Dim - Revenue Band")
    add_ft(wb, "state_do", "Territory multiplier (D&O)", "state", "direct",
           [(c.lower(), STATE_TIERS[STATE_TIER[c]][0]) for c in sorted(STATE_TIER)], note="Dim - State")
    add_ft(wb, "state_gl", "Territory multiplier (GL)", "state", "direct",
           [(c.lower(), STATE_TIERS[STATE_TIER[c]][1]) for c in sorted(STATE_TIER)], note="Dim - State")
    add_ft(wb, "subsection_do", "Subsection multiplier (D&O)", "subsection_type", "direct",
           [(i, d) for i, _l, d, _g, _n in SUBSECTION], note="Dim - Subsection Type")
    add_ft(wb, "subsection_gl", "Subsection multiplier (GL)", "subsection_type", "direct",
           [(i, g) for i, _l, _d, g, _n in SUBSECTION], note="Dim - Subsection Type")
    add_ft(wb, "payroll_do", "Payroll modifier (D&O)", "payroll_band", "binned",
           [(i, f) for i, _l, _lo, _hi, f in PAYROLL_BANDS], default=PAYROLL_DEFAULT,
           note="Dim - Payroll Band")
    add_ft(wb, "stress_do", "Financial-stress modifier (D&O)", "stress_band", "binned",
           [(i, f) for i, _l, _lo, _hi, f in STRESS_BANDS], default=STRESS_DEFAULT,
           note="Dim - Stress Band")
    add_ft(wb, "occupancy_gl", "Occupancy modifier (GL)", "occupancy_intensity", "binned",
           [(i, f) for i, _l, _lo, _hi, f in OCCUPANCY_BANDS], default=OCCUPANCY_DEFAULT,
           note="Dim - Occupancy Intensity")

    ws = wb.create_sheet("chains")
    ws.append(["coverage", "order", "stage_kind", "stage_id", "factor_table", "dimension",
               "input_binding", "value", "exposure_divisor", "citation_rule", "citation_page"])
    sheet_rows(
        ws,
        [
            ("do", 0, "base", "do_base", "", "", "literal:600", 600, "", CITE, "Chain - D&O"),
            ("do", 1, "lookup.direct", "do_ntee", "ft.ntee_do", "ntee_major", "", "", "", CITE, "Chain - D&O"),
            ("do", 2, "lookup.range", "do_revenue", "ft.revenue_do", "revenue_band", "", "", "", CITE, "Chain - D&O"),
            ("do", 3, "lookup.direct", "do_state", "ft.state_do", "state", "", "", "", CITE, "Chain - D&O"),
            ("do", 4, "lookup.direct", "do_subsection", "ft.subsection_do", "subsection_type", "", "", "", CITE, "Chain - D&O"),
            ("do", 5, "lookup.range", "do_payroll", "ft.payroll_do", "payroll_band", "", "", "", CITE, "Chain - D&O"),
            ("do", 6, "lookup.range", "do_stress", "ft.stress_do", "stress_band", "", "", "", CITE, "Chain - D&O"),
            ("do", 7, "lcm", "do_lcm", "", "", "", 1.35, "", "(carrier-set)", ""),
            ("gl", 0, "base", "gl_base", "", "", "literal:300", 300, "", CITE, "Chain - GL"),
            ("gl", 1, "lookup.direct", "gl_ntee", "ft.ntee_gl", "ntee_major", "", "", "", CITE, "Chain - GL"),
            ("gl", 2, "lookup.range", "gl_revenue", "ft.revenue_gl", "revenue_band", "", "", "", CITE, "Chain - GL"),
            ("gl", 3, "lookup.direct", "gl_state", "ft.state_gl", "state", "", "", "", CITE, "Chain - GL"),
            ("gl", 4, "lookup.direct", "gl_subsection", "ft.subsection_gl", "subsection_type", "", "", "", CITE, "Chain - GL"),
            ("gl", 5, "lookup.range", "gl_occupancy", "ft.occupancy_gl", "occupancy_intensity", "", "", "", CITE, "Chain - GL"),
            ("gl", 6, "lcm", "gl_lcm", "", "", "", 1.35, "", "(carrier-set)", ""),
        ],
    )

    ws = wb.create_sheet("gates")
    ws.append(["order", "rule_id", "variable", "op", "value",
               "variable_2", "op_2", "value_2", "variable_3", "op_3", "value_3",
               "tier", "reasoning", "citation_rule", "citation_page"])
    sheet_rows(
        ws,
        [
            (1, "decline_distress_micro", "expense_ratio", "gt", 1.15,
             "revenue", "lt", 50000, "", "", "",
             "decline", "Financial distress + micro revenue — auto-decline; offer broker referral.",
             CITE, "Eligibility Tiers"),
            (2, "submit_unknown_ntee", "ntee_major", "in", "unknown_unclassified,unknown_no_ntee",
             "", "", "", "", "", "",
             "submit", "Unknown classification — send to underwriter queue.", CITE, "Eligibility Tiers"),
            (3, "submit_stress", "expense_ratio", "gt", 1.15, "", "", "", "", "", "",
             "submit", "Financial distress — send to underwriter queue.", CITE, "Eligibility Tiers"),
            (4, "submit_large", "revenue", "ge", 5000000, "", "", "", "", "", "",
             "submit", "Above the $5M band ceiling — refer to non-standard rating.", CITE, "README — scope"),
            (5, "preferred_core", "expense_ratio", "lt", 0.85,
             "ntee_major", "in", "philanthropy,science_tech,social_science",
             "revenue", "ge", 250000,
             "preferred", "Healthy surplus + preferred classes + $250K+ revenue — auto-bind; 5% schedule credit eligible.",
             CITE, "Eligibility Tiers"),
            (99, "__default__", "", "", "", "", "", "", "", "", "",
             "standard", "Standard appetite — auto-bind; standard schedule.", CITE, "Eligibility Tiers"),
        ],
    )

    ws = wb.create_sheet("final_adjustments")
    ws.append(["adjustment_id", "kind", "order", "applies_to", "min_value", "max_value",
               "round_increment", "round_min", "citation_rule", "citation_page"])
    sheet_rows(
        ws,
        [
            # ONE package-level round — the engine's round is the plan-tail
            # total-rounder (registry r2 per_coverage_rounding). The source's
            # per-coverage rounding is recorded in gaps_and_assumptions and
            # absorbed by test-case tolerances.
            ("round_total", "round", 1, "", "", "", 1, "", CITE,
             "Chains (round.nearest_dollar; per-coverage in source — see gaps)"),
        ],
    )

    ws = wb.create_sheet("outputs")
    ws.append(["output_id", "field_name", "display_name", "source", "description"])
    sheet_rows(
        ws,
        [
            ("out_do", "do_premium", "D&O premium", "do_lcm",
             "The D&O tower (unrounded; filed value rounds to $1 — tolerance 0.5)."),
            ("out_gl", "gl_premium", "GL premium", "gl_lcm",
             "The GL tower (unrounded; filed value rounds to $1 — tolerance 0.5)."),
            ("out_total", "total_premium", "Total premium", "round_total",
             "Package total, floored + rounded once."),
        ],
    )

    ws = wb.create_sheet("test_cases")
    header = ["case_id", "name", "ntee_major", "revenue", "state", "subsection_type",
              "employee_count", "expense_ratio", "occupancy_ratio",
              "expected_do_premium", "expected_gl_premium", "expected_total_premium",
              "expected_tier", "tolerance_do_premium", "tolerance_gl_premium"]
    if divergence:
        header.append("tolerance_total_premium")
    ws.append(header)
    for c in cases:
        row = [
            c["case_id"], c["name"], c["ntee_major"], c["revenue"], c["state"],
            c["subsection_type"],
            "" if c["employee_count"] is None else c["employee_count"],
            c["expense_ratio"], c["occupancy_ratio"],
            c["expected_do_premium"], c["expected_gl_premium"], c["expected_total_premium"],
            c["expected_tier"], 0.5, 0.5,
        ]
        if divergence:
            row.append(1.0)
        ws.append(row)

    ws = wb.create_sheet("gaps_and_assumptions")
    ws.append(["kind", "description", "citation_rule", "citation_page", "impact", "related"])
    sheet_rows(
        ws,
        [
            ("unsupported",
             "The source derives stress (total_expenses / total_revenue) and occupancy intensity "
             "(occupancy_expense / total_revenue) inside the plan; the workbook format has no "
             "derivation step (capability registry: formula_stage).",
             CITE, "Plan Spine — derived",
             "Datasets must supply expense_ratio and occupancy_ratio pre-computed; test_cases carry them.",
             "inputs!expense_ratio"),
            ("unsupported",
             "The source's Submit trigger includes 'payroll_band = Unknown' — a test for a missing "
             "field, which gates cannot express.",
             CITE, "Eligibility Tiers",
             "Unknown-payroll accounts rate with the filed 1.10 default factor and tier by the "
             "remaining rules.",
             "gates"),
            ("assumption",
             "The subsection table notes '501(c) other — refer if unknown', but the source's "
             "Eligibility Tiers has no such rule. Transcribed as default_value=501c_other with the "
             "filed 1.10/1.05 factors and no submit rule.",
             CITE, "Dim - Subsection Type",
             "Unknown-subsection accounts rate at the 501(c)-other factors and tier normally.",
             "inputs!subsection_type"),
            ("gap",
             "Source discrepancy: the 'Worked Examples' sheet bins Lakeshore Arts occupancy at "
             "02_03_06 (GL 680) while 'Test Cases' bins 01_under_03 (GL 578). The stated binning "
             "rule (50000 / 2400000 = 2.08% < 3%) supports Test Cases — transcribed accordingly.",
             CITE, "Worked Examples vs Test Cases",
             "np_004 expects GL 578 per the source's own Test Cases sheet.",
             "test_cases!np_004"),
            ("unsupported",
             "The source rounds each coverage to $1 before summing; the platform's round is the "
             "plan-tail total-rounder (capability registry r2: per_coverage_rounding).",
             CITE, "Chain - D&O / Chain - GL (round.nearest_dollar)",
             "Per-coverage expected premiums carry tolerance 0.5 (unrounded towers vs filed "
             "rounded values); the package total rounds once at the tail.",
             "final_adjustments!round_total"),
        ],
    )

    wb.save(OUT)
    print(f"verified 20/20 cases (bands + premiums + tiers) — wrote {OUT.relative_to(REPO)}")
    print(f"sheets: {len(wb.sheetnames)} -> {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    build(*load_cases())
